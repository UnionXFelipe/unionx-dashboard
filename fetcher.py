"""
fetcher.py — Actualizador de datos en tiempo real para el dashboard de planificación
═══════════════════════════════════════════════════════════════════════════════════════

Arquitectura de dos capas:
  ESTÁTICA  (semanal):  REPORTE CST/UNID FLAT del Excel maestro → forecast_sku
  DINÁMICA  (10 min):   Google Sheets stock + tránsitos → stock_actual + transitos
  CALCULADA (en cada ciclo): JOIN de ambas → analisis_rt (cobertura real, críticos, sobrestock)

Uso:
  python fetcher.py                     # un ciclo y muestra estado
  python fetcher.py --loop              # loop continuo (Ctrl+C para parar)
  python fetcher.py --import-forecast   # importar datos estáticos del Excel maestro
  python fetcher.py --status            # ver estado de la DB sin correr nada
  python fetcher.py --loop --interval 300   # loop cada 5 min
"""

import sys, os, io, glob, time, logging, datetime, argparse
import sqlite3
import pandas as pd
import numpy as np

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

# ── Rutas ────────────────────────────────────────────────────────────────────
BASE_DIR      = r'C:\Users\felip\Desktop\UnionX Cloude'
EXCEL_DIR     = r'C:\Users\felip\Desktop\UNIONX\FORECAST FINAL SKU'
ANALISIS_DIR  = os.path.join(EXCEL_DIR, 'Analisis Planificacion')
EXCEL_MAESTRO = os.path.join(EXCEL_DIR, 'FORECAST FINAL SKU 26-27 V2.xlsx')
CREDS_FILE    = os.path.join(BASE_DIR, 'credentials.json')
DB_PATH       = os.path.join(BASE_DIR, 'planificacion.db')
LOG_FILE      = os.path.join(BASE_DIR, 'fetcher.log')

# ── Google Sources ────────────────────────────────────────────────────────────
SHEET_STOCK_ID  = '1N5TpIQrFCJwzxyxtueyNi2UoASONd_TbiLcz17a3fag'
SHEET_STOCK_TAB = 'BaseStk'
SHEET_TRANS_ID  = '1RpxZ69Wnfcots006Hp5fzawYxhUsscW03O_hD3psjHA'
DRIVE_VENTAS_ID = '1K11y6icDm9M3X3glGUVCOe4HsbpWpEBm'

SCOPES = [
    'https://spreadsheets.google.com/feeds',
    'https://www.googleapis.com/auth/drive',
]

# ── Parámetros ───────────────────────────────────────────────────────────────
FETCH_INTERVAL  = 600    # segundos entre fetches de stock/tránsitos (10 min)
VENTAS_INTERVAL = 1800   # segundos entre fetches de ventas Drive (30 min)
COB_CRITICO_MAX = 1.0    # < 1 mes → crítico
COB_SOBRE_MIN   = 6.0    # > 6 meses → sobrestock
OPTIMO_MESES    = 4.0    # stock óptimo = 4 meses de venta (para capital exceso)

# ── Mes actual (dinámico) ─────────────────────────────────────────────────────
_TODAY     = datetime.date.today()
_CUR_MONTH = _TODAY.month   # 4=ABR … 10=OCT

# Índice de VentaCst en REPORTE CST FLAT (0-based):
#   ABR(4)=10  |  MAY+(n): (n-5)*6+16
def _cst_venta_idx(m: int) -> int:
    return 10 if m == 4 else (m - 5) * 6 + 16

# Índice de VentaPPTO en REPORTE UNID FLAT (0-based):
#   ABR(4)=11  |  MAY+(n): (n-5)*6+17
def _unid_venta_idx(m: int) -> int:
    return 11 if m == 4 else (m - 5) * 6 + 17

# ── Logging ──────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s  %(levelname)-8s  %(message)s',
    handlers=[
        logging.FileHandler(LOG_FILE, encoding='utf-8', mode='a'),
        logging.StreamHandler(),
    ]
)
log = logging.getLogger('fetcher')


# ════════════════════════════════════════════════════════════════════════════
# HELPERS
# ════════════════════════════════════════════════════════════════════════════
def _fmt_sku(v) -> str:
    """Normaliza cualquier valor de SKU a string limpio (sin '.0')."""
    if v is None:
        return ''
    if isinstance(v, float):
        return str(int(v)) if not np.isnan(v) else ''
    s = str(v).strip()
    if s.endswith('.0') and s[:-2].lstrip('-').isdigit():
        return s[:-2]
    return s


def _flt(v, default: float = 0.0) -> float:
    """Convierte a float tolerando comas y espacios."""
    if v is None or v == '':
        return default
    try:
        return float(str(v).replace(',', '.').replace(' ', '').replace('\xa0', ''))
    except Exception:
        return default


def _int(v, default: int = 0) -> int:
    try:
        return int(float(str(v).replace(',', '')))
    except Exception:
        return default


def _now() -> str:
    return datetime.datetime.now().isoformat(timespec='seconds')


def _find_col(headers: list, candidates: list):
    """Busca índice de columna por nombre (case-insensitive, substring)."""
    for name in candidates:
        for i, h in enumerate(headers):
            if name.lower() in str(h).lower():
                return i
    return None


# ════════════════════════════════════════════════════════════════════════════
# SCHEMA SQLite
# ════════════════════════════════════════════════════════════════════════════
SCHEMA = """
-- ── Estático: forecast ventas por SKU (del Excel maestro, actualizado semanalmente) ──
CREATE TABLE IF NOT EXISTS forecast_sku (
    sku            TEXT PRIMARY KEY,
    marca          TEXT,
    cat_padre      TEXT,
    cat_hijo       TEXT,
    cat_com        TEXT,
    descripcion    TEXT,
    ranking        INTEGER,
    puerto_origen  TEXT,
    -- Venta PPTO en unidades (para denominar cobertura) — 3 meses desde el actual
    venta_unid_m0  REAL,   -- mes actual
    venta_unid_m1  REAL,   -- mes+1
    venta_unid_m2  REAL,   -- mes+2
    -- Venta CST (para valor monetario de stock)
    venta_cst_m0   REAL,
    venta_cst_m1   REAL,
    venta_cst_m2   REAL,
    -- Valores base del último análisis semanal (fallback si no hay dato de Sheets)
    stock_cst_base  REAL,
    stock_unid_base REAL,
    cobert_base     REAL,
    imported_at    TEXT
);

-- ── Dinámico: stock actual por SKU (de Google Sheets, actualizado cada 10 min) ──
CREATE TABLE IF NOT EXISTS stock_actual (
    sku         TEXT PRIMARY KEY,
    stock_unid  REAL,
    stock_cst   REAL,
    fetched_at  TEXT
);

-- ── Dinámico: tránsitos abiertos (de Google Sheets, actualizado cada 10 min) ──
CREATE TABLE IF NOT EXISTS transitos (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    sku         TEXT NOT NULL,
    descripcion TEXT,
    pi          TEXT,
    cantidad    REAL,
    eta_bodega  TEXT,
    mes         INTEGER,
    marca       TEXT,
    valor_usd   REAL,
    tipo_cat    TEXT,
    fetched_at  TEXT
);

-- ── Dinámico: ventas acumuladas por marca/mes (de Google Drive, cada 30 min) ──
CREATE TABLE IF NOT EXISTS vta_acum_marca (
    marca      TEXT,
    mes        INTEGER,
    ano        INTEGER,
    venta_cst  REAL,
    fetched_at TEXT,
    PRIMARY KEY (marca, mes, ano)
);

-- ── Calculado: resultado en tiempo real (regenerado en cada ciclo) ──
CREATE TABLE IF NOT EXISTS analisis_rt (
    sku            TEXT PRIMARY KEY,
    marca          TEXT,
    cat_padre      TEXT,
    cat_hijo       TEXT,
    cat_com        TEXT,
    descripcion    TEXT,
    ranking        INTEGER,
    -- Stock real (Google Sheets, o base si no hay dato fresco)
    stock_unid     REAL,
    stock_cst      REAL,
    -- Tránsitos reales
    qty_transito   REAL,
    prox_pi        TEXT,
    prox_eta       TEXT,
    prox_mes       INTEGER,
    -- Cobertura calculada con stock real
    avg_venta_3m   REAL,    -- promedio venta PPTO unidades (3 meses)
    cobert_actual  REAL,    -- stock_unid / avg_venta_3m  (sin tránsitos)
    cobert_proyect REAL,    -- (stock_unid + qty_transito) / avg_venta_3m  (con tránsitos)
    -- Flags
    es_critico     INTEGER, -- 1 si cobert_actual < 1
    es_sobrestock  INTEGER, -- 1 si cobert_actual > 6
    capital_exceso REAL,    -- stock_cst - (venta_cst * 4)  (solo si sobrestock)
    venta_cst_m0   REAL,
    computed_at    TEXT
);

-- ── Log de fetches ──
CREATE TABLE IF NOT EXISTS fetch_log (
    id      INTEGER PRIMARY KEY AUTOINCREMENT,
    source  TEXT,
    status  TEXT,
    rows    INTEGER,
    error   TEXT,
    ts      TEXT DEFAULT (datetime('now','localtime'))
);

-- ── Metadatos (timestamps, versión, etc.) ──
CREATE TABLE IF NOT EXISTS meta (
    key   TEXT PRIMARY KEY,
    value TEXT
);

CREATE INDEX IF NOT EXISTS idx_trans_sku    ON transitos(sku);
CREATE INDEX IF NOT EXISTS idx_art_marca    ON analisis_rt(marca);
CREATE INDEX IF NOT EXISTS idx_art_critico  ON analisis_rt(es_critico);
CREATE INDEX IF NOT EXISTS idx_art_sobre    ON analisis_rt(es_sobrestock);
"""


def init_db(db_path: str = DB_PATH):
    conn = sqlite3.connect(db_path)
    conn.executescript(SCHEMA)
    conn.commit()
    conn.close()
    log.info(f"DB lista: {db_path}")


def _conn(db_path: str = DB_PATH) -> sqlite3.Connection:
    c = sqlite3.connect(db_path)
    c.row_factory = sqlite3.Row
    return c


def _set_meta(conn, key: str, value: str):
    conn.execute("INSERT OR REPLACE INTO meta VALUES (?,?)", (key, value))


def _get_meta(conn, key: str, default=None):
    r = conn.execute("SELECT value FROM meta WHERE key=?", (key,)).fetchone()
    return r[0] if r else default


def _log_fetch(conn, source: str, status: str, rows: int = 0, error: str = None):
    conn.execute(
        "INSERT INTO fetch_log (source, status, rows, error) VALUES (?,?,?,?)",
        (source, status, rows, error)
    )


# ════════════════════════════════════════════════════════════════════════════
# IMPORTACIÓN ESTÁTICA — REPORTE CST/UNID FLAT (openpyxl, sin xlwings)
# ════════════════════════════════════════════════════════════════════════════
def import_forecast(excel_path: str = EXCEL_MAESTRO, db_path: str = DB_PATH) -> bool:
    """
    Lee REPORTE CST FLAT + REPORTE UNID FLAT + FCST BASE del Excel maestro
    usando openpyxl read-only (no requiere Excel.exe corriendo).

    Llamar UNA VEZ después de la actualización semanal. Los datos de forecast
    se mantienen estáticos hasta la próxima llamada.

    Returns True si fue exitoso.
    """
    import openpyxl

    log.info("── Importando forecast estático desde Excel maestro ──")
    if not os.path.exists(excel_path):
        log.error(f"Excel maestro no encontrado: {excel_path}")
        return False

    try:
        # data_only=True lee valores calculados (no fórmulas)
        log.info("  Abriendo Excel con openpyxl (puede tomar 30-90 seg)...")
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

        # ── REPORTE CST FLAT ──────────────────────────────────────────────
        if 'REPORTE CST FLAT' not in wb.sheetnames:
            log.error("Hoja 'REPORTE CST FLAT' no encontrada en el Excel")
            wb.close(); return False

        log.info("  Leyendo REPORTE CST FLAT...")
        data_cst = [row for row in wb['REPORTE CST FLAT'].values]
        data_cst = data_cst[1:]  # saltar encabezado

        # ── REPORTE UNID FLAT ─────────────────────────────────────────────
        if 'REPORTE UNID FLAT' not in wb.sheetnames:
            log.error("Hoja 'REPORTE UNID FLAT' no encontrada en el Excel")
            wb.close(); return False

        log.info("  Leyendo REPORTE UNID FLAT...")
        data_unid = [row for row in wb['REPORTE UNID FLAT'].values]
        data_unid = data_unid[1:]

        # ── FCST BASE SKU MACRO — sólo cols ranking (J=9) y puerto (S=18) ─
        fcst_extra: dict = {}
        if 'FCST BASE SKU MACRO' in wb.sheetnames:
            log.info("  Leyendo FCST BASE SKU MACRO (ranking + puerto)...")
            for row in wb['FCST BASE SKU MACRO'].iter_rows(min_row=4, values_only=True):
                if not row or len(row) < 5 or row[4] is None:
                    continue
                sku = _fmt_sku(row[4])
                if sku:
                    fcst_extra[sku] = {
                        'ranking': _int(row[9])        if len(row) > 9  else 0,
                        'puerto':  str(row[18]).strip() if len(row) > 18 and row[18] else '',
                    }

        wb.close()

        # ── Índices de ventas según mes actual ────────────────────────────
        m0, m1, m2 = _CUR_MONTH, _CUR_MONTH + 1, _CUR_MONTH + 2
        cst_v0, cst_v1, cst_v2 = (
            _cst_venta_idx(m0),
            _cst_venta_idx(m1) if m1 <= 10 else None,
            _cst_venta_idx(m2) if m2 <= 10 else None,
        )
        unid_v0, unid_v1, unid_v2 = (
            _unid_venta_idx(m0),
            _unid_venta_idx(m1) if m1 <= 10 else None,
            _unid_venta_idx(m2) if m2 <= 10 else None,
        )
        log.info(f"  Mes actual: {m0} | Índices VentaCST: {cst_v0},{cst_v1},{cst_v2} "
                 f"| VentaUnid: {unid_v0},{unid_v1},{unid_v2}")

        # ── Construir lookup UNID por SKU ─────────────────────────────────
        unid_por_sku: dict = {}
        for ru in data_unid:
            if not ru or len(ru) < 5 or ru[4] is None:
                continue
            sk = _fmt_sku(ru[4])
            if sk:
                unid_por_sku[sk] = ru

        # ── Combinar CST + UNID ───────────────────────────────────────────
        def _safe(row, idx, default=0.0):
            try:
                v = row[idx]
                return _flt(v, default) if v is not None else default
            except (IndexError, TypeError):
                return default

        EXCLUIR = {'(sin clasificar)', ''}
        records = []

        for rc in data_cst:
            if not rc or rc[0] is None:
                continue
            marca = str(rc[0]).strip()
            if not marca or marca.lower().startswith('total '):
                continue
            cat_com = str(rc[3]).strip() if rc[3] else ''
            if cat_com in EXCLUIR or 'descontinua' in cat_com.lower():
                continue

            sku = _fmt_sku(rc[4])
            if not sku:
                continue

            ru    = unid_por_sku.get(sku)
            extra = fcst_extra.get(sku, {})

            records.append({
                'sku':           sku,
                'marca':         marca,
                'cat_padre':     str(rc[1]).strip() if rc[1] else '',
                'cat_hijo':      str(rc[2]).strip() if rc[2] else '',
                'cat_com':       cat_com,
                'descripcion':   str(rc[5]).strip() if rc[5] else '',
                'ranking':       extra.get('ranking', 0),
                'puerto_origen': extra.get('puerto', ''),
                # Venta PPTO unidades (para cobertura)
                'venta_unid_m0': _safe(ru, unid_v0)            if ru else 0.0,
                'venta_unid_m1': _safe(ru, unid_v1)            if (ru and unid_v1) else 0.0,
                'venta_unid_m2': _safe(ru, unid_v2)            if (ru and unid_v2) else 0.0,
                # Venta CST (valor monetario)
                'venta_cst_m0':  _safe(rc, cst_v0),
                'venta_cst_m1':  _safe(rc, cst_v1)             if cst_v1 else 0.0,
                'venta_cst_m2':  _safe(rc, cst_v2)             if cst_v2 else 0.0,
                # Stock base (fallback si Google Sheets no responde)
                'stock_cst_base':   _safe(rc, 6),
                'stock_unid_base':  _safe(ru, 6)               if ru else 0.0,
                'cobert_base':      _safe(rc, 7),
                'imported_at':  _now(),
            })

        # ── Guardar ───────────────────────────────────────────────────────
        conn = _conn(db_path)
        conn.execute("DELETE FROM forecast_sku")
        conn.executemany("""
            INSERT OR REPLACE INTO forecast_sku VALUES (
                :sku, :marca, :cat_padre, :cat_hijo, :cat_com, :descripcion,
                :ranking, :puerto_origen,
                :venta_unid_m0, :venta_unid_m1, :venta_unid_m2,
                :venta_cst_m0,  :venta_cst_m1,  :venta_cst_m2,
                :stock_cst_base, :stock_unid_base, :cobert_base, :imported_at
            )
        """, records)
        _set_meta(conn, 'forecast_imported_at', _now())
        _set_meta(conn, 'forecast_mes', str(_CUR_MONTH))
        _log_fetch(conn, 'forecast_import', 'ok', len(records))
        conn.commit(); conn.close()

        log.info(f"  ✓ Forecast importado: {len(records)} SKUs activos")
        return True

    except Exception as e:
        log.exception(f"Error importando forecast: {e}")
        return False


# ════════════════════════════════════════════════════════════════════════════
# FETCH DINÁMICO — STOCK (Google Sheets BaseStk)
# ════════════════════════════════════════════════════════════════════════════
def fetch_stock(db_path: str = DB_PATH) -> int:
    """
    Lee stock actual de Google Sheets 'Matriz stock' / BaseStk.
    Guarda en stock_actual (reemplaza completamente la tabla).
    Retorna cantidad de SKUs guardados.
    """
    import gspread
    from google.oauth2.service_account import Credentials

    log.info("Fetching stock de Google Sheets BaseStk...")
    try:
        creds = Credentials.from_service_account_file(CREDS_FILE, scopes=SCOPES)
        gc    = gspread.authorize(creds)
        ws    = gc.open_by_key(SHEET_STOCK_ID).worksheet(SHEET_STOCK_TAB)
        data  = ws.get_all_values()

        if not data or len(data) < 2:
            log.warning("BaseStk vacío o sin datos")
            return 0

        headers = [str(h).strip() for h in data[0]]
        log.info(f"  Headers BaseStk (primeros 10): {headers[:10]}")

        # Descubrir columnas — el nombre exacto varía, buscamos por substring
        i_sku   = _find_col(headers, ['sku', 'SKU'])
        i_sunid = _find_col(headers, [
            'stock hoy unid', 'stock unid', 'stk hoy unid', 'stk unid',
            'unidades', 'stock actual unid',
        ])
        i_scst  = _find_col(headers, [
            'stock hoy cst', 'stock cst', 'stk hoy cst', 'stk cst',
            'stock hoy cost', 'costo stock',
        ])

        if i_sku is None:
            log.error(f"No se encontró columna SKU en BaseStk. Headers: {headers}")
            return 0

        log.info(f"  Columnas → SKU:{i_sku} | Stock Unid:{i_sunid} | Stock CST:{i_scst}")

        now = _now()
        records = []
        for row in data[1:]:
            if not row or len(row) <= i_sku or not row[i_sku]:
                continue
            sku = _fmt_sku(row[i_sku])
            if not sku:
                continue
            records.append({
                'sku':        sku,
                'stock_unid': _flt(row[i_sunid]) if i_sunid is not None and i_sunid < len(row) else 0.0,
                'stock_cst':  _flt(row[i_scst])  if i_scst  is not None and i_scst  < len(row) else 0.0,
                'fetched_at': now,
            })

        conn = _conn(db_path)
        conn.execute("DELETE FROM stock_actual")
        conn.executemany(
            "INSERT OR REPLACE INTO stock_actual VALUES (:sku, :stock_unid, :stock_cst, :fetched_at)",
            records
        )
        _set_meta(conn, 'stock_fetched_at', now)
        _log_fetch(conn, 'stock', 'ok', len(records))
        conn.commit(); conn.close()

        log.info(f"  ✓ Stock: {len(records)} SKUs guardados")
        return len(records)

    except Exception as e:
        log.exception(f"Error fetching stock: {e}")
        conn = _conn(db_path)
        _log_fetch(conn, 'stock', 'error', 0, str(e)[:500])
        conn.commit(); conn.close()
        return 0


# ════════════════════════════════════════════════════════════════════════════
# FETCH DINÁMICO — TRÁNSITOS (Google Sheets Importaciones UnionX)
# ════════════════════════════════════════════════════════════════════════════
def fetch_transitos(db_path: str = DB_PATH) -> int:
    """
    Lee tránsitos de Google Sheets 'Importaciones UnionX'.
    Los datos empiezan en fila 14 (índice 13). Encabezados en fila 1.
    Retorna cantidad de filas guardadas.
    """
    import gspread
    from google.oauth2.service_account import Credentials

    log.info("Fetching tránsitos de Google Sheets Importaciones...")
    try:
        creds = Credentials.from_service_account_file(CREDS_FILE, scopes=SCOPES)
        gc    = gspread.authorize(creds)
        ws    = gc.open_by_key(SHEET_TRANS_ID).get_worksheet(0)
        data  = ws.get_all_values()

        if not data or len(data) < 15:
            log.warning(f"Sheet tránsitos insuficiente ({len(data)} filas)")
            return 0

        headers = [str(h).strip() for h in data[0]]

        # Mapeo dinámico de columnas (igual que analisis_stock_critico.py)
        i_sku   = _find_col(headers, ['sku', 'SKU'])                          or 0
        i_desc  = _find_col(headers, ['descripcion', 'descripción'])           or 1
        i_pi    = _find_col(headers, ['pi', 'número pi', 'num pi'])            or 2
        i_qty   = _find_col(headers, ['cantidad'])                             or 6
        i_eta   = _find_col(headers, ['fecha eta bodega', 'eta bodega'])       or 12
        i_mes   = _find_col(headers, ['mes', 'MES'])                           or 13
        i_marca = _find_col(headers, ['marca', 'MARCA'])                       or 17
        i_usd   = _find_col(headers, ['valor usd total', 'valor usd'])         or 16
        i_tipo  = _find_col(headers, ['tipo categoria', 'tipo categoría'])     or 15

        log.info(f"  SKU:{i_sku} PI:{i_pi} Qty:{i_qty} ETA:{i_eta} Mes:{i_mes} Marca:{i_marca}")

        now = _now()
        records = []
        for row in data[13:]:   # fila 14 en adelante (índice 13)
            if not row:
                continue
            def _g(i):
                return row[i] if i < len(row) else ''

            sku = _fmt_sku(_g(i_sku))
            if not sku:
                continue
            pi  = str(_g(i_pi)).strip()
            eta = str(_g(i_eta)).strip()
            if not pi and not eta:
                continue    # fila sin tránsito real

            records.append({
                'sku':        sku,
                'descripcion':str(_g(i_desc)).strip(),
                'pi':         pi,
                'cantidad':   _flt(_g(i_qty)),
                'eta_bodega': eta,
                'mes':        _int(_g(i_mes)),
                'marca':      str(_g(i_marca)).strip(),
                'valor_usd':  _flt(_g(i_usd)),
                'tipo_cat':   str(_g(i_tipo)).strip(),
                'fetched_at': now,
            })

        conn = _conn(db_path)
        conn.execute("DELETE FROM transitos")
        conn.executemany("""
            INSERT INTO transitos
              (sku, descripcion, pi, cantidad, eta_bodega, mes, marca, valor_usd, tipo_cat, fetched_at)
            VALUES
              (:sku, :descripcion, :pi, :cantidad, :eta_bodega, :mes, :marca, :valor_usd, :tipo_cat, :fetched_at)
        """, records)
        _set_meta(conn, 'transitos_fetched_at', now)
        _log_fetch(conn, 'transitos', 'ok', len(records))
        conn.commit(); conn.close()

        log.info(f"  ✓ Tránsitos: {len(records)} filas guardadas")
        return len(records)

    except Exception as e:
        log.exception(f"Error fetching tránsitos: {e}")
        conn = _conn(db_path)
        _log_fetch(conn, 'transitos', 'error', 0, str(e)[:500])
        conn.commit(); conn.close()
        return 0


# ════════════════════════════════════════════════════════════════════════════
# FETCH DINÁMICO — VENTAS ACUM (Google Drive Raw ventas)
# Más pesado (~340k filas) — correr cada 30 min, no cada 10
# ════════════════════════════════════════════════════════════════════════════
def fetch_ventas_acum(db_path: str = DB_PATH) -> int:
    """
    Descarga 'Raw ventas Y.xlsx' de Google Drive, agrega por Marca × Mes,
    guarda en vta_acum_marca. Sólo el año actual.
    """
    from google.oauth2.service_account import Credentials
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaIoBaseDownload

    log.info("Fetching ventas acumuladas de Google Drive...")
    try:
        creds   = Credentials.from_service_account_file(CREDS_FILE, scopes=SCOPES)
        service = build('drive', 'v3', credentials=creds)

        log.info("  Descargando archivo (puede tardar 20-60 seg)...")
        request = service.files().get_media(fileId=DRIVE_VENTAS_ID)
        buf = io.BytesIO()
        dl  = MediaIoBaseDownload(buf, request)
        done = False
        while not done:
            _, done = dl.next_chunk()
        buf.seek(0)

        log.info("  Procesando hoja RAW...")
        df = pd.read_excel(buf, sheet_name='RAW', engine='openpyxl')
        cols = [str(c).strip() for c in df.columns]

        # Descubrir columnas clave
        marca_col = next((c for c in cols if 'marca' in c.lower()), None)
        fecha_col = (next((c for c in cols if 'fecha' in c.lower() and 'venta' in c.lower()), None)
                     or next((c for c in cols if 'fecha' in c.lower()), None))
        monto_col = (next((c for c in cols if 'costo' in c.lower() and 'total' in c.lower()), None)
                     or next((c for c in cols if 'neto' in c.lower() and ('cst' in c.lower() or 'costo' in c.lower())), None)
                     or next((c for c in cols if 'cst' in c.lower()), None))

        if not marca_col or not fecha_col or not monto_col:
            log.warning(f"  Columnas RAW (primeras 20): {cols[:20]}")
            log.error("  No se encontraron columnas marca/fecha/costo")
            return 0

        log.info(f"  Columnas: Marca={marca_col} Fecha={fecha_col} Monto={monto_col}")

        df[fecha_col] = pd.to_datetime(df[fecha_col], errors='coerce')
        df = df.dropna(subset=[fecha_col, marca_col])
        df = df[df[fecha_col].dt.year == _TODAY.year]   # sólo año actual
        df['mes'] = df[fecha_col].dt.month
        df['ano'] = df[fecha_col].dt.year
        df[monto_col] = pd.to_numeric(df[monto_col], errors='coerce').fillna(0)

        agr = (df.groupby([marca_col, 'mes', 'ano'], as_index=False)[monto_col]
                 .sum()
                 .rename(columns={marca_col: 'marca', monto_col: 'venta_cst'}))

        now = _now()
        records = agr.to_dict('records')
        for r in records:
            r['fetched_at'] = now

        conn = _conn(db_path)
        conn.execute("DELETE FROM vta_acum_marca WHERE ano=?", (_TODAY.year,))
        conn.executemany(
            "INSERT OR REPLACE INTO vta_acum_marca VALUES (:marca, :mes, :ano, :venta_cst, :fetched_at)",
            records
        )
        _set_meta(conn, 'ventas_fetched_at', now)
        _log_fetch(conn, 'ventas_acum', 'ok', len(records))
        conn.commit(); conn.close()

        log.info(f"  ✓ Ventas acum: {len(records)} combinaciones marca×mes")
        return len(records)

    except Exception as e:
        log.exception(f"Error fetching ventas: {e}")
        conn = _conn(db_path)
        _log_fetch(conn, 'ventas_acum', 'error', 0, str(e)[:500])
        conn.commit(); conn.close()
        return 0


# ════════════════════════════════════════════════════════════════════════════
# RECÁLCULO EN TIEMPO REAL
# ════════════════════════════════════════════════════════════════════════════
def recalculate(db_path: str = DB_PATH) -> int:
    """
    JOIN forecast_sku (estático) + stock_actual (dinámico) + transitos (dinámico)
    → analisis_rt (tabla calculada, reemplazada en cada ciclo).

    Cobertura calculada:
      cobert_actual  = stock_unid_real / avg_venta_ppto(m0, m1, m2)
      cobert_proyect = (stock_unid_real + qty_transito) / avg_venta_ppto(...)
    """
    log.info("Recalculando análisis RT...")
    conn = _conn(db_path)

    n_fc = conn.execute("SELECT COUNT(*) FROM forecast_sku").fetchone()[0]
    n_st = conn.execute("SELECT COUNT(*) FROM stock_actual").fetchone()[0]

    if n_fc == 0:
        log.warning("forecast_sku vacío — ejecuta: python fetcher.py --import-forecast")
        conn.close(); return 0
    if n_st == 0:
        log.warning("stock_actual vacío — ejecuta primero fetch_stock()")
        conn.close(); return 0

    # Agregar tránsitos por SKU
    trans_df = pd.read_sql("""
        SELECT sku,
               SUM(cantidad)   AS qty_total,
               MIN(eta_bodega) AS prox_eta,
               MIN(mes)        AS prox_mes
        FROM   transitos
        WHERE  cantidad > 0
        GROUP  BY sku
    """, conn)
    trans_idx = trans_df.set_index('sku').to_dict('index') if len(trans_df) else {}

    # PI más próximo por SKU
    pi_df = pd.read_sql("""
        SELECT sku, pi
        FROM   transitos
        WHERE  cantidad > 0 AND pi IS NOT NULL AND pi != ''
        GROUP  BY sku
        HAVING MIN(COALESCE(eta_bodega,'9999'))
    """, conn)
    pi_idx = dict(zip(pi_df['sku'], pi_df['pi'])) if len(pi_df) else {}

    # Merge forecast + stock
    fc = pd.read_sql("SELECT * FROM forecast_sku", conn)
    st = pd.read_sql("SELECT * FROM stock_actual", conn)
    df = fc.merge(st, on='sku', how='left')

    now = _now()
    records = []

    for _, r in df.iterrows():
        sku = r['sku']

        # Stock real o fallback al base del Excel
        s_unid = r['stock_unid'] if pd.notna(r.get('stock_unid')) else r['stock_unid_base']
        s_cst  = r['stock_cst']  if pd.notna(r.get('stock_cst'))  else r['stock_cst_base']

        if s_unid is None or pd.isna(s_unid):
            s_unid = 0.0
        if s_cst is None or pd.isna(s_cst):
            s_cst = 0.0

        # Tránsitos reales
        tr         = trans_idx.get(sku, {})
        qty_trans  = float(tr.get('qty_total',  0) or 0)
        prox_eta   = str(tr.get('prox_eta', '') or '')
        prox_mes   = tr.get('prox_mes', None)
        prox_pi    = pi_idx.get(sku, '')

        # Promedio venta 3 meses (PPTO unidades)
        v0 = float(r['venta_unid_m0'] or 0)
        v1 = float(r['venta_unid_m1'] or 0)
        v2 = float(r['venta_unid_m2'] or 0)
        ventas_pos = [v for v in (v0, v1, v2) if v > 0]
        avg_v = sum(ventas_pos) / len(ventas_pos) if ventas_pos else 0.0

        # Coberturas
        cobert_act  = (s_unid / avg_v)              if avg_v > 0 else None
        cobert_proy = ((s_unid + qty_trans) / avg_v) if avg_v > 0 else cobert_act

        # Flags
        es_critico    = 1 if cobert_act is not None and cobert_act < COB_CRITICO_MAX  else 0
        es_sobrestock = 1 if cobert_act is not None and cobert_act > COB_SOBRE_MIN    else 0

        # Capital inmovilizado excedente
        venta_cst  = float(r['venta_cst_m0'] or 0)
        stock_opt  = venta_cst * OPTIMO_MESES
        cap_exceso = max(0.0, s_cst - stock_opt) if es_sobrestock else 0.0

        records.append({
            'sku':           sku,
            'marca':         r['marca'],
            'cat_padre':     r['cat_padre'],
            'cat_hijo':      r['cat_hijo'],
            'cat_com':       r['cat_com'],
            'descripcion':   r['descripcion'],
            'ranking':       int(r['ranking'] or 0),
            'stock_unid':    float(s_unid),
            'stock_cst':     float(s_cst),
            'qty_transito':  qty_trans,
            'prox_pi':       prox_pi,
            'prox_eta':      prox_eta,
            'prox_mes':      int(prox_mes) if prox_mes else None,
            'avg_venta_3m':  avg_v,
            'cobert_actual': cobert_act,
            'cobert_proyect':cobert_proy,
            'es_critico':    es_critico,
            'es_sobrestock': es_sobrestock,
            'capital_exceso':cap_exceso,
            'venta_cst_m0':  venta_cst,
            'computed_at':   now,
        })

    conn.execute("DELETE FROM analisis_rt")
    conn.executemany("""
        INSERT OR REPLACE INTO analisis_rt VALUES (
            :sku, :marca, :cat_padre, :cat_hijo, :cat_com, :descripcion, :ranking,
            :stock_unid, :stock_cst, :qty_transito, :prox_pi, :prox_eta, :prox_mes,
            :avg_venta_3m, :cobert_actual, :cobert_proyect,
            :es_critico, :es_sobrestock, :capital_exceso, :venta_cst_m0, :computed_at
        )
    """, records)
    _set_meta(conn, 'last_computed_at', now)
    conn.commit()

    n_crit = sum(1 for r in records if r['es_critico'])
    n_sob  = sum(1 for r in records if r['es_sobrestock'])
    log.info(f"  ✓ analisis_rt: {len(records)} SKUs | Críticos: {n_crit} | Sobrestock: {n_sob}")

    conn.close()
    return len(records)


# ════════════════════════════════════════════════════════════════════════════
# LOOP PRINCIPAL
# ════════════════════════════════════════════════════════════════════════════
def run_once(db_path: str = DB_PATH, fetch_ventas: bool = False):
    """Un ciclo completo: fetch stock + tránsitos → recalcular."""
    t0 = time.time()
    fetch_stock(db_path)
    fetch_transitos(db_path)
    if fetch_ventas:
        fetch_ventas_acum(db_path)
    recalculate(db_path)
    log.info(f"Ciclo completado en {time.time() - t0:.1f}s")


def run_loop(interval: int = FETCH_INTERVAL, db_path: str = DB_PATH):
    """
    Loop continuo. Stock + tránsitos cada `interval` seg.
    Ventas Drive cada VENTAS_INTERVAL seg.
    Ctrl+C para parar.
    """
    log.info(f"═══ Iniciando loop — intervalo {interval}s ({interval//60} min) ═══")
    ventas_every   = max(1, VENTAS_INTERVAL // interval)
    ventas_counter = 0

    while True:
        ventas_counter += 1
        try:
            run_once(db_path, fetch_ventas=(ventas_counter % ventas_every == 0))
        except KeyboardInterrupt:
            log.info("Detenido por usuario")
            break
        except Exception as e:
            log.exception(f"Error en ciclo: {e}")
        log.info(f"Próximo fetch en {interval}s  ({datetime.datetime.now() + datetime.timedelta(seconds=interval):%H:%M:%S})")
        time.sleep(interval)


# ════════════════════════════════════════════════════════════════════════════
# STATUS
# ════════════════════════════════════════════════════════════════════════════
def status(db_path: str = DB_PATH):
    """Muestra el estado actual de la base de datos."""
    if not os.path.exists(db_path):
        print("⚠  Base de datos no existe. Ejecuta: python fetcher.py --init-db")
        return
    conn = _conn(db_path)
    print("\n═══ Estado planificacion.db ═══")
    for tbl, lbl in [
        ('forecast_sku',   'Forecast SKUs  (Excel maestro, estático)'),
        ('stock_actual',   'Stock actual   (Google Sheets, dinámico) '),
        ('transitos',      'Tránsitos      (Google Sheets, dinámico) '),
        ('analisis_rt',    'Análisis RT    (calculado)               '),
        ('vta_acum_marca', 'Ventas acum    (Drive, dinámico)         '),
    ]:
        n = conn.execute(f"SELECT COUNT(*) FROM {tbl}").fetchone()[0]
        print(f"  {lbl:<48} {n:>6} filas")

    print("\n─── Últimas actualizaciones ───")
    for key in ['forecast_imported_at', 'stock_fetched_at', 'transitos_fetched_at',
                'ventas_fetched_at', 'last_computed_at']:
        v = _get_meta(conn, key, '—')
        print(f"  {key:<35} {v}")

    if conn.execute("SELECT COUNT(*) FROM analisis_rt").fetchone()[0] > 0:
        n_crit = conn.execute("SELECT COUNT(*) FROM analisis_rt WHERE es_critico=1").fetchone()[0]
        n_sob  = conn.execute("SELECT COUNT(*) FROM analisis_rt WHERE es_sobrestock=1").fetchone()[0]
        n_sin  = conn.execute("SELECT COUNT(*) FROM analisis_rt WHERE cobert_actual IS NULL").fetchone()[0]
        print(f"\n─── Resultados RT ───")
        print(f"  Críticos (cob < 1m):      {n_crit}")
        print(f"  Sobrestock (cob > 6m):    {n_sob}")
        print(f"  Sin venta PPTO (cobert=?): {n_sin}")
    conn.close()


# ════════════════════════════════════════════════════════════════════════════
# ENTRY POINT
# ════════════════════════════════════════════════════════════════════════════
if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='UnionX — Fetcher de datos en tiempo real')
    parser.add_argument('--init-db',         action='store_true', help='Inicializar DB y salir')
    parser.add_argument('--import-forecast', action='store_true', help='Importar forecast del Excel maestro (una vez, ~1-2 min)')
    parser.add_argument('--loop',            action='store_true', help='Correr en loop continuo (Ctrl+C para parar)')
    parser.add_argument('--status',          action='store_true', help='Ver estado de la DB')
    parser.add_argument('--with-ventas',     action='store_true', help='Incluir fetch de ventas Drive')
    parser.add_argument('--interval',        type=int, default=FETCH_INTERVAL, help=f'Segundos entre fetches (default: {FETCH_INTERVAL})')
    args = parser.parse_args()

    init_db()   # siempre asegurar que la DB exista

    if args.status:
        status()
    elif args.init_db:
        print("DB inicializada.")
    elif args.import_forecast:
        ok = import_forecast()
        if ok:
            status()
    elif args.loop:
        run_loop(interval=args.interval)
    else:
        # Single run por defecto
        run_once(fetch_ventas=args.with_ventas)
        status()
