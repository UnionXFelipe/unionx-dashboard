"""
Actualización automática semanal - UnionX Reportes
Corre cada lunes a las 13:00 vía Programador de tareas de Windows.

Fuentes:
  1. Google Sheets "Matriz stock" / BaseStk         → Excel STOCK
  2. Google Sheets "Importaciones UnionX Integrada" → Excel BASE TRANSITOS
  3. Google Drive  "Raw ventas Y.xlsx"              → Excel Raw

También:
  - Compara datos nuevos vs anteriores y reporta cambios.
  - Expande rangos de tablas dinámicas que apunten a Raw.
  - Refresca todas las tablas dinámicas del libro.
"""

import sys, io as _io
# Forzar UTF-8 en consola Windows
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
if hasattr(sys.stderr, 'reconfigure'):
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')

from format_td_reportes_cst import create_reporte_cst_formato

import gspread
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
import openpyxl
import openpyxl.utils
import xlwings as xw
import io, sys
from datetime import datetime

# ─── CONFIGURACIÓN ────────────────────────────────────────────────────────────

CREDENTIALS_FILE = r'C:\Users\felip\Desktop\UnionX Cloude\credentials.json'
EXCEL_PATH       = r'C:\Users\felip\Desktop\UNIONX\FORECAST FINAL SKU\FORECAST FINAL SKU 26-27 V2.xlsx'
LOG_FILE         = r'C:\Users\felip\Desktop\UnionX Cloude\log.txt'

SHEET1_ID  = '1N5TpIQrFCJwzxyxtueyNi2UoASONd_TbiLcz17a3fag'
SHEET1_TAB = 'BaseStk'

SHEET2_ID  = '1RpxZ69Wnfcots006Hp5fzawYxhUsscW03O_hD3psjHA'
SHEET2_TAB = 0   # primera pestaña "Confirmado EN TRANSITO"

DRIVE_FILE_ID    = '1K11y6icDm9M3X3glGUVCOe4HsbpWpEBm'
RAW_DRIVE_SHEET  = 'RAW'   # hoja dentro del xlsx que coincide con los 40 encabezados

EXCEL_TAB_STOCK     = 'STOCK'
EXCEL_TAB_TRANSITOS = 'BASE TRANSITOS'
EXCEL_TAB_RAW       = 'Raw'
RAW_KEY_RANGE       = 'Raw'   # nombre usado en SourceData de pivots

# Columna N de BASE TRANSITOS: fórmula de MES según fecha ETA bodega (col M)
# Día <= 10 → mes actual | Día > 10 → mes siguiente
TRANSITOS_MES_COL       = 14   # columna N (1-based)
TRANSITOS_FECHA_COL     = 13   # columna M (1-based) — Fecha ETA bodega
TRANSITOS_DATA_START    = 14   # primera fila de datos (bajo las 13 filas reservadas)
TRANSITOS_MES_FORMULA   = '=SI({fecha}="";"";SI(DIA({fecha})<=10;MES({fecha});SI(MES({fecha})=12;1;MES({fecha})+1)))'

SCOPES = [
    'https://spreadsheets.google.com/feeds',
    'https://www.googleapis.com/auth/drive',
]

# ─── LOG ──────────────────────────────────────────────────────────────────────

def log(msg, indent=0):
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    prefix = "  " * indent
    line = f"[{timestamp}] {prefix}{msg}"
    print(line)
    with open(LOG_FILE, 'a', encoding='utf-8') as f:
        f.write(line + '\n')

# ─── LECTURA DE DATOS EXCEL (antes de sobrescribir) ───────────────────────────

def normalize_val(val):
    """
    Normaliza un valor para comparación consistente entre Excel y Google Sheets.
    - Floats enteros (3480.0)  → "3480"
    - Fechas/datetimes          → "YYYY-MM-DD"
    - Fechas "DD/M/YYYY"        → "YYYY-MM-DD"  (formato latinoamericano de Sheets)
    - None / vacío              → ""
    - Todo lo demás             → str limpio
    """
    import re
    from datetime import datetime as _dt, date as _date
    if val is None:
        return ''
    if isinstance(val, float):
        if val == int(val):
            return str(int(val))
        return str(round(val, 6))
    if isinstance(val, (_dt, _date)):
        return str(val)[:10]
    s = str(val).strip()
    # Eliminar .0 de strings como "3480.0"
    if s.endswith('.0') and s[:-2].lstrip('-').isdigit():
        s = s[:-2]
        return s
    # Normalizar fecha latinoamericana DD/MM/YYYY o D/M/YYYY → YYYY-MM-DD
    m = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{4})$', s)
    if m:
        d, mo, y = m.groups()
        return f"{y}-{mo.zfill(2)}-{d.zfill(2)}"
    return s


def read_tab_as_dict(ws_xw, key_col_name):
    """
    Lee un tab de Excel via xlwings.
    Retorna (headers, {key: [normalized_values]}) para comparar.
    """
    used = ws_xw.used_range
    data = used.value
    if not data:
        return [], {}
    if not isinstance(data[0], list):
        data = [data]
    headers = [normalize_val(h) for h in data[0]]
    rows = data[1:]

    key_idx = next((i for i, h in enumerate(headers)
                    if h.lower() == key_col_name.lower()), None)
    row_dict = {}
    if key_idx is not None:
        for row in rows:
            key = normalize_val(row[key_idx] if key_idx < len(row) else None)
            if key:
                row_dict[key] = [normalize_val(v) for v in row]
    return headers, row_dict


def read_tab_row_count(ws_xw):
    used = ws_xw.used_range
    return (used.last_cell.row - 1) if used else 0   # -1 para no contar encabezado

# ─── COMPARACIÓN ──────────────────────────────────────────────────────────────

def compare_and_report(tab_name, old_dict, old_headers, new_rows, new_headers, key_col_name):
    """
    Compara usando SOLO las columnas comunes entre Excel y Sheets para evitar falsos positivos.
    Retorna (agregados, eliminados, modificados).
    """
    key_idx_new = next((i for i, h in enumerate(new_headers)
                        if h and str(h).strip().lower() == key_col_name.lower()), None)
    if key_idx_new is None:
        log(f"   ADVERTENCIA: columna '{key_col_name}' no encontrada en datos nuevos.", 1)
        return 0, 0, 0

    # Columnas comunes entre Excel y Sheets (comparación justa, ignora fórmulas)
    old_h_norm = [normalize_val(h).lower() for h in old_headers]
    common_pairs = []
    for n_idx, n_hdr in enumerate(new_headers):
        n_norm = normalize_val(n_hdr).lower()
        if not n_norm:
            continue
        if n_norm in old_h_norm:
            common_pairs.append((old_h_norm.index(n_norm), n_idx))

    def extract(row, indices):
        return [normalize_val(row[i]) if i < len(row) else '' for i in indices]

    old_idxs = [p[0] for p in common_pairs]
    new_idxs = [p[1] for p in common_pairs]

    new_dict = {}
    for row in new_rows:
        key = normalize_val(row[key_idx_new] if key_idx_new < len(row) else None)
        if key:
            new_dict[key] = extract(row, new_idxs)

    old_dict_cmp = {k: extract(v, old_idxs) for k, v in old_dict.items()}

    old_keys = set(old_dict_cmp.keys())
    new_keys = set(new_dict.keys())

    agregados  = sorted(new_keys - old_keys)
    eliminados = sorted(old_keys - new_keys)
    modificados = [k for k in old_keys & new_keys if old_dict_cmp[k] != new_dict[k]]

    total_old = len(old_dict)
    total_new = len(new_dict)

    log(f"── Comparación [{tab_name}] ──────────────────────────────")
    log(f"   Filas anteriores : {total_old}", 1)
    log(f"   Filas nuevas     : {total_new}", 1)

    if not agregados and not eliminados and not modificados:
        log("   ✅ Sin cambios — datos idénticos a la versión anterior.", 1)
    else:
        if agregados:
            log(f"   ➕ {len(agregados)} SKU(s) nuevos:", 1)
            for k in agregados[:20]:
                log(f"      • {k}", 1)
            if len(agregados) > 20:
                log(f"      ... y {len(agregados)-20} más.", 1)
        if eliminados:
            log(f"   ➖ {len(eliminados)} SKU(s) eliminados:", 1)
            for k in eliminados[:20]:
                log(f"      • {k}", 1)
            if len(eliminados) > 20:
                log(f"      ... y {len(eliminados)-20} más.", 1)
        if modificados:
            log(f"   ✏️  {len(modificados)} SKU(s) con valores modificados:", 1)
            for k in modificados[:20]:
                log(f"      • {k}", 1)
            if len(modificados) > 20:
                log(f"      ... y {len(modificados)-20} más.", 1)

    return len(agregados), len(eliminados), len(modificados)


def compare_raw(old_count, new_count):
    log(f"── Comparación [Raw ventas Y] ────────────────────────────")
    log(f"   Filas anteriores : {old_count}", 1)
    log(f"   Filas nuevas     : {new_count}", 1)
    diff = new_count - old_count
    if diff == 0:
        log("   ✅ Sin cambios en cantidad de filas.", 1)
    elif diff > 0:
        log(f"   ➕ {diff} filas nuevas agregadas.", 1)
    else:
        log(f"   ➖ {abs(diff)} filas eliminadas.", 1)

# ─── ACTUALIZACIÓN POR COLUMNAS COINCIDENTES ──────────────────────────────────

def parse_sheet_val(val):
    """
    Convierte strings de Google Sheets (formato chileno) a tipos Python nativos,
    para que xlwings los escriba correctamente en Excel.

    Fechas (DD/MM/YYYY o D/M/YYYY):
      "27/2/2026"   → date(2026, 2, 27)   ← evita que Excel invierta día y mes
      "3/4/2026"    → date(2026, 4, 3)
      "8/04/2026"   → date(2026, 4, 8)

    Números enteros con punto como miles:
      "3.480"       → 3480
      "1.234.567"   → 1234567

    Números con punto miles + coma decimal:
      "1.234,56"    → 1234.56

    Decimal con coma sola:
      "3,14"        → 3.14

    Cualquier otro string no reconocido → se devuelve tal cual.
    """
    import re
    from datetime import date as _date
    if not isinstance(val, str):
        return val
    s = val.strip()
    if not s:
        return val

    # Fecha chilena D/M/YYYY o DD/MM/YYYY (con o sin cero al inicio)
    m = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{4})$', s)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        try:
            return _date(y, mo, d)
        except ValueError:
            pass   # fecha inválida, continuar

    # Entero con punto como miles  → \d{1,3}(\.\d{3})+
    if re.match(r'^\d{1,3}(\.\d{3})+$', s):
        return int(s.replace('.', ''))

    # Número con punto miles Y coma decimal
    if re.match(r'^\d{1,3}(\.\d{3})+,\d+$', s):
        return float(s.replace('.', '').replace(',', '.'))

    # Decimal con coma sola
    if re.match(r'^\d+,\d+$', s):
        return float(s.replace(',', '.'))

    # Int / float simples
    try:
        return int(s)
    except ValueError:
        pass
    try:
        return float(s)
    except ValueError:
        pass

    return val


def parse_sheet_dates_only(val):
    """
    Convierte fechas chilenas DD/MM/YY o DD/MM/YYYY → datetime.date.
    No toca números. Usado en BASE TRANSITOS: preserva comas/puntos
    en números pero convierte fechas a date para evitar que Excel
    invierta día y mes al recibirlas como string.
    """
    import re
    from datetime import date as _date
    if not isinstance(val, str):
        return val
    s = val.strip()
    # DD/MM/YYYY (4 dígitos)
    m = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{4})$', s)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        try:
            return _date(y, mo, d)
        except ValueError:
            pass
    # DD/MM/YY (2 dígitos → siglo 21: 26 → 2026)
    m = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{2})$', s)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3)) + 2000
        try:
            return _date(y, mo, d)
        except ValueError:
            pass
    return val


def set_transitos_column_formats(ws_xw, sheet_rows, sheet_headers, data_start_row):
    """
    Establece el formato de las celdas de BASE TRANSITOS ANTES de escribir datos.
    - Columnas con fechas (DD/MM/YY o DD/MM/YYYY) → formato 'DD/MM/YY'
    - Columnas con enteros puros o miles chilenos  → formato '0' (número)
    - Columnas con moneda '$X,YYY'                → conservan su formato existente
    - Resto de columnas                            → formato '@' (texto)

    Las columnas numéricas (SKU, cantidades) se guardan como número entero para
    que las tablas dinámicas puedan leerlas y operar sobre ellas correctamente.
    """
    import re
    last_col = ws_xw.used_range.last_cell.column
    excel_headers = ws_xw.range((1, 1), (1, last_col)).value
    if not isinstance(excel_headers, list):
        excel_headers = [excel_headers]
    mapping = build_column_mapping(excel_headers, sheet_headers)
    date_pattern     = re.compile(r'^\d{1,2}/\d{1,2}/\d{2,4}$')
    currency_pattern = re.compile(r'^\$\d+,\d+$')
    # Moneda con miles chilenos: '$21.888,000', '$1.234,56'
    currency_miles_pattern = re.compile(r'^\$\d{1,3}(\.\d{3})*,\d+$')
    # Entero puro (ej. "12345") o entero con miles chilenos (ej. "1.250", "1.234.567")
    int_pattern      = re.compile(r'^\d{1,3}(\.\d{3})*$')
    # Decimal con coma: "19,65", "9,70" — flotantes en notación chilena
    decimal_comma_pattern = re.compile(r'^-?\d+,\d+$')
    date_cols     = set()
    currency_cols = set()
    numeric_cols  = set()   # SKU, cantidades → se escriben como int, formato '0'
    float_cols    = set()   # decimales con coma → se escriben como float, formato '#.##0,00'

    for s_idx, e_idx in mapping.items():
        for row in sheet_rows:
            val = row[s_idx] if s_idx < len(row) else None
            if not val or not isinstance(val, str):
                continue
            v = val.strip()
            if not v:
                continue
            # Clasificar según el primer valor no vacío encontrado
            if date_pattern.match(v):
                date_cols.add(e_idx)
            elif currency_pattern.match(v) or currency_miles_pattern.match(v):
                currency_cols.add(e_idx)
            elif int_pattern.match(v):
                numeric_cols.add(e_idx)
            elif decimal_comma_pattern.match(v):
                float_cols.add(e_idx)
            # else: columna de texto → sin categoría (recibirá '@')
            break   # clasificar solo con el primer valor representativo

    last_row = data_start_row + len(sheet_rows) - 1
    for s_idx, e_idx in mapping.items():
        col_rng = ws_xw.range((data_start_row, e_idx + 1), (last_row, e_idx + 1))
        if e_idx in date_cols:
            col_rng.number_format = 'DD/MM/YY'
        elif e_idx in currency_cols:
            pass   # conservar formato $#.##0,00 existente — se escribirá float
        elif e_idx in numeric_cols:
            col_rng.number_format = '0'   # entero: SKU, cantidades
        elif e_idx in float_cols:
            col_rng.number_format = '#.##0,00'   # decimal: costo uni USD, etc.
        else:
            col_rng.number_format = '@'   # texto: sin conversión automática
    n_text = len(mapping) - len(date_cols) - len(currency_cols) - len(numeric_cols) - len(float_cols)
    log(f"   Formatos preestablecidos: {len(date_cols)} col(s) fecha, "
        f"{len(numeric_cols)} col(s) número entero, "
        f"{len(float_cols)} col(s) decimal, "
        f"{len(currency_cols)} col(s) moneda, "
        f"{n_text} col(s) texto (@).", 1)


# Columnas de BASE TRANSITOS que son fórmulas propias en Excel y NO deben
# sobreescribirse aunque existan en Google Sheets con el mismo nombre.
TRANSITOS_FORMULA_COLS = {'mes', 'stock actual', 'tipo categoria', 'valor usd total', 'marca'}

def build_column_mapping(excel_headers, sheet_headers, skip_excel_names=None):
    """
    skip_excel_names: set de nombres (lower) de columnas Excel que se excluyen
    aunque coincidan con Google Sheets (ej. columnas fórmula de BASE TRANSITOS).
    """
    skip = {s.lower() for s in skip_excel_names} if skip_excel_names else set()
    mapping = {}
    for s_idx, s_hdr in enumerate(sheet_headers):
        if not s_hdr or str(s_hdr).strip() == '':
            continue
        s_norm = str(s_hdr).strip().lower()
        for e_idx, e_hdr in enumerate(excel_headers):
            if e_hdr and str(e_hdr).strip().lower() == s_norm:
                if s_norm not in skip:
                    mapping[s_idx] = e_idx   # 0-based
                break
    return mapping


def clear_autofilter(ws_xw):
    """
    Quita filtros activos en la hoja para que todas las filas sean visibles
    y used_range devuelva el rango completo real.
    """
    try:
        if ws_xw.api.AutoFilterMode:
            try:
                ws_xw.api.ShowAllData()   # muestra filas ocultas por filtro
            except Exception:
                pass
            ws_xw.api.AutoFilterMode = False   # elimina el autofilter
            log(f"   Filtro eliminado en '{ws_xw.name}'.", 1)
    except Exception:
        pass


def update_transitos_mes_formula(ws_xw, num_rows):
    """
    Extiende TODAS las columnas de fórmula de BASE TRANSITOS (N-R) hasta la
    última fila de datos usando AutoFill desde la fila template (14).
    Cols de fórmula: N=MES, O=STOCK ACTUAL, P=Tipo Categoria, Q=Valor USD TOTAL, R=MARCA.
    El AutoFill es más robusto que formula_local fila-por-fila y ajusta referencias automáticamente.
    """
    last_col = ws_xw.used_range.last_cell.column
    last_data_row = TRANSITOS_DATA_START + num_rows - 1  # última fila con datos

    # Columnas de fórmula: N=14, O=15, P=16, Q=17, R=18
    formula_cols = list(range(TRANSITOS_MES_COL, last_col + 1))   # desde N hasta última col con encabezado

    extended = 0
    for col in formula_cols:
        # Verificar que row 14 tenga fórmula (template)
        tmpl_formula = ws_xw.range((TRANSITOS_DATA_START, col)).api.FormulaLocal
        if not tmpl_formula or not str(tmpl_formula).startswith('='):
            continue   # no hay template — omitir
        if last_data_row <= TRANSITOS_DATA_START:
            continue
        # AutoFill desde row 14 hasta last_data_row
        src = ws_xw.range((TRANSITOS_DATA_START, col), (TRANSITOS_DATA_START, col))
        dst = ws_xw.range((TRANSITOS_DATA_START, col), (last_data_row, col))
        try:
            src.api.AutoFill(dst.api, 0)   # 0 = xlFillDefault
            extended += 1
        except Exception as e:
            log(f"   ADVERTENCIA AutoFill col {col}: {e}", 2)

    # Limpiar filas sobrantes en col N (MES) si la base achicó vs semana anterior
    old_last = ws_xw.used_range.last_cell.row
    end_row  = last_data_row + 1
    if old_last >= end_row:
        for col in formula_cols:
            try:
                ws_xw.range((end_row, col), (old_last, col)).clear_contents()
            except Exception:
                pass

    log(f"   {extended} columnas de fórmula extendidas hasta fila {last_data_row} ({num_rows} filas).", 1)


def update_matched_columns(ws_xw, sheet_rows, sheet_headers, data_start_row=2, parse_values=True):
    """
    data_start_row: fila de Excel donde empieza el pegado (1-based).
    Por defecto 2 (justo después del encabezado).
    BASE TRANSITOS usa 14 para preservar las filas 2-13.
    parse_values: si False, escribe los valores tal como vienen de Sheets
    sin convertir formato chileno (comas/puntos). Usar False en BASE TRANSITOS.
    """
    last_col = ws_xw.used_range.last_cell.column
    excel_headers = ws_xw.range((1, 1), (1, last_col)).value
    if not isinstance(excel_headers, list):
        excel_headers = [excel_headers]

    mapping = build_column_mapping(excel_headers, sheet_headers,
                                   skip_excel_names=TRANSITOS_FORMULA_COLS if data_start_row == 14 else None)
    if not mapping:
        log("   ADVERTENCIA: no se encontraron columnas coincidentes.", 1)
        return

    managed_cols = sorted(set(mapping.values()))
    last_row = ws_xw.used_range.last_cell.row

    # Limpiar columnas gestionadas solo desde data_start_row en adelante
    if last_row >= data_start_row:
        for col_0 in managed_cols:
            ws_xw.range((data_start_row, col_0 + 1), (last_row, col_0 + 1)).clear_contents()

    # Escribir columna por columna desde data_start_row
    for s_idx, e_idx in mapping.items():
        col_data = []
        for row in sheet_rows:
            val = row[s_idx] if s_idx < len(row) else None
            if val == '':
                val = None
            elif parse_values:
                val = parse_sheet_val(val)
            else:
                # 1. Limpiar espacios sobrantes (ej. "  100 " → "100")
                if isinstance(val, str):
                    val = val.strip()
                # 2. Convertir fechas DD/MM/YY → datetime.date (evita inversión día/mes)
                val = parse_sheet_dates_only(val)
                # 3. Si sigue siendo string: manejar patrones especiales
                if isinstance(val, str):
                    import re as _re
                    # '$2,200' → float 2.2 (evita que Excel lo lea como $2.200 = 2000)
                    m = _re.match(r'^\$(\d+),(\d+)$', val)
                    if m:
                        val = float(f"{m.group(1)}.{m.group(2)}")
                    # '$21.888,000' o '$1.234,56' (moneda con miles y decimal chileno) → float
                    elif _re.match(r'^\$\d{1,3}(\.\d{3})*,\d+$', val):
                        # Eliminar $ y puntos de miles, reemplazar coma decimal por punto
                        num_str = val[1:].replace('.', '').replace(',', '.')
                        val = float(num_str)
                    # "1.5" → "1,5" | "12.34" → "12,34" (decimal con punto → coma)
                    # NO toca miles chilenos: "1.234" tiene 3 dígitos → queda igual
                    elif _re.match(r'^-?\d+\.\d{1,2}$', val):
                        val = val.replace('.', ',')
                    # Decimal con coma: "19,65" → 19.65, "9,70" → 9.70 (float)
                    elif _re.match(r'^-?\d+,\d+$', val):
                        val = float(val.replace(',', '.'))
                    # Entero puro → int (SKU, cantidades sin separadores)
                    elif _re.match(r'^\d+$', val):
                        val = int(val)
                    # Entero con miles chilenos → int (ej. "1.250" → 1250, "1.234.567" → 1234567)
                    elif _re.match(r'^\d{1,3}(\.\d{3})+$', val):
                        val = int(val.replace('.', ''))
            col_data.append([val])
        if col_data:
            ws_xw.range((data_start_row, e_idx + 1)).value = col_data
            # Limpiar celdas que quedaron como texto-vacío '' en lugar de blank real.
            # Ocurre en celdas con formato '@' cuando xlwings escribe None → Excel almacena ''.
            # '' en aritmética causa #VALOR!; blank real se trata como 0.
            if not parse_values:
                for row_offset, cell_val in enumerate(col_data):
                    if cell_val[0] is None:
                        r = data_start_row + row_offset
                        raw = ws_xw.range((r, e_idx + 1)).api.Value
                        if raw == '':
                            ws_xw.range((r, e_idx + 1)).api.ClearContents()

    log(f"   {len(sheet_rows)} filas escritas desde fila {data_start_row} en {len(managed_cols)} columnas.", 1)


def sanitize_raw_rows(raw_rows):
    """
    Convierte tipos no compatibles con COM (datetime.time) a string.
    xlwings no puede serializar datetime.time directamente.
    """
    import datetime as _dt
    result = []
    for row in raw_rows:
        new_row = []
        for val in row:
            if isinstance(val, _dt.time):
                new_row.append(val.strftime('%H:%M:%S'))
            else:
                new_row.append(val)
        result.append(new_row)
    return result


def update_raw_tab(ws_xw, raw_rows):
    last_row = ws_xw.used_range.last_cell.row
    last_col = ws_xw.used_range.last_cell.column
    if last_row >= 2:
        ws_xw.range((2, 1), (last_row, last_col)).clear_contents()
    if raw_rows:
        ws_xw.range((2, 1)).value = sanitize_raw_rows(raw_rows)
    log(f"   {len(raw_rows)} filas escritas en Raw.", 1)
    return len(raw_rows)

# ─── PRESERVAR ORDEN TD VENTAS N ─────────────────────────────────────────────

VENTAS_N_SHEET = 'TD VENTAS N'
VENTAS_N_PIVOT = 'TablaDinámica5'
XL_MANUAL      = -4135   # xlManual en Excel VBA


def capture_pivot_sort(wb_xw):
    """
    Captura el orden manual de ítems de todos los campos de fila/columna
    en la TD VENTAS N, justo antes del refresh.
    Retorna un dict: {field_name: [(item_name, visible), ...]}
    """
    snapshot = {}
    try:
        sheet = wb_xw.sheets[VENTAS_N_SHEET]
        pt = sheet.api.PivotTables(VENTAS_N_PIVOT)
        fields = pt.PivotFields()
        for j in range(1, fields.Count + 1):
            try:
                f = fields.Item(j)
                if f.Orientation not in (1, 2):   # solo fila (1) o columna (2)
                    continue
                items = f.PivotItems()
                order = []
                for k in range(1, items.Count + 1):
                    item = items.Item(k)
                    order.append((item.Name, item.Visible))
                snapshot[f.Name] = order
            except Exception:
                pass
        log(f"   Orden de TD VENTAS N capturado ({len(snapshot)} campos).", 1)
    except Exception as e:
        log(f"   ADVERTENCIA al capturar orden de TD VENTAS N: {e}", 1)
    return snapshot


def restore_pivot_sort(wb_xw, snapshot):
    """
    Restaura el orden manual de ítems en TD VENTAS N después del refresh.
    Ítems nuevos (no estaban antes) quedan al final sin alterar.
    """
    if not snapshot:
        return
    try:
        sheet = wb_xw.sheets[VENTAS_N_SHEET]
        pt = sheet.api.PivotTables(VENTAS_N_PIVOT)
        fields = pt.PivotFields()
        for j in range(1, fields.Count + 1):
            try:
                f = fields.Item(j)
                if f.Name not in snapshot:
                    continue
                # Forzar orden manual para que Position sea editable
                f.AutoSort(XL_MANUAL, f.Name)
                saved_order = snapshot[f.Name]
                saved_names = [name for name, _ in saved_order]
                # Asignar Position e Visible según el orden guardado
                pos = 1
                for name, visible in saved_order:
                    try:
                        item = f.PivotItems(name)
                        item.Position = pos
                        item.Visible  = visible
                        pos += 1
                    except Exception:
                        pass   # ítem ya no existe en los datos nuevos
                # Los ítems nuevos (no estaban en snapshot) quedan al final
            except Exception:
                pass
        pt.RefreshTable()
        log(f"   Orden de TD VENTAS N restaurado.", 1)
    except Exception as e:
        log(f"   ADVERTENCIA al restaurar orden de TD VENTAS N: {e}", 1)


# ─── TABLAS DINÁMICAS ─────────────────────────────────────────────────────────

def expand_and_refresh_pivots(wb_xw, raw_tab_name, new_last_row, new_header_cols):
    """
    1. Busca todos los PivotCache cuyo SourceData apunte a la pestaña Raw.
    2. Actualiza solo el número de filas dentro del SourceData existente,
       preservando el formato exacto que Excel usa (R1C1 español o A1).
       Patrones soportados:
         R1C1 español : Raw!F1C1:F336024C40   → Raw!F1C1:F{total}C{cols}
         A1 estándar  : 'Raw'!$A$1:$AN$336024 → 'Raw'!$A$1:${col}${total}
    3. Refresca todas las tablas dinámicas del libro.
    """
    import re
    total_rows = new_last_row + 1   # +1 para incluir fila de encabezado
    pivots_updated = 0

    def build_new_source(src):
        # R1C1 español: Raw!F1C1:F<filas>C<cols>
        m = re.match(r"^(.+!F1C1:F)\d+(C\d+)$", src, re.IGNORECASE)
        if m:
            return f"{m.group(1)}{total_rows}{m.group(2)}"
        # A1 estándar: 'Raw'!$A$1:$XX$<filas>
        m = re.match(r"^(.+!\$[A-Z]+\$1:\$[A-Z]+\$)\d+$", src, re.IGNORECASE)
        if m:
            return f"{m.group(1)}{total_rows}"
        return None

    try:
        # Construir el rango COM de origen
        raw_ws_api = wb_xw.sheets[raw_tab_name].api
        src_range  = raw_ws_api.Range(
            raw_ws_api.Cells(1, 1),
            raw_ws_api.Cells(total_rows, new_header_cols)
        )

        # Crear un nuevo PivotCache con el rango actualizado (xlDatabase = 1)
        new_cache = wb_xw.api.PivotCaches().Create(SourceType=1, SourceData=src_range)

        # Reasignar el nuevo cache a todas las TDs que apuntaban a Raw.
        # Para cada TD se intenta ChangePivotCache y luego se VERIFICA que el
        # SourceData realmente cambió — evita falsos positivos cuando el error
        # no era propagación sino un fallo real (TD con cache independiente).
        for sheet in wb_xw.sheets:
            try:
                pts = sheet.api.PivotTables()
                for i in range(1, pts.Count + 1):
                    try:
                        pt  = pts.Item(i)
                        src = str(pt.PivotCache().SourceData)
                        if raw_tab_name.lower() not in src.lower():
                            continue
                        pt_name = pt.Name
                        # Ya apunta al rango correcto (propagación previa)
                        if str(total_rows) in src:
                            pivots_updated += 1
                            log(f"   TD '{pt_name}' en '{sheet.name}' → ya actualizada.", 1)
                            continue
                        # Intentar cambio de cache (vía nuevo PivotCache COM)
                        try:
                            pt.ChangePivotCache(new_cache)
                        except Exception:
                            pass   # puede fallar por propagación automática; verificar abajo
                        # Verificar que el SourceData cambió realmente
                        try:
                            new_src = str(pt.PivotCache().SourceData)
                        except Exception:
                            new_src = ''
                        # Fallback: si ChangePivotCache no propagó, editar el SourceData
                        # del cache existente directamente (resuelve TablaDinámica5 / TablaDinámica2
                        # que comparten cache propio y rechazan ChangePivotCache con E_INVALIDARG)
                        if str(total_rows) not in new_src:
                            fallback_src = build_new_source(src)  # src = valor ANTES del intento
                            if fallback_src:
                                try:
                                    pt.PivotCache().SourceData = fallback_src
                                    new_src = str(pt.PivotCache().SourceData)
                                except Exception as e_fb:
                                    log(f"   ADVERTENCIA fallback SourceData '{pt_name}': {e_fb}", 1)
                        if str(total_rows) in new_src:
                            pivots_updated += 1
                            log(f"   TD '{pt_name}' en '{sheet.name}' → cache actualizado a {total_rows} filas.", 1)
                        else:
                            log(f"   ADVERTENCIA: TD '{pt_name}' en '{sheet.name}' no se actualizó (src: {new_src[:80]}).", 1)
                    except Exception as e:
                        log(f"   ADVERTENCIA TD: {e}", 1)
            except Exception:
                pass
    except Exception as e:
        log(f"   ADVERTENCIA al actualizar pivots de Raw: {e}", 1)

    # Capturar orden manual de TD VENTAS N antes de refrescar
    ventas_n_snapshot = capture_pivot_sort(wb_xw)

    # Refrescar cada tabla dinámica individualmente (síncrono) para que
    # el nuevo SourceData quede confirmado ANTES de que save() ocurra.
    # RefreshAll() es asíncrono y puede que save() corra antes de que Excel
    # confirme el cambio de cache, dejando el origen sin actualizar en disco.
    refreshed = 0
    for sheet in wb_xw.sheets:
        try:
            pts = sheet.api.PivotTables()
            for i in range(1, pts.Count + 1):
                try:
                    pt = pts.Item(i)
                    pt.RefreshTable()   # bloquea hasta terminar
                    refreshed += 1
                except Exception as e:
                    log(f"   ADVERTENCIA al refrescar TD '{sheet.name}': {e}", 1)
        except Exception:
            pass

    # Esperar a que Excel finalice cualquier cálculo/query pendiente
    try:
        wb_xw.api.Application.CalculateUntilAsyncQueriesDone()
    except Exception:
        pass

    log(f"   {refreshed} tabla(s) dinámica(s) refrescadas individualmente.", 1)

    # Restaurar orden manual de TD VENTAS N
    restore_pivot_sort(wb_xw, ventas_n_snapshot)

    return pivots_updated

# ─── VTA X MARCA META ────────────────────────────────────────────────────────

# Abreviaciones de mes en español → número
MES_ABBR = {
    'ENE':1,'FEB':2,'MAR':3,'ABR':4,'MAY':5,'JUN':6,
    'JUL':7,'AGO':8,'SEP':9,'OCT':10,'NOV':11,'DIC':12
}

VTA_SHEET     = 'VTA X marca meta'
TD_VENTAS_N   = 'TD VENTAS N'

# Columnas de VENTA ACUM en la última tabla (1-based)
VTA_ACUM_COLS = [5, 10, 15]   # E, J, O

# Filas con fórmulas propias que NO se deben tocar
VTA_SKIP_LABELS = {'total general', 'total empresa', 'p. nacionales'}

# Normalización de nombres de marca para matching entre las dos hojas
def _norm_brand(s):
    if not s: return ''
    s = str(s).strip().lower()
    # Quitar sufijos entre paréntesis: "uma (mattel)" → "uma"
    import re as _re
    s = _re.sub(r'\s*\(.*\)', '', s)
    return s.strip()


def update_vta_x_marca(wb_xw):
    """
    Lee TD VENTAS N (pivot por marca × mes 2026) y escribe los valores
    en las columnas VENTA ACUM. de la última tabla de VTA X marca meta.
    Solo rellena celdas cuyo mes ya existe en el pivot (meses futuros quedan vacíos).
    Respeta las filas con fórmulas propias (Total general, TOTAL EMPRESA, P. Nacionales).
    """
    # ── Leer TD VENTAS N ─────────────────────────────────────────────────
    ws_td = wb_xw.sheets[TD_VENTAS_N]
    used  = ws_td.used_range
    lr, lc = used.last_cell.row, used.last_cell.column
    td    = ws_td.range((1,1),(lr,lc)).value
    if not isinstance(td[0], list):
        td = [[v] for v in td]

    # Fila 4 (idx 3): años | Fila 5 (idx 4): meses
    yr_row  = td[3]   # [None, None, 2025, None, ..., 'Total 2025', 2026, None, ..., 'Total 2026', ...]
    mon_row = td[4]   # [Tipo, Marca, 1, 2, ..., 12, None, 1, 2, ..., None, None]

    # Encontrar columnas de 2026 (donde yr_row tiene 2026 o None pero después del primer 2026)
    yr2026_start = None
    for j, v in enumerate(yr_row):
        if isinstance(v, float) and int(v) == 2026:
            yr2026_start = j
            break

    if yr2026_start is None:
        log("   ADVERTENCIA: no se encontró año 2026 en TD VENTAS N.", 1)
        return

    # Mapear mes_numero → columna en td (0-based) dentro del bloque 2026
    mes_col_map = {}   # {4: col_idx, 5: col_idx, ...}
    for j in range(yr2026_start, lc):
        v = mon_row[j]
        if isinstance(v, float) and 1 <= v <= 12:
            mes_col_map[int(v)] = j
        elif v is not None and not isinstance(v, float):
            break   # llegamos a 'Total 2026' u otra sección

    # Construir dict: {marca_norm: {mes_num: valor}}
    td_data = {}
    for row in td[5:]:   # datos empiezan en fila 6 (idx 5)
        marca = row[1] if len(row) > 1 else None
        if not marca:
            continue
        marca_n = _norm_brand(marca)
        if not marca_n or marca_n in ('total propia', 'total general'):
            continue
        # "Otras Marcas" en TD VENTAS N = Proveedores Nacionales en su completitud
        if marca_n == 'otras marcas':
            marca_n = 'p. nacionales'
        vals = {}
        for mes_num, col_j in mes_col_map.items():
            v = row[col_j] if col_j < len(row) else None
            if isinstance(v, (int, float)) and v != 0:
                vals[mes_num] = v
        if vals:
            td_data[marca_n] = vals

    log(f"   TD VENTAS N: {len(td_data)} marcas leídas, meses 2026: {sorted(mes_col_map.keys())}", 1)

    # ── Leer VTA X marca meta — última tabla ─────────────────────────────
    ws_vta = wb_xw.sheets[VTA_SHEET]
    used2  = ws_vta.used_range
    lr2, lc2 = used2.last_cell.row, used2.last_cell.column
    vta   = ws_vta.range((1,1),(lr2,lc2)).value
    if not isinstance(vta[0], list):
        vta = [[v] for v in vta]

    # Encontrar la ÚLTIMA fila de cabecera que contenga "VENTA ACUM."
    last_header_row = None
    for i, row in enumerate(vta):
        if any(cell and 'venta acum' in str(cell).lower() for cell in row):
            last_header_row = i   # 0-based

    if last_header_row is None:
        log("   ADVERTENCIA: no se encontró cabecera 'VENTA ACUM.' en VTA X marca meta.", 1)
        return

    # Obtener mes → columna Excel (1-based) desde esa fila de cabecera
    hdr = vta[last_header_row]
    acum_col_map = {}   # {mes_num: col_excel_1based}
    for j, cell in enumerate(hdr):
        if not cell or 'venta acum' not in str(cell).lower():
            continue
        s = str(cell).upper()
        for abbr, num in MES_ABBR.items():
            if abbr in s:
                acum_col_map[num] = j + 1   # 1-based
                break

    log(f"   VTA X marca meta: columnas VENTA ACUM. encontradas: { {k: v for k,v in acum_col_map.items()} }", 1)

    # Escribir valores: recorrer filas de datos de la última tabla
    header_excel_row = last_header_row + 1   # 1-based
    escritas = 0
    for i in range(last_header_row + 1, lr2):
        row = vta[i]
        marca_cell = row[1] if len(row) > 1 else None   # col B
        if not marca_cell:
            continue
        marca_n = _norm_brand(marca_cell)
        if marca_n in VTA_SKIP_LABELS:
            continue
        if marca_n not in td_data:
            log(f"   No encontrado en TD VENTAS N: '{marca_cell}' (norm: '{marca_n}')", 1)
            continue
        for mes_num, col_excel in acum_col_map.items():
            valor = td_data[marca_n].get(mes_num)
            excel_row = i + 1   # 1-based
            if valor is not None:
                ws_vta.range((excel_row, col_excel)).value = valor
                escritas += 1
            # Si el mes no existe aún en el pivot, no se toca la celda

    log(f"   {escritas} celdas actualizadas en VTA X marca meta.", 1)


# ─── MAIN ─────────────────────────────────────────────────────────────────────

def main(skip_analisis=False):
    log("=" * 60)
    log(f"Actualización semanal UnionX — {datetime.now().strftime('%A %d/%m/%Y %H:%M')}")

    creds = Credentials.from_service_account_file(CREDENTIALS_FILE, scopes=SCOPES)
    gc = gspread.authorize(creds)
    drive_service = build('drive', 'v3', credentials=creds)

    # ── Obtener datos nuevos desde Google ─────────────────────────────────────
    log("[1/3] Descargando Matriz stock / BaseStk...")
    sh1    = gc.open_by_key(SHEET1_ID)
    data1  = sh1.worksheet(SHEET1_TAB).get_all_values()
    hdr1   = data1[0] if data1 else []
    rows1  = data1[1:] if len(data1) > 1 else []
    log(f"   {len(rows1)} filas obtenidas.", 1)

    log("[2/3] Descargando Importaciones UnionX...")
    sh2    = gc.open_by_key(SHEET2_ID)
    data2  = sh2.get_worksheet(SHEET2_TAB).get_all_values()
    hdr2   = data2[0] if data2 else []
    rows2  = data2[1:] if len(data2) > 1 else []
    log(f"   {len(rows2)} filas obtenidas.", 1)

    log("[3/3] Descargando Raw ventas Y.xlsx desde Drive...")
    request = drive_service.files().get_media(fileId=DRIVE_FILE_ID)
    fh = io.BytesIO()
    downloader = MediaIoBaseDownload(fh, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    fh.seek(0)
    wb_raw_src = openpyxl.load_workbook(fh, read_only=True, data_only=True)
    if RAW_DRIVE_SHEET in wb_raw_src.sheetnames:
        ws_raw_src = wb_raw_src[RAW_DRIVE_SHEET]
        log(f"   Usando hoja '{RAW_DRIVE_SHEET}' del archivo Drive.", 1)
    else:
        ws_raw_src = wb_raw_src.active
        log(f"   ADVERTENCIA: hoja '{RAW_DRIVE_SHEET}' no encontrada, usando hoja activa '{ws_raw_src.title}'.", 1)
    raw_rows = [list(r) for r in ws_raw_src.iter_rows(min_row=2, values_only=True)
                if any(c is not None for c in r[:5])]   # basta con que alguna de las 5 primeras columnas tenga dato
    raw_col_count = ws_raw_src.max_column or 1
    wb_raw_src.close()
    log(f"   {len(raw_rows)} filas obtenidas ({raw_col_count} columnas).", 1)

    # ── Abrir Excel ───────────────────────────────────────────────────────────
    log("Abriendo Excel con xlwings...")
    app = xw.App(visible=False, add_book=False)
    try:
        wb = app.books.open(EXCEL_PATH)

        # ── Limpiar filtros ANTES de leer (para que used_range sea completo) ─
        log("Limpiando filtros activos en hojas de datos...")
        clear_autofilter(wb.sheets[EXCEL_TAB_STOCK])
        clear_autofilter(wb.sheets[EXCEL_TAB_TRANSITOS])
        clear_autofilter(wb.sheets[EXCEL_TAB_RAW])

        # ── Leer datos ANTERIORES para comparación ────────────────────────────
        log("Leyendo datos anteriores para comparación...")
        old_stock_hdr,     old_stock_dict     = read_tab_as_dict(wb.sheets[EXCEL_TAB_STOCK],     'SKU')
        old_transitos_hdr, old_transitos_dict = read_tab_as_dict(wb.sheets[EXCEL_TAB_TRANSITOS], 'SKU')
        old_raw_count = read_tab_row_count(wb.sheets[EXCEL_TAB_RAW])

        # ── Comparación ───────────────────────────────────────────────────────
        log("")
        log("╔══════════════ REPORTE DE CAMBIOS ══════════════╗")
        compare_and_report(EXCEL_TAB_STOCK,     old_stock_dict,     old_stock_hdr,     rows1, hdr1, 'SKU')
        compare_and_report(EXCEL_TAB_TRANSITOS, old_transitos_dict, old_transitos_hdr, rows2, hdr2, 'SKU')
        compare_raw(old_raw_count, len(raw_rows))
        log("╚════════════════════════════════════════════════╝")
        log("")

        # ── Actualizar datos ──────────────────────────────────────────────────
        log(f"Actualizando '{EXCEL_TAB_STOCK}'...")
        update_matched_columns(wb.sheets[EXCEL_TAB_STOCK], rows1, hdr1)

        log(f"Actualizando '{EXCEL_TAB_TRANSITOS}'...")
        set_transitos_column_formats(wb.sheets[EXCEL_TAB_TRANSITOS], rows2, hdr2, data_start_row=14)
        update_matched_columns(wb.sheets[EXCEL_TAB_TRANSITOS], rows2, hdr2, data_start_row=14, parse_values=False)
        update_transitos_mes_formula(wb.sheets[EXCEL_TAB_TRANSITOS], len(rows2))

        log(f"Actualizando '{EXCEL_TAB_RAW}'...")
        update_raw_tab(wb.sheets[EXCEL_TAB_RAW], raw_rows)

        # ── Expandir pivots de Raw + refrescar todo ───────────────────────────
        log("Expandiendo rangos de tablas dinámicas y refrescando...")
        n = expand_and_refresh_pivots(wb, EXCEL_TAB_RAW, len(raw_rows), raw_col_count)
        log(f"   {n} tabla(s) dinámica(s) de Raw expandidas.", 1)

        # ── Actualizar VTA X marca meta DESPUÉS del refresh de pivots ─────────
        # IMPORTANTE: debe ir después de expand_and_refresh_pivots() para leer
        # los datos actualizados de TD VENTAS N (si va antes lee datos viejos).
        log("Actualizando 'VTA X marca meta'...")
        update_vta_x_marca(wb)

        # ── Regenerar hojas REPORTE CST FLAT y REPORTE UNID FLAT ─────────────
        log("Regenerando 'REPORTE CST FLAT' y 'REPORTE UNID FLAT'...")
        try:
            create_reporte_cst_formato(wb)
        except Exception as e:
            log(f"   ADVERTENCIA: no se pudo regenerar reportes CST/UNID FLAT — {e}", 1)

        # ── Forzar recálculo completo antes de guardar ───────────────────────
        # El archivo puede estar en modo cálculo manual; CalculateFull() fuerza
        # recálculo de TODAS las celdas (incluye FLAT que referencia FCST BASE)
        log("Recalculando fórmulas antes de guardar...")
        try:
            wb.app.api.CalculateFull()
        except Exception:
            try:
                wb.app.api.CalculateUntilAsyncQueriesDone()
            except Exception:
                wb.app.calculate()

        # ── Guardar ───────────────────────────────────────────────────────────
        log("Guardando Excel...")
        wb.save()
        wb.close()

        # ── Seguimiento PPTO (Cómo Vamos + Comp. Marcas/Canales) ────────────
        log("Generando hojas de seguimiento presupuesto...")
        try:
            import subprocess as _sp, sys as _sys, os as _os2
            _seg_dir  = _os2.path.join(_os2.path.dirname(__file__), "Seguimiento PPTO 2026")
            _leer     = _os2.path.join(_seg_dir, "leer_todo.py")
            _crear    = _os2.path.join(_seg_dir, "crear_seguimiento.py")
            _py       = _sys.executable
            if _os2.path.exists(_leer) and _os2.path.exists(_crear):
                r1 = _sp.run([_py, _leer],  capture_output=True, text=True, encoding='cp1252', errors='replace')
                if r1.returncode == 0:
                    log("   ✅ leer_todo.py OK")
                else:
                    log(f"   ⚠️  leer_todo.py error: {r1.stderr[-300:]}", 1)
                r2 = _sp.run([_py, _crear], capture_output=True, text=True, encoding='cp1252', errors='replace')
                if r2.returncode == 0:
                    log("   ✅ crear_seguimiento.py OK — 5 hojas actualizadas en Metas oficiales")
                else:
                    log(f"   ⚠️  crear_seguimiento.py error: {r2.stderr[-300:]}", 1)
            else:
                log("   ℹ️  Scripts de seguimiento no encontrados en Seguimiento PPTO 2026/")
        except Exception as _e:
            log(f"   ⚠️  No se pudo generar seguimiento: {_e}", 1)

        # ── Análisis de planificación (críticos + sobrestock) ─────────────────
        if skip_analisis:
            log("⏭  Análisis de planificación omitido (--no-analisis).")
        else:
            log("Generando análisis de planificación...")
            try:
                from analisis_stock_critico import run_analisis
                run_analisis(log_fn=log)
            except Exception as e:
                log(f"   ADVERTENCIA: no se pudo generar análisis de planificación — {e}", 1)

        # ── Subir análisis + metas a Google Drive ────────────────────────────
        try:
            import os as _os
            _config_path = _os.path.join(_os.path.dirname(__file__), "drive_config.py")
            if _os.path.exists(_config_path):
                import importlib.util as _ilu
                _spec = _ilu.spec_from_file_location("drive_config", _config_path)
                _dc = _ilu.module_from_spec(_spec); _spec.loader.exec_module(_dc)
                _folder_id  = getattr(_dc, "DRIVE_FOLDER_ID", "")
                _metas_fid  = getattr(_dc, "METAS_FILE_ID", "")
                if _folder_id:
                    from drive_utils import get_service, upload_or_update, find_file
                    from analisis_stock_critico import OUT_PATH
                    _svc = get_service(rw=True)

                    # 1) Analisis planificacion
                    _fname = _os.path.basename(OUT_PATH)
                    _existing = find_file(_svc, _folder_id, _fname)
                    upload_or_update(_svc, OUT_PATH, _fname, _folder_id,
                                     existing_id=_existing)
                    log(f"   ☁️  Análisis subido a Drive: {_fname}")

                    # 2) Metas oficiales (actualiza el archivo existente por ID)
                    _metas_local = r"C:\Users\felip\Desktop\UNIONX\PPTO 2026\Metas oficiales 1SEM Nuevo.xlsx"
                    if _metas_fid and _os.path.exists(_metas_local):
                        upload_or_update(_svc, _metas_local,
                                         _os.path.basename(_metas_local),
                                         _folder_id, existing_id=_metas_fid)
                        log(f"   ☁️  Metas subido a Drive: {_os.path.basename(_metas_local)}")
                    elif not _metas_fid:
                        log("   ℹ️  METAS_FILE_ID vacío en drive_config.py — omitiendo upload metas.")
                    else:
                        log(f"   ⚠️  Metas no encontrado localmente: {_metas_local}")
                else:
                    log("   ℹ️  DRIVE_FOLDER_ID vacío en drive_config.py — omitiendo upload.")
            else:
                log("   ℹ️  drive_config.py no encontrado — ejecuta setup_drive.py para activar upload.")
        except Exception as _e:
            log(f"   ⚠️  No se pudo subir a Drive: {_e}", 1)

        log("✅ Actualización completada exitosamente!")

    except Exception as e:
        raise e
    finally:
        app.quit()

    log("=" * 60)
    log("")


if __name__ == '__main__':
    _skip = '--no-analisis' in sys.argv
    try:
        main(skip_analisis=_skip)
    except Exception as e:
        log(f"❌ ERROR: {e}")
        import traceback
        log(traceback.format_exc())
        sys.exit(1)
