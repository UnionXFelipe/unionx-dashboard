"""
Análisis de planificación de inventario UnionX
  - Hoja 1+2: Stock Crítico   (cobertura < 1 mes)
  - Hoja 3+4: Sobrestock      (cobertura > 6 meses)

Entry point: run_analisis(log_fn=print)
  Llamado desde actualizar_reportes.py cada lunes tras guardar el Excel.
"""
import sys, datetime as _dt, calendar
from collections import defaultdict
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Rutas ─────────────────────────────────────────────────────────────
EXCEL_PATH = (r'C:\Users\felip\Desktop\UNIONX\FORECAST FINAL SKU'
              r'\FORECAST FINAL SKU 26-27 V2.xlsx')

# ── Mes base dinámico (se auto-ajusta cada mes) ────────────────────────
_TODAY      = _dt.date.today()
_CUR_MONTH  = _TODAY.month   # 4=ABR, 5=MAY, 6=JUN, ...
_MES_ABBR_DYN = {4:'ABR', 5:'MAY', 6:'JUN', 7:'JUL', 8:'AGO', 9:'SEP', 10:'OCT'}
_CUR_MES     = _MES_ABBR_DYN.get(_CUR_MONTH, 'ABR')
_CUR_MES_KEY = f'{_CUR_MES}26'       # e.g. 'MAY26'
_CUR_MES_YEAR = f'{_CUR_MES} 2026'   # e.g. 'MAY 2026'

OUT_PATH = (r'C:\Users\felip\Desktop\UNIONX\FORECAST FINAL SKU'
            fr'\Analisis Planificacion\analisis_planificacion_{_CUR_MES}26.xlsx')

# ── Columnas tránsito FCST BASE (0-based) — verificadas en fila 3 de FCST BASE SKU MACRO V2 ──
# V2 con columna "Sku Padre" en idx 4 — todos los índices originales +1.
# idx 4='Sku Padre', idx 5='Sku', idx 10='Ranking Comercial', idx 19='Puerto Origen'
# idx 101=Transito ABR 26, idx 114=Embarcado Mayo (MAY usa Embarcado, no Transito),
# idx 126=Transito JUN 26, idx 138=Transito JUL 26,
# idx 150=Transito AGO 26, idx 162=Transito SEP 26, idx 174=Transito OCT 26
TRANSITO_FCST = {
    'ABR26': 101, 'MAY26': 114, 'JUN26': 126,
    'JUL26': 138, 'AGO26': 150, 'SEP26': 162, 'OCT26': 174,
}
FCST_IDX_SKU_PADRE = 4   # 'Sku Padre'
FCST_IDX_SKU       = 5   # 'Sku'
FCST_IDX_RANKING   = 10  # 'Ranking Comercial'
FCST_IDX_PUERTO    = 19  # 'Puerto Origen'
MESES_ORDEN = ['ABR26','MAY26','JUN26','JUL26','AGO26','SEP26','OCT26']

# ── Posiciones en hojas FLAT (0-based) ────────────────────────────────
CST_STOCK  = 6;  CST_COBERT = 7
UNID_STOCK = 6
# Venta del mes actual — se desplaza 6 cols por mes a partir de MAY
# ABR(4): idx 10 (K)  | MAY+(n): 8+4 + (n-5)*6 + 4 = (n-5)*6 + 16
CST_VENTA  = 10 if _CUR_MONTH == 4 else (_CUR_MONTH - 5) * 6 + 16
# UNID: ABR(4): idx 11 (L) | MAY+(n): (n-5)*6 + 17
UNID_VENTA = 11 if _CUR_MONTH == 4 else (_CUR_MONTH - 5) * 6 + 17

# ── Colores ───────────────────────────────────────────────────────────
C_ROJO  = 'C0392B'; C_ROJO_BG  = 'FADBD8'
C_NARAN = 'CA6F1E'; C_NARAN_BG = 'FDEBD0'
C_VERDE = '1E8449'; C_VERDE_BG = 'D5F5E3'
C_MORA  = '6C3483'; C_MORA_BG  = 'E8DAEF'
C_HDR   = '2C3E50'; C_HDR_FG   = 'FFFFFF'
C_SUB   = 'AED6F1'; C_SUB_FG   = '1A252F'
C_ALT   = 'EBF5FB'; C_BORDE    = 'BDC3C7'

# Colores adicionales para hojas de tránsito
C_AZUL      = '1A5276'; C_AZUL_BG      = 'D6EAF8'   # azul marino — embarques
C_TEAL      = '0E6655'; C_TEAL_BG      = 'D0ECE7'   # teal — nuevos
C_GRIS_SEP  = '4D5656'; C_GRIS_SEP_BG  = 'D5D8DC'   # gris — separadores de mes

thin  = Side(style='thin', color=C_BORDE)
borde = Border(left=thin, right=thin, top=thin, bottom=thin)

# ══════════════════════════════════════════════════════════════════════
# HELPERS DE ESTILO
# ══════════════════════════════════════════════════════════════════════
def hdr(ws, row, col, val, w=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(bold=True, color=C_HDR_FG, name='Calibri', size=9)
    c.fill = PatternFill('solid', fgColor=C_HDR)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = borde
    if w: ws.column_dimensions[get_column_letter(col)].width = w

def cel(ws, row, col, val, align='left', fmt=None, bold=False,
        bg=None, fg='000000', size=9):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(name='Calibri', size=size, bold=bold, color=fg)
    c.alignment = Alignment(horizontal=align, vertical='center')
    c.border = borde
    if bg:  c.fill = PatternFill('solid', fgColor=bg)
    if fmt: c.number_format = fmt

def flt(v, default=0):
    if v is None or v == '': return default
    try: return float(v)
    except: return default

def _fmt_sku(v):
    """Normaliza cualquier valor de SKU a string limpio sin '.0'.
    float 12345.0 → '12345' | str '12345.0' → '12345' | str 'ABC-01' → 'ABC-01'
    """
    if v is None: return ''
    if isinstance(v, float): return str(int(v))
    s = str(v).strip()
    if s.endswith('.0') and s[:-2].lstrip('-').isdigit():
        return s[:-2]
    return s

# ══════════════════════════════════════════════════════════════════════
# FUNCIONES DE COLOR
# ══════════════════════════════════════════════════════════════════════
def cob_color_critico(cob):
    if cob == 0:  return 'F9EBEA', C_ROJO
    if cob < 0.5: return C_ROJO_BG, C_ROJO
    return C_NARAN_BG, C_NARAN

def cob_color_sobre(cob):
    if cob > 12: return C_MORA_BG, C_MORA
    if cob > 9:  return C_ROJO_BG, C_ROJO
    return C_NARAN_BG, C_NARAN

C_AMAR  = 'B7950B'; C_AMAR_BG  = 'FCF3CF'   # amarillo — inquietante 1-1.5 meses

def cob_color_inq(cob):
    """Color para coberturas inquietantes (0 – 1.5 meses)."""
    if cob == 0:   return 'F9EBEA', C_ROJO
    if cob < 0.5:  return C_ROJO_BG,  C_ROJO
    if cob < 1.0:  return C_NARAN_BG, C_NARAN
    return C_AMAR_BG, C_AMAR   # 1 – 1.5: amarillo

def _is_valid_prox(mes_str):
    """
    Un mes es 'válido' como próxima llegada solo si su fecha límite de recepción
    (día 10 del mes, según la regla 10-día) aún no ha pasado.
    Ejemplo: hoy=20-ABR → ABR26 cutoff=10-ABR < hoy → inválido → se omite.
    """
    cutoff = {
        'ABR26': _dt.date(2026, 4, 10), 'MAY26': _dt.date(2026, 5, 10),
        'JUN26': _dt.date(2026, 6, 10), 'JUL26': _dt.date(2026, 7, 10),
        'AGO26': _dt.date(2026, 8, 10), 'SEP26': _dt.date(2026, 9, 10),
        'OCT26': _dt.date(2026, 10, 10),
    }.get(mes_str)
    return cutoff is None or cutoff >= _dt.date.today()


def fecha_max_carga(prox, puerto):
    """
    Calcula la fecha máxima de carga para que el producto llegue dentro del mes indicado.
    Lógica: último día del mes anterior al mes de llegada − días de tránsito según puerto.
      Shenzhen → 70 días | Ningbo → 55 días
    """
    if not prox or prox == 'Sin embarque' or not puerto:
        return ''
    techo_mes = {
        'ABR26': _dt.date(2026, 3, 31),
        'MAY26': _dt.date(2026, 4, 30),
        'JUN26': _dt.date(2026, 5, 31),
        'JUL26': _dt.date(2026, 6, 30),
        'AGO26': _dt.date(2026, 7, 31),
        'SEP26': _dt.date(2026, 8, 31),
        'OCT26': _dt.date(2026, 9, 30),
    }
    techo = techo_mes.get(prox)
    if not techo:
        return ''
    dias = 70 if 'shenzhen' in puerto.lower() else 55  # Ningbo = 55 días
    return (techo - _dt.timedelta(days=dias)).strftime('%d/%m/%Y')


def prox_color_critico(prox):
    if prox == _CUR_MES_KEY:   return C_VERDE_BG, C_VERDE   # llega este mes → bueno
    if prox == 'Sin embarque': return C_ROJO_BG,  C_ROJO
    return C_NARAN_BG, C_NARAN

def prox_color_sobre(prox):
    if prox == 'Sin embarque': return C_VERDE_BG, C_VERDE   # sin llegadas → bueno
    if prox == _CUR_MES_KEY:   return C_ROJO_BG,  C_ROJO   # llega este mes → peor
    return C_NARAN_BG, C_NARAN

# ══════════════════════════════════════════════════════════════════════
# AGRUPACIÓN
# ══════════════════════════════════════════════════════════════════════
def agrupar(results):
    grupos = defaultdict(list)
    for r in results:
        grupos[(r['marca'], r['cat_com'])].append(r)
    resumen = []
    for (marca, cat), rows in sorted(grupos.items()):
        n = len(rows)
        resumen.append({
            'marca': marca, 'cat_com': cat, 'n': n,
            'cob_prom':   round(sum(r['cobert']     for r in rows) / n, 2),
            'stock_unid': sum(r['stock_unid'] for r in rows),
            'stock_cst':  sum(r['stock_cst']  for r in rows),
            'venta_unid': sum(r['venta_unid'] for r in rows),
            'venta_cst':  sum(r['venta_cst']  for r in rows),
            'll_cur':     sum(r['llegadas'][_CUR_MES_KEY] for r in rows),
            'sin_emb':    sum(1 for r in rows if r['prox'] == 'Sin embarque'),
            'con_ll':     sum(1 for r in rows if r['prox'] != 'Sin embarque'),
            'dist': ' | '.join(
                f"{m}: {sum(1 for r in rows if r['prox']==m)}"
                for m in MESES_ORDEN + ['Sin embarque']
                if any(r['prox'] == m for r in rows)
            ),
            '_rows': rows,
        })
    return resumen, grupos

# ══════════════════════════════════════════════════════════════════════
# ESCRITURA DE HOJAS
# ══════════════════════════════════════════════════════════════════════
def write_resumen(owb, resumen, title, title_color, sheet_name, cob_fn,
                  col3_label, col3_val_fn, col3_alert_fn,
                  col11_label, col11_val_fn, col11_alert_fn):
    ws = owb.create_sheet(sheet_name)
    ws.freeze_panes = 'A4'
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 38

    ws.merge_cells('A1:L1')
    t = ws.cell(row=1, column=1, value=title)
    t.font = Font(bold=True, color='FFFFFF', name='Calibri', size=13)
    t.fill = PatternFill('solid', fgColor=title_color)
    t.alignment = Alignment(horizontal='center', vertical='center')

    for c1, c2, label, color in [
        (1,2,'',C_HDR),(3,5,'COBERTURA / STOCK','34495E'),
        (6,7,'UNIDADES','1A5276'),(8,9,'COSTO','6E2F1A'),(10,12,'TRANSITOS','145A32'),
    ]:
        if c1 != c2:
            ws.merge_cells(start_row=2, start_column=c1, end_row=2, end_column=c2)
        c = ws.cell(row=2, column=c1, value=label)
        c.font = Font(bold=True, color='FFFFFF', name='Calibri', size=9)
        c.fill = PatternFill('solid', fgColor=color)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = borde

    for col, (h, w) in enumerate([
        ('Marca',16),('Cat. Comercial',20),('SKUs',7),('Cob. Prom',10),(col3_label,8),
        ('Stock Hoy\nUnidades',13),(f'Venta PPTO\n{_CUR_MES_KEY} Unid',13),
        ('Stock Hoy\nCST',13),(f'Venta CST\n{_CUR_MES_KEY}',13),
        (f'Llegadas\n{_CUR_MES_KEY} Unid',12),(col11_label,10),('Detalle Llegadas',30),
    ], 1):
        hdr(ws, 3, col, h, w)

    cur_marca = None
    rn = 4
    for r in resumen:
        if r['marca'] != cur_marca:
            cur_marca = r['marca']
            ws.merge_cells(f'A{rn}:L{rn}')
            mc = ws.cell(row=rn, column=1, value=f'  {cur_marca}')
            mc.font = Font(bold=True, color=C_SUB_FG, name='Calibri', size=10)
            mc.fill = PatternFill('solid', fgColor=C_SUB)
            mc.alignment = Alignment(horizontal='left', vertical='center')
            mc.border = borde
            ws.row_dimensions[rn].height = 16
            rn += 1

        bg = C_ALT if rn % 2 == 0 else 'FFFFFF'
        cel(ws, rn, 1, r['marca'],   bg=bg)
        cel(ws, rn, 2, r['cat_com'], bg=bg)
        cel(ws, rn, 3, r['n'],       align='center', bold=True, bg=bg)
        cb, cf = cob_fn(r['cob_prom'])
        cel(ws, rn, 4, r['cob_prom'], align='center', fmt='0.00', bold=True, bg=cb, fg=cf)

        v5 = col3_val_fn(r)
        b5, f5, bo5 = col3_alert_fn(v5, bg)
        cel(ws, rn, 5, v5, align='center', bg=b5, fg=f5, bold=bo5)

        cel(ws, rn, 6, r['stock_unid'], align='right', fmt='#,##0', bg=bg)
        cel(ws, rn, 7, r['venta_unid'], align='right', fmt='#,##0', bg=bg)
        cel(ws, rn, 8, r['stock_cst'],  align='right', fmt='#,##0', bg=bg)
        cel(ws, rn, 9, r['venta_cst'],  align='right', fmt='#,##0', bg=bg)

        ll_bg = C_VERDE_BG if r['ll_cur'] > 0 else bg
        cel(ws, rn, 10, r['ll_cur'], align='right', fmt='#,##0', bg=ll_bg)

        v11 = col11_val_fn(r)
        b11, f11, bo11 = col11_alert_fn(v11, bg)
        cel(ws, rn, 11, v11, align='center', bg=b11, fg=f11, bold=bo11)
        cel(ws, rn, 12, r['dist'], bg=bg, size=8)
        ws.row_dimensions[rn].height = 15
        rn += 1

    ws.row_dimensions[rn].height = 18
    for col, (v, f) in enumerate(zip(
        ['TOTAL','', sum(r['n'] for r in resumen),'','',
         sum(r['stock_unid'] for r in resumen), sum(r['venta_unid'] for r in resumen),
         sum(r['stock_cst']  for r in resumen), sum(r['venta_cst']  for r in resumen),
         sum(r['ll_cur']     for r in resumen),'',''],
        [None,None,None,None,None,'#,##0','#,##0','#,##0','#,##0','#,##0',None,None]
    ), 1):
        cel(ws, rn, col, v,
            align='right' if isinstance(v,(int,float)) else 'center',
            fmt=f, bold=True, bg=C_HDR, fg='FFFFFF')
    ws.auto_filter.ref = f'A3:L{rn-1}'


def write_resumen_marca(owb, results, title, title_color, sheet_name,
                        cob_fn, col5_label, col5_val_fn, col5_alert_fn,
                        show_cat_padre=False):
    """Tabla resumen agrupada por Marca (opcionalmente también por Cat. Padre)."""
    # Agrupar por (marca, cat_padre) o solo marca
    por_marca = defaultdict(list)
    for r in results:
        key = (r['marca'], r['cat_padre']) if show_cat_padre else r['marca']
        por_marca[key].append(r)

    ws = owb.create_sheet(sheet_name)
    ws.freeze_panes = 'A3'
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 38

    ncols   = 11 if show_cat_padre else 10
    end_col = get_column_letter(ncols)

    # Título
    ws.merge_cells(f'A1:{end_col}1')
    t = ws.cell(row=1, column=1, value=title)
    t.font = Font(bold=True, color='FFFFFF', name='Calibri', size=13)
    t.fill = PatternFill('solid', fgColor=title_color)
    t.alignment = Alignment(horizontal='center', vertical='center')

    # Encabezados
    headers = [('Marca', 20)]
    if show_cat_padre:
        headers.append(('Cat. Padre', 22))
    headers += [
        ('SKUs', 7), ('Cob.\nProm', 10), (col5_label, 10),
        ('Stock Hoy\nUnidades', 14), (f'Venta PPTO\n{_CUR_MES_KEY} Unid', 14),
        ('Stock Hoy\nCST', 14), (f'Venta CST\n{_CUR_MES_KEY}', 14),
        (f'Llegadas\n{_CUR_MES_KEY} Unid', 13), ('Detalle Llegadas', 35),
    ]
    for col, (h, w) in enumerate(headers, 1):
        hdr(ws, 2, col, h, w)

    rn = 3
    for key in sorted(por_marca):
        rows      = por_marca[key]
        marca     = key[0] if show_cat_padre else key
        cat_padre = key[1] if show_cat_padre else None
        n          = len(rows)
        cob_prom   = round(sum(r['cobert']     for r in rows) / n, 2)
        stock_unid = sum(r['stock_unid'] for r in rows)
        stock_cst  = sum(r['stock_cst']  for r in rows)
        venta_unid = sum(r['venta_unid'] for r in rows)
        venta_cst  = sum(r['venta_cst']  for r in rows)
        ll_abr     = sum(r['llegadas'][_CUR_MES_KEY] for r in rows)
        dist_str   = ' | '.join(
            f"{m}: {sum(1 for r in rows if r['prox']==m)}"
            for m in MESES_ORDEN + ['Sin embarque']
            if any(r['prox'] == m for r in rows)
        )
        v5 = col5_val_fn(rows)

        bg = C_ALT if rn % 2 == 0 else 'FFFFFF'
        c = 1
        cel(ws, rn, c, marca, bold=True, bg=bg); c += 1
        if show_cat_padre:
            cel(ws, rn, c, cat_padre, bg=bg); c += 1
        cel(ws, rn, c, n, align='center', bold=True, bg=bg); c += 1

        cb, cf = cob_fn(cob_prom)
        cel(ws, rn, c, cob_prom, align='center', fmt='0.00', bold=True, bg=cb, fg=cf); c += 1

        b5, f5, bo5 = col5_alert_fn(v5, bg)
        cel(ws, rn, c, v5, align='center', bg=b5, fg=f5, bold=bo5); c += 1

        cel(ws, rn, c, stock_unid, align='right', fmt='#,##0', bg=bg); c += 1
        cel(ws, rn, c, venta_unid, align='right', fmt='#,##0', bg=bg); c += 1
        cel(ws, rn, c, stock_cst,  align='right', fmt='#,##0', bg=bg); c += 1
        cel(ws, rn, c, venta_cst,  align='right', fmt='#,##0', bg=bg); c += 1
        ll_bg = C_VERDE_BG if ll_abr > 0 else bg
        cel(ws, rn, c, ll_abr, align='right', fmt='#,##0', bg=ll_bg); c += 1
        cel(ws, rn, c, dist_str, bg=bg, size=8)
        ws.row_dimensions[rn].height = 15
        rn += 1

    # Totales
    ws.row_dimensions[rn].height = 18
    tot_n  = len(results)
    tot_su = sum(r['stock_unid'] for r in results)
    tot_vu = sum(r['venta_unid'] for r in results)
    tot_sc = sum(r['stock_cst']  for r in results)
    tot_vc = sum(r['venta_cst']  for r in results)
    tot_ll = sum(r['llegadas'][_CUR_MES_KEY] for r in results)
    totales     = ['TOTAL']
    totales_fmt = [None]
    if show_cat_padre:
        totales.append(''); totales_fmt.append(None)
    totales     += [tot_n, '', '', tot_su, tot_vu, tot_sc, tot_vc, tot_ll, '']
    totales_fmt += [None, None, None, '#,##0','#,##0','#,##0','#,##0','#,##0', None]
    for col, (v, f) in enumerate(zip(totales, totales_fmt), 1):
        cel(ws, rn, col, v,
            align='right' if isinstance(v,(int,float)) else 'center',
            fmt=f, bold=True, bg=C_HDR, fg='FFFFFF')
    ws.auto_filter.ref = f'A2:{end_col}{rn-1}'


def write_resumen_cat(owb, results, all_results, title, title_color, sheet_name, cat_field):
    """
    Resumen de coberturas inquietantes agrupado por Marca × categoría
    (cat_padre o cat_hijo).
    - results:     SKUs filtrados (<2 meses) — para conteos y métricas de urgencia
    - all_results: TODOS los SKUs — para calcular cob. prom. real por (marca, cat)
    """
    # ── Todos los SKUs agrupados por (marca, cat) — para métricas totales ──
    all_cat_map = defaultdict(list)
    for r in all_results:
        key = (r['marca'], r[cat_field] or '(sin categoria)')
        all_cat_map[key].append(r)

    # ── Agrupar inquietantes (<2m) por (marca, cat) ──────────────────────
    por_cat = defaultdict(list)
    for r in results:
        key = (r['marca'], r[cat_field] or '(sin categoria)')
        por_cat[key].append(r)

    filas = []
    for (marca, cat), inq_rows in sorted(por_cat.items()):
        # Todos los SKUs de esta (marca × cat) — para métricas de categoría completa
        all_rows_cat = all_cat_map.get((marca, cat), inq_rows)
        todos_cob = [r['cobert'] for r in all_rows_cat]
        # Promedio excluyendo sobrestock (>6m): evita distorsión bimodal
        no_sobre = [c for c in todos_cob if c <= 6]
        base_prom = no_sobre if no_sobre else todos_cob
        filas.append({
            'marca':      marca,
            'cat':        cat,
            # n y métricas financieras → CATEGORÍA COMPLETA
            'n':          len(all_rows_cat),
            'cob_prom':   round(sum(base_prom) / len(base_prom), 2),
            # Distribución de urgencia → solo los SKUs inquietantes (<2m)
            'n_rojo':     sum(1 for r in inq_rows if r['cobert'] < 0.5),
            'n_naran':    sum(1 for r in inq_rows if 0.5 <= r['cobert'] < 1.0),
            'n_amar':     sum(1 for r in inq_rows if 1.0 <= r['cobert'] < 2.0),
            # Stock/venta/llegadas → CATEGORÍA COMPLETA
            'stock_unid': sum(r['stock_unid'] for r in all_rows_cat),
            'venta_unid': sum(r['venta_unid'] for r in all_rows_cat),
            'stock_cst':  sum(r['stock_cst']  for r in all_rows_cat),
            'venta_cst':  sum(r['venta_cst']  for r in all_rows_cat),
            'll_abr':     sum(r['llegadas'][_CUR_MES_KEY] for r in all_rows_cat),
            'sin_emb':    sum(1 for r in all_rows_cat if r['prox'] == 'Sin embarque'),
            'con_ll':     sum(1 for r in all_rows_cat if r['prox'] != 'Sin embarque'),
            'dist': ' | '.join(
                f"{m}: {sum(1 for r in all_rows_cat if r['prox']==m)}"
                for m in MESES_ORDEN + ['Sin embarque']
                if any(r['prox'] == m for r in all_rows_cat)
            ),
        })
    # Solo mostrar filas cuyo promedio (marca × cat, excluyendo sobrestock) también es < 2
    filas = [f for f in filas if f['cob_prom'] < 2]
    # Ordenar por cobertura promedio ascendente (más urgente primero)
    filas.sort(key=lambda x: x['cob_prom'])

    # ── Hoja ─────────────────────────────────────────────────────────────
    ws = owb.create_sheet(sheet_name)
    ws.freeze_panes = 'A4'
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 38

    NCOLS = 15
    end_col = get_column_letter(NCOLS)
    ws.merge_cells(f'A1:{end_col}1')
    t = ws.cell(row=1, column=1, value=title)
    t.font = Font(bold=True, color='FFFFFF', name='Calibri', size=13)
    t.fill = PatternFill('solid', fgColor=title_color)
    t.alignment = Alignment(horizontal='center', vertical='center')

    # Sub-encabezados de grupo
    for c1, c2, label, color in [
        (1, 4, 'MARCA / CATEGORÍA', C_HDR),
        (5, 7, 'URGENCIA (SKUs)', '7B241C'),
        (8, 11, 'STOCK / VENTA', '1A5276'),
        (12, 15, 'TRÁNSITOS', '145A32'),
    ]:
        ws.merge_cells(start_row=2, start_column=c1, end_row=2, end_column=c2)
        c = ws.cell(row=2, column=c1, value=label)
        c.font = Font(bold=True, color='FFFFFF', name='Calibri', size=9)
        c.fill = PatternFill('solid', fgColor=color)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = borde

    cat_label = 'Cat. Padre' if cat_field == 'cat_padre' else 'Cat. Hijo'
    for col, (h, w) in enumerate([
        ('Marca', 15), (cat_label, 26), ('SKUs', 7), ('Cob.\nProm', 9),
        ('SKUs\n<0.5m', 9), ('SKUs\n0.5-1m', 9), ('SKUs\n1-2m', 9),
        ('Stock\nUnid', 13), (f'Venta\nUnid {_CUR_MES}', 13),
        ('Stock\nCST', 13), (f'Venta\nCST {_CUR_MES}', 13),
        (f'Llegadas\n{_CUR_MES} Unid', 12), ('Sin\nEmbarque', 10),
        ('Con\nLlegadas', 10), ('Detalle Llegadas', 32),
    ], 1):
        hdr(ws, 3, col, h, w)

    # ── Datos ─────────────────────────────────────────────────────────────
    for i, f in enumerate(filas, 4):
        bg = C_ALT if i % 2 == 0 else 'FFFFFF'
        cel(ws, i, 1, f['marca'], bg=bg, bold=True)
        cel(ws, i, 2, f['cat'],   bg=bg)
        cel(ws, i, 3, f['n'],     align='center', bold=True, bg=bg)

        cb, cf = cob_color_inq(f['cob_prom'])
        cel(ws, i, 4, f['cob_prom'], align='center', fmt='0.00', bold=True, bg=cb, fg=cf)

        cel(ws, i, 5, f['n_rojo'] or '', align='center', bold=bool(f['n_rojo']),
            bg=C_ROJO_BG  if f['n_rojo']  else bg, fg=C_ROJO  if f['n_rojo']  else '000000')
        cel(ws, i, 6, f['n_naran'] or '', align='center', bold=bool(f['n_naran']),
            bg=C_NARAN_BG if f['n_naran'] else bg, fg=C_NARAN if f['n_naran'] else '000000')
        cel(ws, i, 7, f['n_amar'] or '', align='center', bold=bool(f['n_amar']),
            bg=C_AMAR_BG  if f['n_amar']  else bg, fg=C_AMAR  if f['n_amar']  else '000000')

        cel(ws, i, 8,  f['stock_unid'], align='right', fmt='#,##0', bg=bg)
        cel(ws, i, 9,  f['venta_unid'], align='right', fmt='#,##0', bg=bg)
        cel(ws, i, 10, f['stock_cst'],  align='right', fmt='#,##0', bg=bg)
        cel(ws, i, 11, f['venta_cst'],  align='right', fmt='#,##0', bg=bg)

        ll_bg = C_VERDE_BG if f['ll_abr'] > 0 else bg
        cel(ws, i, 12, f['ll_abr'] or '', align='right', fmt='#,##0' if f['ll_abr'] else None, bg=ll_bg)
        cel(ws, i, 13, f['sin_emb'] or '', align='center', bold=bool(f['sin_emb']),
            bg=C_ROJO_BG if f['sin_emb'] else bg, fg=C_ROJO if f['sin_emb'] else '000000')
        cel(ws, i, 14, f['con_ll'] or '',  align='center', bold=bool(f['con_ll']),
            bg=C_VERDE_BG if f['con_ll'] else bg, fg=C_VERDE if f['con_ll'] else '000000')
        cel(ws, i, 15, f['dist'], bg=bg, size=8)
        ws.row_dimensions[i].height = 15

    # ── Fila total ────────────────────────────────────────────────────────
    rn = 4 + len(filas)
    ws.row_dimensions[rn].height = 18
    totales = ['TOTAL', '', sum(f['n'] for f in filas), '',
               sum(f['n_rojo'] for f in filas), sum(f['n_naran'] for f in filas),
               sum(f['n_amar'] for f in filas),
               sum(f['stock_unid'] for f in filas), sum(f['venta_unid'] for f in filas),
               sum(f['stock_cst']  for f in filas), sum(f['venta_cst']  for f in filas),
               sum(f['ll_abr']     for f in filas), sum(f['sin_emb'] for f in filas),
               sum(f['con_ll']     for f in filas), '']
    fmts = [None,None,None,None,None,None,None,
            '#,##0','#,##0','#,##0','#,##0','#,##0',None,None,None]
    for col, (v, fmt) in enumerate(zip(totales, fmts), 1):
        cel(ws, rn, col, v,
            align='right' if isinstance(v,(int,float)) else 'center',
            fmt=fmt, bold=True, bg=C_HDR, fg='FFFFFF')

    ws.auto_filter.ref = f'A3:{end_col}{rn-1}'


def write_detalle(owb, results, title, title_color, sheet_name,
                  cob_fn, prox_fn, sort_key_fn, show_carga=False):
    ws = owb.create_sheet(sheet_name)
    ws.freeze_panes = 'A3'
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 38

    ncols   = 22 if show_carga else 21
    end_col = get_column_letter(ncols)
    ws.merge_cells(f'A1:{end_col}1')
    t = ws.cell(row=1, column=1, value=title)
    t.font = Font(bold=True, color='FFFFFF', name='Calibri', size=12)
    t.fill = PatternFill('solid', fgColor=title_color)
    t.alignment = Alignment(horizontal='center', vertical='center')

    # 21 cols base + col 22 opcional: Fecha Máx. de Carga (solo Detalle Critico)
    headers = [
        ('Marca',15),('Cat. Comercial',18),('Cat. Padre',16),('Cat. Hijo',16),
        ('Ranking\nComercial',10),('SKU',16),('Descripcion',38),
        ('Cobert.\nACT 26',9),('Stock\nUnid',11),(f'Venta PPTO\n{_CUR_MES} Unid',12),
        ('Stock\nCST',11),(f'Venta\nCST {_CUR_MES}',12),
        ('Llegadas\nABR Unid',11),('Llegadas\nMAY Unid',11),
        ('Llegadas\nJUN Unid',11),('Llegadas\nJUL Unid',11),
        ('Llegadas\nAGO Unid',11),('Llegadas\nSEP Unid',11),
        ('Prox.\nLlegada',13),('PI\nEmbarque',15),('ETA\nBodega',13),
    ]
    if show_carga:
        headers.append(('Fecha Máx.\nde Carga', 14))
    for col, (h, w) in enumerate(headers, 1):
        hdr(ws, 2, col, h, w)

    for i, r in enumerate(sorted(results, key=sort_key_fn), 3):
        bg = C_ALT if i % 2 == 0 else 'FFFFFF'
        cel(ws, i,  1, r['marca'],      bg=bg)
        cel(ws, i,  2, r['cat_com'],    bg=bg)
        cel(ws, i,  3, r['cat_padre'],  bg=bg)
        cel(ws, i,  4, r['cat_hijo'],   bg=bg)
        rnk = r['ranking']
        cel(ws, i,  5, rnk, align='center', bg=bg)
        cel(ws, i,  6, r['sku'],        align='center', bg=bg)
        cel(ws, i,  7, r['desc'],       bg=bg)

        cb, cf = cob_fn(r['cobert'])
        cel(ws, i,  8, r['cobert'], align='center', fmt='0.000', bold=True, bg=cb, fg=cf)

        cel(ws, i,  9, r['stock_unid'], align='right', fmt='#,##0', bg=bg)
        cel(ws, i, 10, r['venta_unid'], align='right', fmt='#,##0', bg=bg)
        cel(ws, i, 11, r['stock_cst'],  align='right', fmt='#,##0', bg=bg)
        cel(ws, i, 12, r['venta_cst'],  align='right', fmt='#,##0', bg=bg)

        for j, mes in enumerate(MESES_ORDEN[:-1], 13):
            v = r['llegadas'][mes]
            cel(ws, i, j, v if v > 0 else '', align='right',
                fmt='#,##0' if v > 0 else None,
                bg=C_VERDE_BG if v > 0 else bg,
                fg=C_VERDE    if v > 0 else 'AAAAAA')

        pb, pf = prox_fn(r['prox'])
        cel(ws, i, 19, r['prox'], align='center', bold=True, bg=pb, fg=pf)

        if r['prox'] != 'Sin embarque':
            cel(ws, i, 20, r['pi'], align='center', bg=pb, fg=pf, bold=bool(r['pi']))
            eta = r['eta']
            eta_str = eta.strftime('%d/%m/%Y') if isinstance(eta,(_dt.datetime,_dt.date)) else (str(eta) if eta else '')
            cel(ws, i, 21, eta_str, align='center', bg=pb, fg=pf, bold=bool(eta_str))
        else:
            cel(ws, i, 20, '', bg=pb, fg=pf)
            cel(ws, i, 21, '', bg=pb, fg=pf)

        if show_carga:
            # Solo se calcula si el SKU NO tiene OC registrada (sin embarque = necesita orden nueva)
            if not r.get('has_emb', False):
                fmc = fecha_max_carga(r['prox'], r.get('puerto', ''))
            else:
                fmc = ''
            if fmc:
                today = _dt.date.today()
                fmc_date = _dt.datetime.strptime(fmc, '%d/%m/%Y').date()
                if fmc_date < today:
                    fmc_bg, fmc_fg = C_ROJO_BG, C_ROJO
                elif (fmc_date - today).days <= 7:
                    fmc_bg, fmc_fg = C_NARAN_BG, C_NARAN
                else:
                    fmc_bg, fmc_fg = C_VERDE_BG, C_VERDE
                cel(ws, i, 22, fmc, align='center', bold=True, bg=fmc_bg, fg=fmc_fg)
            else:
                cel(ws, i, 22, '', bg=bg)

        ws.row_dimensions[i].height = 15
    ws.auto_filter.ref = f'A2:{end_col}{2+len(results)}'


# ══════════════════════════════════════════════════════════════════════
# HOJAS DE TRÁNSITO (recreación de hojas moradas)
# ══════════════════════════════════════════════════════════════════════
def write_transitos_embarque(owb, trans_rows, all_rows=None):
    """
    'Tránsitos por Embarque' — UNIFICADA con Riesgo por Embarque.
    Resumen por PI con nivel de riesgo, críticos e inquietantes.
    Filas SKU desplegables con cobertura real desde all_rows.
    """
    from openpyxl.worksheet.properties import WorksheetProperties, Outline

    # cob_map_full: dict sku → ('cobert', valor) | ('sin_venta', None) | None si no existe
    cob_map_full = all_rows or {}   # se recibe ya construido desde run_analisis
    desc_por_sku = {}

    ws = owb.create_sheet('Tránsitos por Embarque')
    ws.freeze_panes = 'A3'
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 38
    ws.sheet_properties = WorksheetProperties(
        outlinePr=Outline(summaryBelow=False, summaryRight=False)
    )

    NCOLS   = 11
    end_col = get_column_letter(NCOLS)

    ws.merge_cells(f'A1:{end_col}1')
    t = ws.cell(row=1, column=1,
                value=f'TRÁNSITOS POR EMBARQUE  —  [+] para ver SKUs  |  '
                      f'Cobertura: 🔴<1m  🟠1-2m  🟢2-4m  🔵4-6m  🟣>6m  |  {_CUR_MES_YEAR}')
    t.font      = Font(bold=True, color='FFFFFF', name='Calibri', size=12)
    t.fill      = PatternFill('solid', fgColor=C_AZUL)
    t.alignment = Alignment(horizontal='center', vertical='center')

    for col, (h, w) in enumerate([
        ('PI / Embarque',          22),
        ('ETA Chile  |  Descripción SKU', 38),
        ('ETA\nBodega',            14),
        ('Mes\nLlegada',           12),
        ('Marcas',                 20),
        ('SKUs\nDistintos',        11),
        ('Críticos\n<1m',          11),
        ('Inquiet.\n1-2m',         11),
        ('Unidades',               14),
        ('Valor USD',              14),
        ('Nivel\nRiesgo',          13),
    ], 1):
        hdr(ws, 2, col, h, w)

    # Agrupar por PI
    por_pi = defaultdict(list)
    for r in trans_rows:
        if r['pi']:
            sk_clean = _fmt_sku(r['sku'])
            por_pi[r['pi']].append({**r, 'sku': sk_clean})

    def _earliest_bod(rows):
        dates = [r['eta_bod'] for r in rows
                 if isinstance(r['eta_bod'], (_dt.datetime, _dt.date))]
        return min(dates) if dates else _dt.date(2099, 1, 1)

    def _fmt_date(val):
        if isinstance(val, _dt.datetime): return val.strftime('%d/%m/%Y')
        if isinstance(val, _dt.date):     return val.strftime('%d/%m/%Y')
        return str(val) if val else ''

    mm = {4:'ABR26',5:'MAY26',6:'JUN26',7:'JUL26',8:'AGO26',9:'SEP26',10:'OCT26'}
    MES_COLOR = {
        'ABR26': (C_VERDE_BG, C_VERDE),  'MAY26': (C_VERDE_BG, C_VERDE),
        'JUN26': (C_NARAN_BG, C_NARAN),  'JUL26': (C_NARAN_BG, C_NARAN),
        'AGO26': (C_ROJO_BG,  C_ROJO),   'SEP26': (C_ROJO_BG,  C_ROJO),
        'OCT26': (C_MORA_BG,  C_MORA),
    }

    def _cob_style_det(sk, mes_num=None):
        """Color para cobertura en fila de detalle SKU.
        mes_num: mes de llegada del embarque (4=ABR … 8=AGO); None → usa cobertura actual."""
        entry = cob_map_full.get(sk)
        if entry is None: return ('F4F6F7', '95A5A6', '—')
        if entry['tipo'] == 'sin_venta': return ('EBF5FB', '1A5276', 'Sin venta')
        # Prioridad: cobertura proyectada al mes de llegada; fallback a actual
        cob = entry.get(mes_num, entry['actual']) if mes_num else entry['actual']
        if cob == 0: return (C_ROJO_BG,  C_ROJO,  'Sin stock')   # stock=0, más crítico que <1m
        if cob < 1:  return (C_ROJO_BG,  C_ROJO,  f'{cob:.2f}')
        if cob < 2:   return (C_NARAN_BG, C_NARAN, f'{cob:.2f}')
        if cob < 4:   return (C_VERDE_BG, C_VERDE, f'{cob:.2f}')
        if cob < 6:   return (C_AZUL_BG,  C_AZUL,  f'{cob:.2f}')
        return                ('E8DAEF',   '6C3483', f'{cob:.2f}')

    def _riesgo(n_crit, n_inq):
        if n_crit >= 5: return ('FADBD8', 'CB4335', 'ALTO')
        if n_crit >= 2: return ('FDEBD0', 'E67E22', 'MEDIO')
        if n_crit >= 1: return ('FEF9E7', 'B7950B', 'BAJO')
        if n_inq  >= 3: return ('FEF9E7', 'B7950B', 'VIGILAR')
        return                 ('D5F5E3', '1E8449', 'OK')

    DET_BG = 'F4F6F7'
    DET_FG = '4D5656'

    sorted_pis = sorted(por_pi.items(), key=lambda kv: _earliest_bod(kv[1]))

    rn = 3
    for pi, rows in sorted_pis:
        bg = C_ALT if rn % 2 == 0 else 'FFFFFF'

        eta_chl_vals = [r['eta_chl'] for r in rows
                        if isinstance(r['eta_chl'], (_dt.datetime, _dt.date))]
        eta_chl = eta_chl_vals[0] if eta_chl_vals else None
        eta_bod = _earliest_bod(rows)
        eta_bod = None if eta_bod == _dt.date(2099, 1, 1) else eta_bod

        mes_counts     = defaultdict(int)
        mes_num_counts = defaultdict(int)
        for r in rows:
            try:
                mn_num = int(r['mes'])
                mn     = mm.get(mn_num)
                if mn:
                    mes_counts[mn]     += 1
                    mes_num_counts[mn_num] += 1
            except: pass
        mes_name   = max(mes_counts,     key=mes_counts.get)     if mes_counts     else ''
        mes_num_pi = max(mes_num_counts, key=mes_num_counts.get) if mes_num_counts else None
        mc_bg, mc_fg = MES_COLOR.get(mes_name, (bg, '000000'))

        # Agrupar SKUs dentro del PI
        sku_det = defaultdict(lambda: {'desc':'','eta_bod':None,'mes':None,
                                       'cantidad':0,'valor_usd':0,'tipo_cat':''})
        marca_set = set()
        for r in rows:
            sk = r['sku']
            if ' ' in sk or '\n' in sk or len(sk) > 25: continue
            d = sku_det[sk]
            if not d['desc']:    d['desc']    = desc_por_sku.get(sk, r['desc'])
            if not d['eta_bod']: d['eta_bod'] = r['eta_bod']
            if not d['mes']:     d['mes']     = r['mes']
            if not d['tipo_cat']: d['tipo_cat'] = r.get('tipo_cat','')
            d['cantidad']  += r['cantidad']
            d['valor_usd'] += r['valor_usd']
            marca_set.add(r['marca'])

        skus_validos = list(sku_det.keys())
        # Cobertura del mes ANTERIOR al de llegada — cuánto stock queda justo antes de que llegue
        # Ej: llega JUL → mostrar cobertura JUN (mes en que se agotaría si no llegara)
        mes_cob_pi = (mes_num_pi - 1) if mes_num_pi else None
        def _cob_val(s, mn=None):
            e = cob_map_full.get(s)
            if e is None or e['tipo'] == 'sin_venta': return None
            return e.get(mn, e['actual']) if mn else e['actual']
        n_crit = sum(1 for s in skus_validos
                     if _cob_val(s, mes_cob_pi) is not None and _cob_val(s, mes_cob_pi) < 1)
        n_inq  = sum(1 for s in skus_validos
                     if _cob_val(s, mes_cob_pi) is not None and 1 <= _cob_val(s, mes_cob_pi) < 2)
        n_sku   = len(skus_validos)
        unid    = int(sum(sku_det[s]['cantidad']  for s in skus_validos))
        usd_tot = sum(sku_det[s]['valor_usd'] for s in skus_validos)
        marcas  = ', '.join(sorted(marca_set))
        rb, rf, rlabel = _riesgo(n_crit, n_inq)

        # ── Fila resumen PI (nivel 0) ─────────────────────────────────
        cel(ws, rn, 1,  pi,                  bold=True, bg=bg)
        cel(ws, rn, 2,  _fmt_date(eta_chl),  align='center', bg=bg)
        cel(ws, rn, 3,  _fmt_date(eta_bod),  align='center', bg=mc_bg, fg=mc_fg, bold=True)
        cel(ws, rn, 4,  mes_name,            align='center', bold=True, bg=mc_bg, fg=mc_fg)
        cel(ws, rn, 5,  marcas,              bg=bg, fg='566573')
        cel(ws, rn, 6,  n_sku,               align='center', bg=bg)
        crit_bg = C_ROJO_BG if n_crit > 0 else bg
        crit_fg = C_ROJO    if n_crit > 0 else '000000'
        cel(ws, rn, 7,  n_crit,  align='center', bg=crit_bg, fg=crit_fg, bold=n_crit > 0)
        inq_bg = C_NARAN_BG if n_inq > 0 else bg
        inq_fg = C_NARAN    if n_inq > 0 else '000000'
        cel(ws, rn, 8,  n_inq,   align='center', bg=inq_bg, fg=inq_fg)
        cel(ws, rn, 9,  unid,    align='right', fmt='#,##0', bg=bg)
        cel(ws, rn, 10, usd_tot, align='right', fmt='$#,##0', bg=bg)
        cel(ws, rn, 11, rlabel,  align='center', bg=rb, fg=rf, bold=True)
        ws.row_dimensions[rn].height = 16
        rn += 1

        # ── Filas detalle SKU (nivel 1, colapsadas) ───────────────────
        # Ordenar por cobertura del mes anterior al de llegada
        def _sku_sort(sk):
            cv = _cob_val(sk, mes_cob_pi)
            if cv is None: return 3
            if cv < 1:     return 0
            if cv < 2:     return 1
            return 2
        for sk in sorted(skus_validos, key=_sku_sort):
            d = sku_det[sk]
            # Mes anterior al de llegada del SKU (o del PI como fallback)
            try:    det_mes_num = int(d['mes']) if d['mes'] else mes_num_pi
            except: det_mes_num = mes_num_pi
            det_cob_mes = (det_mes_num - 1) if det_mes_num else None
            cb, cf, cob_txt = _cob_style_det(sk, det_cob_mes)
            is_crit = (cb == C_ROJO_BG)

            det_mes = mm.get(det_mes_num, '')
            det_mc_bg, det_mc_fg = MES_COLOR.get(det_mes, (DET_BG, DET_FG))

            c = ws.cell(row=rn, column=1, value=f'  ↳  {sk}')
            c.font      = Font(name='Calibri', size=8, italic=True,
                               bold=is_crit, color=C_AZUL)
            c.fill      = PatternFill('solid', fgColor=cb if is_crit else DET_BG)
            c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
            c.border    = borde

            c2 = ws.cell(row=rn, column=2, value=d['desc'])
            c2.font      = Font(name='Calibri', size=8, color=DET_FG)
            c2.fill      = PatternFill('solid', fgColor=DET_BG)
            c2.alignment = Alignment(horizontal='left', vertical='center')
            c2.border    = borde

            cel(ws, rn, 3,  _fmt_date(d['eta_bod']), align='center',
                bg=det_mc_bg, fg=det_mc_fg, size=8)
            cel(ws, rn, 4,  det_mes, align='center',
                bg=det_mc_bg, fg=det_mc_fg, size=8)
            cel(ws, rn, 5,  '', bg=DET_BG)
            cel(ws, rn, 6,  '', bg=DET_BG)
            # Cobertura real
            cel(ws, rn, 7,  cob_txt, align='center', bg=cb, fg=cf, size=8,
                bold=is_crit)
            cel(ws, rn, 8,  '', bg=DET_BG)
            cel(ws, rn, 9,  int(d['cantidad']),  align='right', fmt='#,##0',
                bg=DET_BG, fg=DET_FG, size=8)
            cel(ws, rn, 10, d['valor_usd'], align='right', fmt='$#,##0',
                bg=DET_BG, fg=DET_FG, size=8)
            cel(ws, rn, 11, '', bg=DET_BG)

            ws.row_dimensions[rn].outline_level = 1
            ws.row_dimensions[rn].hidden        = True
            ws.row_dimensions[rn].height        = 14
            rn += 1

    # Fila total
    ws.row_dimensions[rn].height = 18
    tot_pi   = len(por_pi)
    tot_unid = int(sum(r['cantidad']  for r in trans_rows))
    tot_usd  = sum(r['valor_usd'] for r in trans_rows)
    for col, (v, f) in enumerate(zip(
        ['TOTAL','','','',f'{tot_pi} embarques','','','', tot_unid, tot_usd,''],
        [None,None,None,None,None,None,None,None,'#,##0','$#,##0',None]
    ), 1):
        cel(ws, rn, col, v,
            align='right' if isinstance(v,(int,float)) else 'center',
            fmt=f, bold=True, bg=C_HDR, fg='FFFFFF')
    ws.auto_filter.ref = f'A2:{end_col}{rn-1}'


def write_nuevos_transito(owb, trans_rows):
    """
    'Nuevos en Tránsito': SKUs con Tipo Categoria = NUEVO, agrupados por mes de llegada.
    Columnas: SKU | Descripción | Marca | Mes Llegada | ETA Bodega | Cantidad
    """
    ws = owb.create_sheet('Nuevos en Tránsito')
    ws.freeze_panes = 'A3'
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 38

    NCOLS = 6
    end_col = get_column_letter(NCOLS)

    ws.merge_cells(f'A1:{end_col}1')
    t = ws.cell(row=1, column=1,
                value=f'NUEVOS EN TRÁNSITO — Tipo Categoría: NUEVO | {_CUR_MES_YEAR}')
    t.font      = Font(bold=True, color='FFFFFF', name='Calibri', size=13)
    t.fill      = PatternFill('solid', fgColor=C_TEAL)
    t.alignment = Alignment(horizontal='center', vertical='center')

    for col, (h, w) in enumerate([
        ('SKU', 14), ('Descripción', 44), ('Marca', 18),
        ('Mes\nLlegada', 11), ('Fecha ETA\nBodega', 14), ('Cantidad', 12),
    ], 1):
        hdr(ws, 2, col, h, w)

    # Filtrar NUEVO
    nuevos = [r for r in trans_rows if 'nuevo' in r['tipo_cat'].lower()]

    mm = {4:'ABR26',5:'MAY26',6:'JUN26',7:'JUL26',8:'AGO26',9:'SEP26',10:'OCT26'}

    def _sort_key(r):
        try:   m = int(r['mes']) if r['mes'] is not None else 99
        except: m = 99
        d = r['eta_bod']
        if isinstance(d, _dt.datetime): d = d.date()
        elif not isinstance(d, _dt.date): d = _dt.date(2099, 1, 1)
        return (m, d, str(r['sku']))

    nuevos.sort(key=_sort_key)

    def _fmt_date(val):
        if isinstance(val, _dt.datetime): return val.strftime('%d/%m/%Y')
        if isinstance(val, _dt.date):     return val.strftime('%d/%m/%Y')
        return str(val) if val else ''

    cur_mes = None
    rn = 3
    for r in nuevos:
        try:   mes_name = mm.get(int(r['mes']), str(r['mes']))
        except: mes_name = str(r['mes']) if r['mes'] else ''

        # Separador de mes
        if mes_name != cur_mes:
            cur_mes = mes_name
            ws.merge_cells(f'A{rn}:{end_col}{rn}')
            mc = ws.cell(row=rn, column=1, value=f'  ▶  Llegada {mes_name}')
            mc.font      = Font(bold=True, color='FFFFFF', name='Calibri', size=10)
            mc.fill      = PatternFill('solid', fgColor='196F3D')
            mc.alignment = Alignment(horizontal='left', vertical='center')
            mc.border    = borde
            ws.row_dimensions[rn].height = 17
            rn += 1

        bg = C_ALT if rn % 2 == 0 else 'FFFFFF'
        cel(ws, rn, 1, r['sku'],          align='center', bg=bg)
        cel(ws, rn, 2, r['desc'],         bg=bg)
        cel(ws, rn, 3, r['marca'],        bg=bg)
        cel(ws, rn, 4, mes_name,          align='center', bold=True,
            bg=C_VERDE_BG, fg=C_VERDE)
        cel(ws, rn, 5, _fmt_date(r['eta_bod']), align='center', bg=bg)
        cel(ws, rn, 6, int(r['cantidad']), align='right', fmt='#,##0',
            bg=C_VERDE_BG, fg=C_VERDE, bold=True)
        ws.row_dimensions[rn].height = 15
        rn += 1

    # Fila total
    ws.row_dimensions[rn].height = 18
    tot_n    = len(nuevos)
    tot_unid = int(sum(r['cantidad'] for r in nuevos))
    for col, (v, f) in enumerate(zip(
        ['TOTAL','', f'{tot_n} items','','', tot_unid],
        [None,None,None,None,None,'#,##0']
    ), 1):
        cel(ws, rn, col, v,
            align='right' if isinstance(v, (int, float)) else 'center',
            fmt=f, bold=True, bg=C_HDR, fg='FFFFFF')
    ws.auto_filter.ref = f'A2:{end_col}{rn-1}'


def write_stock_transitos(owb, trans_rows):
    """
    'Stock + Tránsitos': por SKU — stock actual y llegadas programadas por mes.
    Columnas: SKU | Descripción | Marca | Stock Actual | MAY26 | JUN26 | ... | OCT26 | Total
    """
    ws = owb.create_sheet('Stock + Tránsitos')
    ws.freeze_panes = 'E3'
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 38

    MESES_TRANS = ['MAY26','JUN26','JUL26','AGO26','SEP26','OCT26']
    MESES_LABEL = ['MAY 26','JUN 26','JUL 26','AGO 26','SEP 26','OCT 26']
    mm = {4:'ABR26',5:'MAY26',6:'JUN26',7:'JUL26',8:'AGO26',9:'SEP26',10:'OCT26'}

    NCOLS = 4 + len(MESES_TRANS) + 1   # 4 fijos + 6 meses + 1 total
    end_col = get_column_letter(NCOLS)

    ws.merge_cells(f'A1:{end_col}1')
    t = ws.cell(row=1, column=1,
                value=f'STOCK + TRÁNSITOS — Unidades en tránsito por mes | {_CUR_MES_YEAR}')
    t.font      = Font(bold=True, color='FFFFFF', name='Calibri', size=13)
    t.fill      = PatternFill('solid', fgColor='2C3E50')
    t.alignment = Alignment(horizontal='center', vertical='center')

    # Sub-encabezados de sección
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)
    sc1 = ws.cell(row=2, column=1, value='PRODUCTO')
    sc1.font      = Font(bold=True, color='FFFFFF', name='Calibri', size=9)
    sc1.fill      = PatternFill('solid', fgColor='2C3E50')
    sc1.alignment = Alignment(horizontal='center', vertical='center')
    sc1.border    = borde

    ws.merge_cells(start_row=2, start_column=5, end_row=2, end_column=10)
    sc2 = ws.cell(row=2, column=5, value='TRÁNSITOS POR MES (unidades)')
    sc2.font      = Font(bold=True, color='FFFFFF', name='Calibri', size=9)
    sc2.fill      = PatternFill('solid', fgColor='145A32')
    sc2.alignment = Alignment(horizontal='center', vertical='center')
    sc2.border    = borde

    ws.cell(row=2, column=11, value='').fill  = PatternFill('solid', fgColor='2C3E50')
    ws.cell(row=2, column=11).border = borde

    # Fila de encabezados de columna
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 38

    for col, (h, w) in enumerate([
        ('SKU', 16), ('Descripción', 38), ('Marca', 18), ('Stock\nActual', 12),
        *[(m, 12) for m in MESES_LABEL],
        ('Total\nTránsito', 13),
    ], 1):
        hdr(ws, 3, col, h, w)

    # Agregar dimensión extra de fila para los sub-encabezados
    ws.freeze_panes = 'E4'

    # Construir por SKU
    por_sku = {}
    for r in trans_rows:
        sku = r['sku']
        if sku not in por_sku:
            por_sku[sku] = {
                'desc':     r['desc'],
                'marca':    r['marca'],
                'stk_act':  r['stk_act'],
                'llegadas': defaultdict(float),
            }
        else:
            # Actualizar stock si viene mayor (distintas filas del mismo SKU)
            if r['stk_act'] > por_sku[sku]['stk_act']:
                por_sku[sku]['stk_act'] = r['stk_act']
        try:
            mn = mm.get(int(r['mes']))
            if mn and mn in MESES_TRANS:
                por_sku[sku]['llegadas'][mn] += r['cantidad']
        except: pass

    # Filtrar SKUs inválidos (muestras, notas largas, etc.) — SKU real: ≤25 chars, sin espacios
    por_sku = {k: v for k, v in por_sku.items()
               if len(k) <= 25 and ' ' not in k and '\n' not in k}

    # Ordenar por marca → SKU
    sorted_skus = sorted(por_sku.items(), key=lambda kv: (kv[1]['marca'], kv[0]))

    rn = 4
    for sku, d in sorted_skus:
        bg = C_ALT if rn % 2 == 0 else 'FFFFFF'
        stk = int(d['stk_act'])

        if stk == 0:    stk_bg, stk_fg = C_ROJO_BG,  C_ROJO
        elif stk < 50:  stk_bg, stk_fg = C_NARAN_BG, C_NARAN
        else:           stk_bg, stk_fg = bg,          '000000'

        total_trans = int(sum(d['llegadas'].get(m, 0) for m in MESES_TRANS))

        cel(ws, rn, 1, sku,        align='center', bg=bg)
        cel(ws, rn, 2, d['desc'],  bg=bg)
        cel(ws, rn, 3, d['marca'], bg=bg)
        cel(ws, rn, 4, stk,         align='right', fmt='#,##0',
            bg=stk_bg, fg=stk_fg, bold=(stk == 0))

        for j, mes in enumerate(MESES_TRANS, 5):
            v = int(d['llegadas'].get(mes, 0))
            if v > 0:
                cel(ws, rn, j, v, align='right', fmt='#,##0',
                    bg=C_VERDE_BG, fg=C_VERDE, bold=True)
            else:
                cel(ws, rn, j, '', bg=bg)

        if total_trans > 0:
            cel(ws, rn, 11, total_trans, align='right', fmt='#,##0',
                bg=C_AZUL_BG, fg=C_AZUL, bold=True)
        else:
            cel(ws, rn, 11, '', bg=bg)

        ws.row_dimensions[rn].height = 15
        rn += 1

    # Fila total
    ws.row_dimensions[rn].height = 18
    tot_stk    = int(sum(d['stk_act'] for d in por_sku.values()))
    tots_mes   = {m: int(sum(d['llegadas'].get(m, 0) for d in por_sku.values()))
                  for m in MESES_TRANS}
    tot_trans  = sum(tots_mes.values())
    totales    = ['TOTAL', f'{len(por_sku)} SKUs', '', tot_stk,
                  *[tots_mes[m] for m in MESES_TRANS], tot_trans]
    fmts_tot   = [None, None, None, '#,##0',
                  *['#,##0']*len(MESES_TRANS), '#,##0']
    for col, (v, f) in enumerate(zip(totales, fmts_tot), 1):
        cel(ws, rn, col, v,
            align='right' if isinstance(v, (int, float)) else 'center',
            fmt=f, bold=True, bg=C_HDR, fg='FFFFFF')
    ws.auto_filter.ref = f'A3:{end_col}{rn-1}'


# ══════════════════════════════════════════════════════════════════════
# HOJA VTA X MARCA — mes en curso
# ══════════════════════════════════════════════════════════════════════
_MES_ABBR = ['ENE','FEB','MAR','ABR','MAY','JUN',
             'JUL','AGO','SEP','OCT','NOV','DIC']
_MES_NOMBRE = ['Enero','Febrero','Marzo','Abril','Mayo','Junio',
               'Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre']

def write_vta_marca_mes(owb, data_vta):
    """
    Hoja 'VTA x Marca MAY 26':
    Solo el mes en curso (Mayo) con linealidad.
    6 columnas: Marca | Meta MAY | Meta al día X | Venta Acum. MAY | vs Lineal $ | % vs Lineal
    """
    today      = _dt.date.today()
    año        = today.year

    # MAY — en curso, con linealidad
    ayer        = today - _dt.timedelta(days=1)
    dia_ayer    = ayer.day
    total_dias  = calendar.monthrange(año, 5)[1]   # mayo
    linealidad  = dia_ayer / total_dias

    sheet_name = f'VTA x Marca MAY 26'

    if not data_vta or not isinstance(data_vta[0], list):
        return

    VTA_SKIP = {'total general', 'total empresa'}

    # Encontrar última fila de encabezados con "VENTA ACUM."
    last_hdr = None
    for i, row in enumerate(data_vta):
        if any(cell and 'venta acum' in str(cell).lower() for cell in (row or [])):
            last_hdr = i
    if last_hdr is None:
        return

    hdr_row = data_vta[last_hdr]

    # Detectar columnas MAY
    col_ppto_may = col_acum_may = None
    for j, cell in enumerate(hdr_row):
        if not cell: continue
        s = str(cell).upper()
        if 'MAY' in s:
            if 'VENTA ACUM' in s and col_acum_may is None: col_acum_may = j
            elif 'PPTO' in s and col_ppto_may is None:     col_ppto_may = j

    if col_acum_may is None:
        return

    def _g(row, col):
        if col is None or col >= len(row): return 0.0
        return float(row[col] or 0)

    # Leer filas de marcas
    vta_rows = []
    for i in range(last_hdr + 1, len(data_vta)):
        row = data_vta[i]
        if not row: continue
        marca = row[1] if len(row) > 1 else None
        if not marca: continue
        marca_s = str(marca).strip()
        if not marca_s or marca_s.lower() in VTA_SKIP: continue

        ppto_may   = _g(row, col_ppto_may)
        acum_may   = _g(row, col_acum_may)
        meta_lin   = ppto_may * linealidad
        vs_lin     = acum_may - meta_lin
        pct_lin    = acum_may / meta_lin if meta_lin else 0

        vta_rows.append({
            'marca':    marca_s,
            'ppto_may': ppto_may, 'meta_lin': meta_lin,
            'acum_may': acum_may, 'vs_lin':   vs_lin,
            'pct_lin':  pct_lin,
        })

    if not vta_rows:
        return

    pnac_rows  = [r for r in vta_rows if r['marca'].lower() == 'p. nacionales']
    marca_rows = [r for r in vta_rows if r['marca'].lower() != 'p. nacionales']
    marca_rows.sort(key=lambda r: -r['pct_lin'])   # ordenar por % vs Lineal desc

    def _lin_style(pct):
        if pct >= 1.10: return ('D2B4DE', '6C3483')
        if pct >= 0.90: return (C_VERDE_BG,  C_VERDE)
        if pct >= 0.70: return (C_AMAR_BG,   C_AMAR)
        if pct >= 0.50: return (C_NARAN_BG,  C_NARAN)
        return               (C_ROJO_BG,   C_ROJO)

    # ── Crear hoja ────────────────────────────────────────────────────
    ws = owb.create_sheet(sheet_name)
    ws.freeze_panes = 'A4'
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 40

    NCOLS   = 6
    end_col = get_column_letter(NCOLS)

    # Fila 1: título
    ws.merge_cells(f'A1:{end_col}1')
    t = ws.cell(row=1, column=1,
                value=f'VENTA x MARCA — MAY {año} en curso: '
                      f'día {dia_ayer} de {total_dias} = {linealidad:.1%} linealidad')
    t.font      = Font(bold=True, color='FFFFFF', name='Calibri', size=13)
    t.fill      = PatternFill('solid', fgColor='1B2631')
    t.alignment = Alignment(horizontal='center', vertical='center')

    # Fila 2: encabezado de bloque MAY
    ws.merge_cells(f'A2:{end_col}2')
    b2 = ws.cell(row=2, column=1,
                 value=f'MAYO {año} — EN CURSO (día {dia_ayer}/{total_dias})')
    b2.font      = Font(bold=True, color='FFFFFF', name='Calibri', size=10)
    b2.fill      = PatternFill('solid', fgColor='145A32')
    b2.alignment = Alignment(horizontal='center', vertical='center')

    # Fila 3: encabezados de columna
    for col, (h, w) in enumerate([
        ('Marca',                              22),
        (f'Meta MAY\n{año}',                  16),
        (f'Meta al día {dia_ayer}\n({linealidad:.1%})', 16),
        (f'Venta Acum.\nMAY {año}',           16),
        ('vs Lineal\n($)',                     15),
        ('% vs Lineal\nMAY',                  13),
    ], 1):
        hdr(ws, 3, col, h, w)

    rn = 4
    for r in marca_rows:
        bg = C_ALT if rn % 2 == 0 else 'FFFFFF'
        mb, mf = _lin_style(r['pct_lin'])
        vl_bg = C_VERDE_BG if r['vs_lin'] >= 0 else C_ROJO_BG
        vl_fg = C_VERDE    if r['vs_lin'] >= 0 else C_ROJO

        cel(ws, rn, 1, r['marca'],    bold=True, bg=bg)
        cel(ws, rn, 2, r['ppto_may'], align='right', fmt='$ #,##0', bg=bg)
        cel(ws, rn, 3, r['meta_lin'], align='right', fmt='$ #,##0', bg='F4F6F7', fg='566573')
        cel(ws, rn, 4, r['acum_may'], align='right', fmt='$ #,##0', bg=mb, fg=mf, bold=True)
        cel(ws, rn, 5, r['vs_lin'],   align='right', fmt='$ #,##0;[Red]$ -#,##0', bg=vl_bg, fg=vl_fg, bold=True)
        cel(ws, rn, 6, r['pct_lin'],  align='center', fmt='0.0%', bg=mb, fg=mf, bold=True)

        ws.row_dimensions[rn].height = 17
        rn += 1

    # Separador
    ws.row_dimensions[rn].height = 4
    for c in range(1, NCOLS + 1):
        ws.cell(row=rn, column=c).fill = PatternFill('solid', fgColor='1B2631')
    rn += 1

    # TOTAL PROPIA
    def _s(rows, k): return sum(r[k] for r in rows)
    tp_ppto_may = _s(marca_rows, 'ppto_may'); tp_acum_may = _s(marca_rows, 'acum_may')
    tp_meta_lin = tp_ppto_may * linealidad
    tp_vs_lin   = tp_acum_may - tp_meta_lin
    tp_pct_lin  = tp_acum_may / tp_meta_lin if tp_meta_lin else 0

    ws.row_dimensions[rn].height = 18
    for col, (v, f) in enumerate(zip(
        ['TOTAL PROPIA', tp_ppto_may, tp_meta_lin, tp_acum_may, tp_vs_lin, tp_pct_lin],
        [None, '$ #,##0', '$ #,##0', '$ #,##0', '$ #,##0;[Red]$ -#,##0', '0.0%']
    ), 1):
        c = ws.cell(row=rn, column=col, value=v)
        c.font      = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
        c.fill      = PatternFill('solid', fgColor='1B2631')
        c.alignment = Alignment(horizontal='right' if isinstance(v,(int,float)) else 'center', vertical='center')
        c.border    = borde
        if f: c.number_format = f
    rn += 1

    # P. Nacionales
    for pn in pnac_rows:
        pn_ml      = pn['ppto_may'] * linealidad
        pn_vs      = pn['acum_may'] - pn_ml
        pn_pct_lin = pn['acum_may'] / pn_ml if pn_ml else 0
        mb, mf = _lin_style(pn_pct_lin)
        vl_bg = C_VERDE_BG if pn_vs >= 0 else C_ROJO_BG
        vl_fg = C_VERDE    if pn_vs >= 0 else C_ROJO

        ws.row_dimensions[rn].height = 17
        cel(ws, rn, 1, 'P. Nacionales', bold=True, bg='EBF5FB', fg='1A5276')
        cel(ws, rn, 2, pn['ppto_may'],  align='right', fmt='$ #,##0', bg='EBF5FB', fg='1A5276')
        cel(ws, rn, 3, pn_ml,           align='right', fmt='$ #,##0', bg='EBF5FB', fg='566573')
        cel(ws, rn, 4, pn['acum_may'],  align='right', fmt='$ #,##0', bg=mb, fg=mf, bold=True)
        cel(ws, rn, 5, pn_vs,           align='right', fmt='$ #,##0;[Red]$ -#,##0', bg=vl_bg, fg=vl_fg, bold=True)
        cel(ws, rn, 6, pn_pct_lin,      align='center', fmt='0.0%',   bg=mb, fg=mf, bold=True)
        rn += 1

    # TOTAL EMPRESA
    if pnac_rows:
        te_ppto_may = tp_ppto_may + _s(pnac_rows, 'ppto_may')
        te_acum_may = tp_acum_may + _s(pnac_rows, 'acum_may')
        te_ml       = te_ppto_may * linealidad
        te_vs       = te_acum_may - te_ml
        te_pct_lin  = te_acum_may / te_ml if te_ml else 0

        ws.row_dimensions[rn].height = 18
        for col, (v, f) in enumerate(zip(
            ['TOTAL EMPRESA', te_ppto_may, te_ml, te_acum_may, te_vs, te_pct_lin],
            [None, '$ #,##0', '$ #,##0', '$ #,##0', '$ #,##0;[Red]$ -#,##0', '0.0%']
        ), 1):
            c = ws.cell(row=rn, column=col, value=v)
            c.font      = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
            c.fill      = PatternFill('solid', fgColor='0B2341')
            c.alignment = Alignment(horizontal='right' if isinstance(v,(int,float)) else 'center', vertical='center')
            c.border    = borde
            if f: c.number_format = f
        rn += 1

    ws.auto_filter.ref = f'A3:{end_col}{rn - 2}'


# ══════════════════════════════════════════════════════════════════════
# ANÁLISIS 1 — Capital Inmovilizado en Sobrestock
# ══════════════════════════════════════════════════════════════════════

def write_capital_inmovilizado(owb, sobrestock):
    """
    Hoja 'Capital Inmovilizado': jerarquía desplegable 3 niveles.
    Nivel 0 (siempre visible): Marca
    Nivel 1 (colapsado):       Cat. Padre
    Nivel 2 (colapsado):       Cat. Hijo
    Nivel 3 (colapsado):       SKU + descripción
    Ordenado por capital inmovilizado descendente en cada nivel.
    """
    if not sobrestock:
        return

    from openpyxl.worksheet.properties import WorksheetProperties, Outline

    OPTIMO_MESES = 4
    # 3 meses rolling para el promedio: mes actual + próximos 2
    _m1, _m2, _m3 = _CUR_MONTH, _CUR_MONTH + 1, _CUR_MONTH + 2
    _m3_abbr = _MES_ABBR_DYN.get(_m3, '')

    rows = []
    for r in sobrestock:
        # Promedio venta CST rolling 3 meses (mes actual + próximos 2)
        v1 = r['venta_mes'].get(_m1, 0)
        v2 = r['venta_mes'].get(_m2, 0)
        v3 = r['venta_mes'].get(_m3, 0)
        venta_prom = (v1 + v2 + v3) / 3
        stock  = r['stock_cst']
        cobert = r['cobert']
        if venta_prom <= 0:
            continue
        stock_optimo   = venta_prom * OPTIMO_MESES
        capital_exceso = max(0, stock - stock_optimo)
        meses_exceso   = max(0, cobert - OPTIMO_MESES)
        rows.append({**r,
                     'venta_cst':      venta_prom,     # promedio 3 meses para cálculo y display
                     'stock_optimo':   stock_optimo,
                     'capital_exceso': capital_exceso,
                     'meses_exceso':   round(meses_exceso, 1)})
    if not rows:
        return

    # ── Construir jerarquía marca → cat_padre → cat_hijo → [skus] ────────
    hier = {}
    for r in rows:
        m, cp, ch = r['marca'], r['cat_padre'], r['cat_hijo']
        hier.setdefault(m, {}).setdefault(cp, {}).setdefault(ch, []).append(r)

    def _sum(lst, key):   return sum(r[key] for r in lst)
    def _flat(cp_dict):   return [r for ch_dict in cp_dict.values()
                                  for sku_list in ch_dict.values() for r in sku_list]

    # Ordenar cada nivel por capital_exceso desc
    for m in hier:
        for cp in hier[m]:
            for ch in hier[m][cp]:
                hier[m][cp][ch].sort(key=lambda r: -r['capital_exceso'])
        hier[m] = dict(sorted(hier[m].items(),
                              key=lambda kv: -sum(r['capital_exceso']
                                                  for ch in kv[1].values() for r in ch)))
        for cp in hier[m]:
            hier[m][cp] = dict(sorted(hier[m][cp].items(),
                                      key=lambda kv: -sum(r['capital_exceso'] for r in kv[1])))
    hier = dict(sorted(hier.items(),
                       key=lambda kv: -_sum(_flat(kv[1]), 'capital_exceso')))


    # ── Hoja ─────────────────────────────────────────────────────────────
    ws = owb.create_sheet('Capital Inmovilizado')
    ws.sheet_properties = WorksheetProperties(
        outlinePr=Outline(summaryBelow=False, summaryRight=False)
    )
    ws.freeze_panes = 'A3'
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 36

    COLS = [
        ('Marca / Cat. Padre / Cat. Hijo / SKU', 36),
        ('Descripción',           32),
        ('SKUs',                   7),
        ('Cobert.\nACT (m)',      11),
        ('Meses\nExceso',         11),
        ('Stock CST\n($)',         18),
        (f'Venta CST\nProm {_CUR_MES}-{_m3_abbr} ($)', 18),
        ('Stock\nÓptimo ($)',      18),
        ('Capital\nInmovil.($)',  18),
        ('Tiene\nLlegadas',       12),
    ]
    NCOLS   = len(COLS)
    end_col = get_column_letter(NCOLS)

    tot_cap = sum(r['capital_exceso'] for r in rows)
    ws.merge_cells(f'A1:{end_col}1')
    t = ws.cell(row=1, column=1,
                value=f'CAPITAL INMOVILIZADO EN SOBRESTOCK  |  '
                      f'Total exceso sobre {OPTIMO_MESES} meses óptimos: ${tot_cap:,.0f}')
    t.font      = Font(bold=True, color='FFFFFF', name='Calibri', size=12)
    t.fill      = PatternFill('solid', fgColor='4A235A')
    t.alignment = Alignment(horizontal='center', vertical='center')

    for col, (h, w) in enumerate(COLS, 1):
        hdr(ws, 2, col, h, w)

    def _cob_bg(meses):
        if meses > 8: return ('E8DAEF', '6C3483')
        if meses > 4: return ('FADBD8', 'CB4335')
        return               ('FDEBD0', 'E67E22')

    MARCA_BG = '4A235A'   # morado muy oscuro
    CP_BG    = '7D3C98'   # morado medio
    CH_BG    = 'D7BDE2'   # morado claro
    SKU_BG1  = 'F4ECF7'
    SKU_BG2  = 'FFFFFF'

    def _row_marca(rn, label, n_skus, stk, vta, opt, cap, lleg):
        cel(ws, rn, 1, f'▶  {label}', bold=True, bg=MARCA_BG, fg='FFFFFF')
        cel(ws, rn, 2, '',            bg=MARCA_BG)
        cel(ws, rn, 3, n_skus,        align='center',  bg=MARCA_BG, fg='FFFFFF', bold=True)
        cel(ws, rn, 4, '',            bg=MARCA_BG)
        cel(ws, rn, 5, '',            bg=MARCA_BG)
        cel(ws, rn, 6, stk,  align='right', fmt='$ #,##0', bg=MARCA_BG, fg='FFFFFF', bold=True)
        cel(ws, rn, 7, vta,  align='right', fmt='$ #,##0', bg=MARCA_BG, fg='D7BDE2')
        cel(ws, rn, 8, opt,  align='right', fmt='$ #,##0', bg=MARCA_BG, fg='D7BDE2')
        cel(ws, rn, 9, cap,  align='right', fmt='$ #,##0', bg=MARCA_BG, fg='FFFFFF', bold=True)
        ltxt = f'{lleg}/{n_skus} c/llegada' if lleg else 'Sin llegadas'
        cel(ws, rn, 10, ltxt, align='center', bg=MARCA_BG, fg='FFFFFF')
        ws.row_dimensions[rn].height = 17

    def _row_cp(rn, label, n_skus, stk, vta, opt, cap, lleg):
        cel(ws, rn, 1, f'   ▸  {label}', bold=True, bg=CP_BG, fg='FFFFFF')
        cel(ws, rn, 2, '',               bg=CP_BG)
        cel(ws, rn, 3, n_skus, align='center', bg=CP_BG, fg='FFFFFF', bold=True)
        cel(ws, rn, 4, '',     bg=CP_BG)
        cel(ws, rn, 5, '',     bg=CP_BG)
        cel(ws, rn, 6, stk,  align='right', fmt='$ #,##0', bg=CP_BG, fg='FFFFFF', bold=True)
        cel(ws, rn, 7, vta,  align='right', fmt='$ #,##0', bg=CP_BG, fg='D7BDE2')
        cel(ws, rn, 8, opt,  align='right', fmt='$ #,##0', bg=CP_BG, fg='D7BDE2')
        cel(ws, rn, 9, cap,  align='right', fmt='$ #,##0', bg=CP_BG, fg='FFFFFF', bold=True)
        ltxt = f'{lleg}/{n_skus}' if lleg else '—'
        cel(ws, rn, 10, ltxt, align='center', bg=CP_BG, fg='FFFFFF')
        ws.row_dimensions[rn].outline_level = 1
        ws.row_dimensions[rn].hidden        = True
        ws.row_dimensions[rn].height        = 16

    def _row_ch(rn, label, n_skus, stk, vta, opt, cap, lleg):
        cel(ws, rn, 1, f'         ▹  {label}', bold=True, bg=CH_BG, fg='4A235A')
        cel(ws, rn, 2, '',                     bg=CH_BG)
        cel(ws, rn, 3, n_skus, align='center', bg=CH_BG, fg='4A235A', bold=True)
        cel(ws, rn, 4, '',     bg=CH_BG)
        cel(ws, rn, 5, '',     bg=CH_BG)
        cel(ws, rn, 6, stk,  align='right', fmt='$ #,##0', bg=CH_BG, fg='4A235A', bold=True)
        cel(ws, rn, 7, vta,  align='right', fmt='$ #,##0', bg=CH_BG, fg='566573')
        cel(ws, rn, 8, opt,  align='right', fmt='$ #,##0', bg=CH_BG, fg='566573')
        cel(ws, rn, 9, cap,  align='right', fmt='$ #,##0', bg=CH_BG, fg='4A235A', bold=True)
        ltxt = f'{lleg}/{n_skus}' if lleg else '—'
        cel(ws, rn, 10, ltxt, align='center', bg=CH_BG, fg='4A235A')
        ws.row_dimensions[rn].outline_level = 2
        ws.row_dimensions[rn].hidden        = True
        ws.row_dimensions[rn].height        = 16

    rn = 3
    for marca, cp_dict in hier.items():
        m_rows = _flat(cp_dict)
        _row_marca(rn, marca,
                   len(m_rows),
                   _sum(m_rows,'stock_cst'), _sum(m_rows,'venta_cst'),
                   _sum(m_rows,'stock_optimo'), _sum(m_rows,'capital_exceso'),
                   sum(1 for r in m_rows if r['has_emb']))
        rn += 1

        for cat_padre, ch_dict in cp_dict.items():
            cp_rows = [r for ch in ch_dict.values() for r in ch]
            _row_cp(rn, cat_padre,
                    len(cp_rows),
                    _sum(cp_rows,'stock_cst'), _sum(cp_rows,'venta_cst'),
                    _sum(cp_rows,'stock_optimo'), _sum(cp_rows,'capital_exceso'),
                    sum(1 for r in cp_rows if r['has_emb']))
            rn += 1

            for cat_hijo, sku_rows in ch_dict.items():
                _row_ch(rn, cat_hijo,
                        len(sku_rows),
                        _sum(sku_rows,'stock_cst'), _sum(sku_rows,'venta_cst'),
                        _sum(sku_rows,'stock_optimo'), _sum(sku_rows,'capital_exceso'),
                        sum(1 for r in sku_rows if r['has_emb']))
                rn += 1

                for i, r in enumerate(sku_rows):
                    bg = SKU_BG1 if i % 2 == 0 else SKU_BG2
                    cb, cf = _cob_bg(r['meses_exceso'])
                    cel(ws, rn, 1, f'               ↳  {r["sku"]}',
                        bg=bg, fg=C_AZUL, size=8)
                    cel(ws, rn, 2, r['desc'],          bg=bg, size=8)
                    cel(ws, rn, 3, '',                 bg=bg)
                    cel(ws, rn, 4, r['cobert'],        align='center', fmt='0.00',
                        bg=cb, fg=cf, size=8, bold=True)
                    cel(ws, rn, 5, r['meses_exceso'],  align='center', fmt='0.0',
                        bg=cb, fg=cf, size=8)
                    cel(ws, rn, 6, r['stock_cst'],     align='right', fmt='$ #,##0',
                        bg=bg, size=8)
                    cel(ws, rn, 7, r['venta_cst'],     align='right', fmt='$ #,##0',
                        bg=bg, size=8)
                    cel(ws, rn, 8, r['stock_optimo'],  align='right', fmt='$ #,##0',
                        bg=bg, fg='566573', size=8)
                    cel(ws, rn, 9, r['capital_exceso'],align='right', fmt='$ #,##0',
                        bg=cb, fg=cf, size=8, bold=True)
                    tiene = '⚠ SÍ' if r['has_emb'] else 'No'
                    cel(ws, rn, 10, tiene, align='center', size=8,
                        bg=C_NARAN_BG if r['has_emb'] else bg,
                        fg=C_NARAN    if r['has_emb'] else '566573',
                        bold=r['has_emb'])
                    ws.row_dimensions[rn].outline_level = 3
                    ws.row_dimensions[rn].hidden        = True
                    ws.row_dimensions[rn].height        = 14
                    rn += 1

    # ── Total general ─────────────────────────────────────────────────────
    ws.row_dimensions[rn].height = 4
    for c in range(1, NCOLS + 1):
        ws.cell(row=rn, column=c).fill = PatternFill('solid', fgColor='4A235A')
    rn += 1
    ws.row_dimensions[rn].height = 17
    tot_stk = sum(r['stock_cst']    for r in rows)
    tot_vta = sum(r['venta_cst']    for r in rows)
    tot_opt = sum(r['stock_optimo'] for r in rows)
    for col, (v, f) in enumerate(zip(
        ['TOTAL', '', len(rows), '', '', tot_stk, tot_vta, tot_opt, tot_cap, ''],
        [None,None,'0',None,None,'$ #,##0','$ #,##0','$ #,##0','$ #,##0',None]
    ), 1):
        c = ws.cell(row=rn, column=col, value=v)
        c.font      = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
        c.fill      = PatternFill('solid', fgColor='4A235A')
        c.alignment = Alignment(
            horizontal='right' if isinstance(v,(int,float)) else 'center',
            vertical='center')
        c.border    = borde
        if f: c.number_format = f


# ══════════════════════════════════════════════════════════════════════
# ANÁLISIS — Sobrestock por Marca → Cat.Padre → Cat.Hijo → SKU Padre → SKU
# ══════════════════════════════════════════════════════════════════════

def write_sobrestock_sku_padre(owb, sobrestock):
    """
    Hoja 'Sobrestock x SKU Padre': jerarquía 4 niveles desplegable.
      Nivel 0 (siempre visible): Marca          — morado oscuro
      Nivel 1 (colapsado):       Cat. Padre     — morado medio
      Nivel 2 (colapsado):       Cat. Hijo      — morado claro
      Nivel 3 (colapsado):       SKU Padre      — azul marino
      Nivel 4 (colapsado):       SKU + Desc     — filas alternas
    Ordenado por capital excedente desc en cada nivel.
    SKUs sin sku_padre se agrupan bajo '(Sin Padre)'.
    """
    if not sobrestock:
        return

    from openpyxl.worksheet.properties import WorksheetProperties, Outline

    OPTIMO_MESES = 4
    _m1, _m2, _m3 = _CUR_MONTH, _CUR_MONTH + 1, _CUR_MONTH + 2
    _m3_abbr = _MES_ABBR_DYN.get(_m3, '')

    # ── Calcular capital excedente por SKU ────────────────────────────
    rows = []
    for r in sobrestock:
        v1 = r['venta_mes'].get(_m1, 0)
        v2 = r['venta_mes'].get(_m2, 0)
        v3 = r['venta_mes'].get(_m3, 0)
        venta_prom = (v1 + v2 + v3) / 3
        if venta_prom <= 0:
            continue
        stock_optimo   = venta_prom * OPTIMO_MESES
        capital_exceso = max(0, r['stock_cst'] - stock_optimo)
        meses_exceso   = max(0, r['cobert'] - OPTIMO_MESES)
        sp = r.get('sku_padre', '') or ''
        sp = str(sp).strip() if sp else ''
        if not sp: sp = '(Sin Padre)'
        rows.append({**r,
                     'sku_padre_key': sp,
                     'venta_cst':      venta_prom,
                     'stock_optimo':   stock_optimo,
                     'capital_exceso': capital_exceso,
                     'meses_exceso':   round(meses_exceso, 1)})
    if not rows:
        return

    # ── Construir jerarquía marca → cat_padre → cat_hijo → sku_padre → [skus] ──
    hier = {}
    for r in rows:
        m, cp, ch, sp = r['marca'], r['cat_padre'], r['cat_hijo'], r['sku_padre_key']
        hier.setdefault(m, {}).setdefault(cp, {}).setdefault(ch, {}).setdefault(sp, []).append(r)

    def _sum(lst, k):  return sum(r[k] for r in lst)
    def _flat_sp(sp_d):  return [r for lst in sp_d.values()  for r in lst]
    def _flat_ch(ch_d):  return [r for sp_d in ch_d.values() for lst in sp_d.values() for r in lst]
    def _flat_cp(cp_d):  return [r for ch_d in cp_d.values() for sp_d in ch_d.values()
                                 for lst in sp_d.values() for r in lst]

    # Ordenar cada nivel por capital_exceso desc
    for m in hier:
        for cp in hier[m]:
            for ch in hier[m][cp]:
                for sp in hier[m][cp][ch]:
                    hier[m][cp][ch][sp].sort(key=lambda r: -r['capital_exceso'])
                hier[m][cp][ch] = dict(sorted(hier[m][cp][ch].items(),
                    key=lambda kv: -_sum(kv[1], 'capital_exceso')))
            hier[m][cp] = dict(sorted(hier[m][cp].items(),
                key=lambda kv: -_sum(_flat_sp(kv[1]), 'capital_exceso')))
        hier[m] = dict(sorted(hier[m].items(),
            key=lambda kv: -_sum(_flat_ch(kv[1]), 'capital_exceso')))
    hier = dict(sorted(hier.items(),
        key=lambda kv: -_sum(_flat_cp(kv[1]), 'capital_exceso')))

    # ── Hoja ─────────────────────────────────────────────────────────
    ws = owb.create_sheet('Sobrestock x SKU Padre')
    ws.sheet_properties = WorksheetProperties(
        outlinePr=Outline(summaryBelow=False, summaryRight=False)
    )
    ws.freeze_panes = 'A3'
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 36

    COLS = [
        ('Marca / Cat. Padre / Cat. Hijo / SKU Padre / SKU', 40),
        ('Descripción',                   32),
        ('SKUs',                           7),
        ('Cobert.\nACT (m)',              11),
        ('Meses\nExceso',                 11),
        ('Stock CST\n($)',                18),
        (f'Venta CST\nProm {_CUR_MES}-{_m3_abbr} ($)', 18),
        ('Stock\nÓptimo ($)',             18),
        ('Capital\nInmovil. ($)',         18),
        ('Tiene\nLlegadas',               12),
    ]
    NCOLS   = len(COLS)
    end_col = get_column_letter(NCOLS)

    tot_cap  = sum(r['capital_exceso'] for r in rows)
    tot_skus = len(rows)

    ws.merge_cells(f'A1:{end_col}1')
    t = ws.cell(row=1, column=1,
                value=f'CAPITAL INMOVILIZADO — VISTA POR SKU PADRE  |  '
                      f'Total exceso sobre {OPTIMO_MESES} meses óptimos: ${tot_cap:,.0f}')
    t.font      = Font(bold=True, color='FFFFFF', name='Calibri', size=12)
    t.fill      = PatternFill('solid', fgColor='4A235A')
    t.alignment = Alignment(horizontal='center', vertical='center')

    for col, (h, w) in enumerate(COLS, 1):
        hdr(ws, 2, col, h, w)

    # ── Colores (misma paleta que Capital Inmovilizado + nivel SKU Padre en azul) ──
    MARCA_BG   = '4A235A'   # morado muy oscuro
    CP_BG      = '7D3C98'   # morado medio
    CH_BG      = 'D7BDE2'   # morado claro
    SP_BG      = '1A3A5C'   # azul marino oscuro — SKU Padre
    SKU_BG1    = 'EAF2FB'   # azul muy claro — fila impar SKU
    SKU_BG2    = 'FFFFFF'   # blanco — fila par SKU

    def _cob_bg(meses):
        if meses > 8: return ('E8DAEF', '6C3483')
        if meses > 4: return ('FADBD8', 'CB4335')
        return               ('FDEBD0', 'E67E22')

    def _base_desc(sku_rows):
        """Descripción base del SKU Padre: prefijo común + sufijo común de palabras.
        Elimina las palabras variables entre variantes (ej. colores).
        Ejemplo:
          'Parlante Tune Up Blue Lhotse'  ┐
          'Parlante Tune Up Red Lhotse'   ├→ 'Parlante Tune Up Lhotse'
          'Parlante Tune Up Green Lhotse' ┘
        """
        descs = [r['desc'] for r in sku_rows if r.get('desc')]
        if not descs:
            return ''
        if len(descs) == 1:
            return descs[0]
        word_lists = [d.split() for d in descs]
        # Prefijo común palabra a palabra
        prefix = []
        for words in zip(*word_lists):
            if len(set(w.lower() for w in words)) == 1:
                prefix.append(words[0])
            else:
                break
        # Sufijo común palabra a palabra (desde el final)
        suffix = []
        for words in zip(*[list(reversed(wl)) for wl in word_lists]):
            if len(set(w.lower() for w in words)) == 1:
                suffix.append(words[0])
            else:
                break
        suffix.reverse()
        # Evitar solapamiento entre prefijo y sufijo
        p_len = len(prefix)
        s_len = len(suffix)
        min_len = min(len(wl) for wl in word_lists)
        if p_len + s_len >= min_len:
            s_len = max(0, min_len - p_len)
            suffix = suffix[-s_len:] if s_len else []
        return ' '.join(prefix + suffix).strip()

    def _write_agg(rn, label, lst, outline_lvl, bg, fg_main, fg_num, fg_dim,
                   hidden=False, height=17, desc=''):
        """Fila de agrupación genérica (Marca/CP/CH/SP)."""
        n    = len(lst)
        stk  = _sum(lst, 'stock_cst')
        vta  = _sum(lst, 'venta_cst')
        opt  = _sum(lst, 'stock_optimo')
        cap  = _sum(lst, 'capital_exceso')
        lleg = sum(1 for r in lst if r['has_emb'])
        ltxt = f'{lleg}/{n} c/llegada' if lleg else 'Sin llegadas'
        cel(ws, rn, 1, label,  bold=True, bg=bg, fg=fg_main)
        cel(ws, rn, 2, desc,   bg=bg, fg='AED6F1' if desc else fg_main, size=9)
        cel(ws, rn, 3, n,      align='center', bg=bg, fg=fg_main, bold=True)
        cel(ws, rn, 4, '',                bg=bg)
        cel(ws, rn, 5, '',                bg=bg)
        cel(ws, rn, 6, stk,  align='right', fmt='$ #,##0', bg=bg, fg=fg_main, bold=True)
        cel(ws, rn, 7, vta,  align='right', fmt='$ #,##0', bg=bg, fg=fg_dim)
        cel(ws, rn, 8, opt,  align='right', fmt='$ #,##0', bg=bg, fg=fg_dim)
        cel(ws, rn, 9, cap,  align='right', fmt='$ #,##0', bg=bg, fg=fg_main, bold=True)
        cel(ws, rn, 10, ltxt, align='center', bg=bg, fg=fg_num)
        ws.row_dimensions[rn].outline_level = outline_lvl
        ws.row_dimensions[rn].hidden        = hidden
        ws.row_dimensions[rn].height        = height

    def _row_sku(rn, r, i, outline_lvl):
        bg = SKU_BG1 if i % 2 == 0 else SKU_BG2
        cb, cf = _cob_bg(r['meses_exceso'])
        cel(ws, rn, 1, f'               ↳  {r["sku"]}',
            bg=bg, fg='1A5276', size=8)
        cel(ws, rn, 2, r['desc'],           bg=bg, size=8)
        cel(ws, rn, 3, '',                  bg=bg)
        cel(ws, rn, 4, r['cobert'],         align='center', fmt='0.00',
            bg=cb, fg=cf, size=8, bold=True)
        cel(ws, rn, 5, r['meses_exceso'],   align='center', fmt='0.0',
            bg=cb, fg=cf, size=8)
        cel(ws, rn, 6, r['stock_cst'],      align='right', fmt='$ #,##0', bg=bg, size=8)
        cel(ws, rn, 7, r['venta_cst'],      align='right', fmt='$ #,##0', bg=bg, size=8)
        cel(ws, rn, 8, r['stock_optimo'],   align='right', fmt='$ #,##0',
            bg=bg, fg='566573', size=8)
        cel(ws, rn, 9, r['capital_exceso'], align='right', fmt='$ #,##0',
            bg=cb, fg=cf, size=8, bold=True)
        tiene = '⚠ SÍ' if r['has_emb'] else 'No'
        cel(ws, rn, 10, tiene, align='center', size=8,
            bg=C_NARAN_BG if r['has_emb'] else bg,
            fg=C_NARAN    if r['has_emb'] else '566573',
            bold=r['has_emb'])
        ws.row_dimensions[rn].outline_level = outline_lvl
        ws.row_dimensions[rn].hidden        = True
        ws.row_dimensions[rn].height        = 14

    # ── Render ───────────────────────────────────────────────────────
    rn = 3
    for marca, cp_dict in hier.items():
        m_rows = _flat_cp(cp_dict)
        _write_agg(rn, f'▶  {marca}', m_rows,
                   outline_lvl=0, bg=MARCA_BG,
                   fg_main='FFFFFF', fg_num='D7BDE2', fg_dim='D7BDE2',
                   hidden=False, height=17)
        rn += 1

        for cat_padre, ch_dict in cp_dict.items():
            cp_rows = _flat_sp(ch_dict) if False else _flat_ch(ch_dict)
            # cp_rows = todos los SKUs dentro de este cat_padre
            cp_rows = [r for ch in ch_dict.values() for sp_d in ch.values() for r in sp_d]
            _write_agg(rn, f'   ▸  {cat_padre}', cp_rows,
                       outline_lvl=1, bg=CP_BG,
                       fg_main='FFFFFF', fg_num='D7BDE2', fg_dim='D7BDE2',
                       hidden=True, height=16)
            rn += 1

            for cat_hijo, sp_dict in ch_dict.items():
                ch_rows = [r for lst in sp_dict.values() for r in lst]
                _write_agg(rn, f'         ▹  {cat_hijo}', ch_rows,
                           outline_lvl=2, bg=CH_BG,
                           fg_main='4A235A', fg_num='566573', fg_dim='566573',
                           hidden=True, height=16)
                rn += 1

                for sku_padre, sku_rows in sp_dict.items():
                    _write_agg(rn, f'               ◆  {sku_padre}', sku_rows,
                               outline_lvl=3, bg=SP_BG,
                               fg_main='FFFFFF', fg_num='AED6F1', fg_dim='AED6F1',
                               hidden=True, height=15,
                               desc=_base_desc(sku_rows))
                    rn += 1

                    for i, r in enumerate(sku_rows):
                        _row_sku(rn, r, i, outline_lvl=4)
                        rn += 1

    # ── Total general ────────────────────────────────────────────────
    ws.row_dimensions[rn].height = 4
    for c in range(1, NCOLS + 1):
        ws.cell(row=rn, column=c).fill = PatternFill('solid', fgColor='4A235A')
    rn += 1
    ws.row_dimensions[rn].height = 17
    tot_stk = sum(r['stock_cst']    for r in rows)
    tot_vta = sum(r['venta_cst']    for r in rows)
    tot_opt = sum(r['stock_optimo'] for r in rows)
    for col, (v, f) in enumerate(zip(
        ['TOTAL', '', tot_skus, '', '', tot_stk, tot_vta, tot_opt, tot_cap, ''],
        [None,None,'0',None,None,'$ #,##0','$ #,##0','$ #,##0','$ #,##0',None]
    ), 1):
        c = ws.cell(row=rn, column=col, value=v)
        c.font      = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
        c.fill      = PatternFill('solid', fgColor='4A235A')
        c.alignment = Alignment(
            horizontal='right' if isinstance(v, (int, float)) else 'center',
            vertical='center')
        c.border    = borde
        if f: c.number_format = f

    # Nota al pie
    rn += 2
    nota = ws.cell(row=rn, column=1,
                   value=f'Capital excedente = Stock CST − (Venta Prom {_CUR_MES}‑{_m3_abbr} × {OPTIMO_MESES} meses óptimos)  |  '
                         f'SKUs sin venta en los próximos 3 meses se excluyen del cálculo.')
    nota.font = Font(name='Calibri', size=8, italic=True, color='7F8C8D')
    ws.merge_cells(f'A{rn}:{end_col}{rn}')


# ══════════════════════════════════════════════════════════════════════
# ANÁLISIS 4 — Sobrestock con Llegada Encima
# ══════════════════════════════════════════════════════════════════════

def write_sobrestock_con_llegada(owb, sobrestock):
    """
    Hoja 'Sobrestock c/Llegada': SKUs con >6 meses de cobertura
    que además tienen un embarque en tránsito llegando.
    Ordenado por cobertura descendente (peor primero).
    Venta mostrada = promedio 3 meses (mes actual + 2 futuros).
    """
    rows = [r for r in sobrestock if r['has_emb'] and r['prox'] != 'Sin embarque']
    if not rows:
        return
    rows.sort(key=lambda r: -r['cobert'])

    # Meses para promedio de venta: mes actual + 2 futuros
    _m1, _m2, _m3 = _CUR_MONTH, _CUR_MONTH + 1, _CUR_MONTH + 2
    _m1_lbl = _MES_ABBR_DYN.get(_m1, '')
    _m3_lbl = _MES_ABBR_DYN.get(_m3, '')

    ws = owb.create_sheet('Sobrestock c-Llegada')
    ws.freeze_panes = 'A3'
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 36

    COLS = [
        ('Marca',            18), ('Cat. Comercial',  22),
        ('SKU',              12), ('Descripción',     35),
        ('Cobert.\nACT (m)', 11), ('Stock\nCST ($)',  18),
        (f'Venta Prom.\n{_m1_lbl}-{_m3_lbl} ($)', 18), ('Prox.\nLlegada', 12),
        ('PI\nEmbarque',     16), ('ETA\nBodega',     13),
    ]
    NCOLS   = len(COLS)
    end_col = get_column_letter(NCOLS)

    ws.merge_cells(f'A1:{end_col}1')
    t = ws.cell(row=1, column=1,
                value=f'SOBRESTOCK CON LLEGADA ENCIMA  |  '
                      f'{len(rows)} SKUs con >6 meses de cobertura recibiendo más stock')
    t.font      = Font(bold=True, color='FFFFFF', name='Calibri', size=12)
    t.fill      = PatternFill('solid', fgColor='784212')
    t.alignment = Alignment(horizontal='center', vertical='center')

    for col, (h, w) in enumerate(COLS, 1):
        hdr(ws, 2, col, h, w)

    def _cob_style(c):
        if c > 12: return ('E8DAEF', '6C3483')
        if c > 9:  return ('FADBD8', 'CB4335')
        return            ('FDEBD0', 'E67E22')

    rn = 3
    for r in rows:
        bg = C_ALT if rn % 2 == 0 else 'FFFFFF'
        cb, cf = _cob_style(r['cobert'])
        eta_str = r['eta'].strftime('%d/%m/%Y') if isinstance(r['eta'], (_dt.date, _dt.datetime)) else (r['eta'] or '—')
        # Promedio venta 3 meses: mes actual + 2 futuros
        v1 = r['venta_mes'].get(_m1, 0)
        v2 = r['venta_mes'].get(_m2, 0)
        v3 = r['venta_mes'].get(_m3, 0)
        venta_prom = (v1 + v2 + v3) / 3
        cel(ws, rn, 1,  r['marca'],     bold=True, bg=bg)
        cel(ws, rn, 2,  r['cat_com'],   bg=bg)
        cel(ws, rn, 3,  r['sku'],       align='center', bg=bg)
        cel(ws, rn, 4,  r['desc'],      bg=bg)
        cel(ws, rn, 5,  r['cobert'],    align='center', fmt='0.00', bg=cb, fg=cf, bold=True)
        cel(ws, rn, 6,  r['stock_cst'], align='right',  fmt='$ #,##0', bg=bg)
        cel(ws, rn, 7,  venta_prom,     align='right',  fmt='$ #,##0', bg=bg)
        cel(ws, rn, 8,  r['prox'],      align='center', bg='FDEBD0', fg='784212', bold=True)
        cel(ws, rn, 9,  r['pi'],        align='center', bg=bg, fg='566573')
        cel(ws, rn, 10, eta_str,        align='center', bg='FADBD8', fg='CB4335')
        ws.row_dimensions[rn].height = 15
        rn += 1

    ws.auto_filter.ref = f'A2:{end_col}{rn - 1}'


# ══════════════════════════════════════════════════════════════════════
# ANÁLISIS 2 — Fecha de Quiebre Proyectada (SKUs Críticos)
# ══════════════════════════════════════════════════════════════════════

def write_fecha_quiebre(owb, criticos):
    """
    Hoja 'Fecha de Quiebre': para cada SKU crítico calcula el día
    exacto en que se agota el stock, y si la OC en tránsito llega a tiempo.

    LÓGICA:
      Días Restantes = Stock CST actual ÷ (Venta CST ABR ÷ 30 días)
      → Asume que la venta diaria es constante al ritmo del mes ABR.
      Fecha Quiebre = Hoy + Días Restantes.
      Si la ETA de la OC es ≤ Fecha Quiebre → llega a tiempo.
      Si la ETA es > Fecha Quiebre → la OC llega DESPUÉS del quiebre.

    Columnas: Marca | Cat.Comercial | Cat.Padre | Cat.Hijo | SKU | Desc | ...
    Ordenado por días restantes ascendente (más urgente primero).
    """
    import calendar as _cal
    today       = _dt.date.today()
    YEAR        = today.year
    today_month = today.month
    today_day   = today.day

    MES_NOMBRES = {4:'ABR 26',5:'MAY 26',6:'JUN 26',7:'JUL 26',8:'AGO 26'}

    def _proyectar_quiebre_mes(r):
        """
        Proyecta en qué MES se agota el stock.
        ABR: stock hoy vs venta proporcional a días restantes del mes.
        MAY en adelante: StkIni del FLAT (ya incluye tránsitos) vs VentaMes.
        Retorna nombre_mes (ej: 'MAY 26') o None si no quiebra en horizonte.
        """
        stock = float(r['stock_cst'])

        for month_num in [4, 5, 6, 7, 8]:
            if month_num < today_month:
                continue

            venta = r['venta_mes'].get(month_num, 0)

            if month_num == today_month:
                # ABR: stock hoy vs consumo de los días que quedan
                if venta <= 0:
                    continue
                days_total    = _cal.monthrange(YEAR, month_num)[1]
                dias_restantes = days_total - today_day + 1
                consumo_resto = venta * (dias_restantes / days_total)
                if stock < consumo_resto:
                    return MES_NOMBRES[month_num]
                # Sobrevive ABR — para MAY en adelante usa Stk+Ped del FLAT
            else:
                # MAY–AGO: usar Stk+Ped (StkIni + Llegadas confirmadas)
                stk_ped = r['stk_ini_mes'].get(month_num, 0.0)
                if stk_ped <= 0:
                    return MES_NOMBRES[month_num]
                if venta > 0 and stk_ped < venta:
                    return MES_NOMBRES[month_num]

        return None  # Sin quiebre en horizonte ABR–AGO

    rows = []
    for r in criticos:
        stock = r['stock_cst']
        eta   = r['eta']
        if isinstance(eta, _dt.datetime):
            eta = eta.date()

        if stock <= 0:
            mes_quiebre = None   # "⛔ Quebrado"
            ya_quebrado = True
        else:
            mes_quiebre = _proyectar_quiebre_mes(r)
            ya_quebrado = False

        # Estado: evalúa OC independientemente de si ya quebró o no
        # Para comparar OC con mes de quiebre usamos el primer día del mes
        _MES_NUM = {'ABR 26':4,'MAY 26':5,'JUN 26':6,'JUL 26':7,'AGO 26':8}
        if not ya_quebrado and mes_quiebre is None:
            estado  = '✔ Sin quiebre proyectado'
            est_bg, est_fg = C_VERDE_BG, C_VERDE
        elif r['has_emb'] and eta:
            if ya_quebrado:
                # Ya quebrado — OC llega tarde de todas formas
                estado  = f'⚠ Quebrado — OC ETA {eta.strftime("%d/%m")}'
                est_bg, est_fg = C_NARAN_BG, C_NARAN
            else:
                mn  = _MES_NUM.get(mes_quiebre, 12)
                fq1 = _dt.date(YEAR, mn, 1)   # primer día del mes de quiebre
                if eta <= fq1:
                    estado  = f'✔ OC a tiempo — ETA {eta.strftime("%d/%m")}'
                    est_bg, est_fg = C_VERDE_BG, C_VERDE
                else:
                    estado = (f'⚠ Quiebre {mes_quiebre} — '
                              f'OC {eta.strftime("%d/%m")}')
                    est_bg, est_fg = C_NARAN_BG, C_NARAN
        else:
            prox = r.get('prox', 'Sin embarque')
            if prox and prox != 'Sin embarque':
                estado  = f'✘ Sin OC — llegada proy. {prox}'
            else:
                estado  = '✘ Sin OC abierta'
            est_bg, est_fg = C_ROJO_BG, C_ROJO

        rows.append({**r,
                     'mes_quiebre': mes_quiebre,
                     'ya_quebrado': ya_quebrado,
                     'eta_d':       eta,
                     'estado':      estado,
                     'est_bg':      est_bg,
                     'est_fg':      est_fg})

    _MES_ORD = {'ABR 26':0,'MAY 26':1,'JUN 26':2,'JUL 26':3,'AGO 26':4}
    rows.sort(key=lambda r: (-1 if r['ya_quebrado'] else
                              _MES_ORD.get(r['mes_quiebre'], 99)))

    ws = owb.create_sheet('Fecha de Quiebre')
    ws.freeze_panes = 'A4'
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 36

    COLS = [
        ('Marca',              18), ('Cat. Comercial',   18),
        ('Cat. Padre',         16), ('Cat. Hijo',        16),
        ('SKU',                12), ('Descripción',      33),
        ('Cobert.\nACT (m)',   11), ('Stock\nCST ($)',   18),
        ('Venta CST\nABR ($)', 18), ('Mes de\nQuiebre',  13),
        ('ETA\nOC',            13), ('Estado',           28),
    ]
    NCOLS   = len(COLS)
    end_col = get_column_letter(NCOLS)

    sin_oc = sum(1 for r in rows if '✘' in r['estado'])
    tarde  = sum(1 for r in rows if '⚠' in r['estado'])

    ws.merge_cells(f'A1:{end_col}1')
    t = ws.cell(row=1, column=1,
                value=f'FECHA DE QUIEBRE PROYECTADA  |  '
                      f'{sin_oc} sin OC  |  {tarde} con OC tardía  |  '
                      f'{len(rows) - sin_oc - tarde} con OC a tiempo')
    t.font      = Font(bold=True, color='FFFFFF', name='Calibri', size=12)
    t.fill      = PatternFill('solid', fgColor='922B21')
    t.alignment = Alignment(horizontal='center', vertical='center')

    # Fila 2: nota metodológica
    ws.merge_cells(f'A2:{end_col}2')
    nota = ws.cell(row=2, column=1,
                   value='Proyección mes a mes: ABR con stock real hoy → MAY-AGO con StkIni del FLAT '
                         '(tránsitos incluidos)  |  ⛔ Quebrado = sin stock hoy  '
                         '|  ⚠ Quiebre DD/MM — OC DD/MM = gap antes que llegue el embarque')
    nota.font      = Font(italic=True, size=9, color='F5B7B1', name='Calibri')
    nota.fill      = PatternFill('solid', fgColor='922B21')
    nota.alignment = Alignment(horizontal='center', vertical='center')

    for col, (h, w) in enumerate(COLS, 1):
        hdr(ws, 3, col, h, w)

    _MES_BG = {
        'ABR 26': ('FADBD8', 'CB4335'),   # rojo — urgente
        'MAY 26': ('FDEBD0', 'E67E22'),   # naranja
        'JUN 26': ('FEF9E7', 'B7950B'),   # amarillo
        'JUL 26': ('EBF5FB', '1A5276'),   # azul claro
        'AGO 26': ('E8F8F5', '1E8449'),   # verde claro
    }

    rn = 4
    for r in rows:
        bg  = C_ALT if rn % 2 == 0 else 'FFFFFF'
        mq  = r['mes_quiebre']
        yq  = r['ya_quebrado']
        eta_str = r['eta_d'].strftime('%d/%m/%Y') if r['eta_d'] else '—'

        if yq:
            mq_str = '⛔ Quebrado'
            mb, mf = ('C0392B', 'FFFFFF')
        elif mq is None:
            mq_str = '∞ sin quiebre'
            mb, mf = (C_VERDE_BG, C_VERDE)
        else:
            mq_str = mq
            mb, mf = _MES_BG.get(mq, ('FFFFFF', '000000'))

        cel(ws, rn, 1,  r['marca'],     bold=True,     bg=bg)
        cel(ws, rn, 2,  r['cat_com'],   bg=bg)
        cel(ws, rn, 3,  r['cat_padre'], bg=bg, fg='566573')
        cel(ws, rn, 4,  r['cat_hijo'],  bg=bg, fg='566573')
        cel(ws, rn, 5,  r['sku'],       align='center', bg=bg)
        cel(ws, rn, 6,  r['desc'],      bg=bg)
        cel(ws, rn, 7,  r['cobert'],    align='center', fmt='0.000', bg=mb, fg=mf)
        cel(ws, rn, 8,  r['stock_cst'], align='right',  fmt='$ #,##0', bg=bg)
        cel(ws, rn, 9,  r['venta_cst'], align='right',  fmt='$ #,##0', bg=bg)
        cel(ws, rn, 10, mq_str,         align='center', bg=mb, fg=mf, bold=True)
        cel(ws, rn, 11, eta_str,        align='center', bg=bg, fg='566573')
        cel(ws, rn, 12, r['estado'],    align='center', bg=r['est_bg'], fg=r['est_fg'], bold=True)

        ws.row_dimensions[rn].height = 15
        rn += 1

    ws.auto_filter.ref = f'A3:{end_col}{rn - 1}'


# ══════════════════════════════════════════════════════════════════════
# ANÁLISIS 3 — Concentración de Riesgo por Embarque (PI)
# ══════════════════════════════════════════════════════════════════════

def write_riesgo_embarque(owb, trans_rows_full, all_rows):
    """
    Hoja 'Riesgo por Embarque': por cada PI muestra cuántos SKUs
    críticos e inquietantes dependen de ese embarque, con filas de detalle
    SKU expandibles (outline grouping). SKUs críticos primero.
    Ordenado por SKUs críticos descendente.
    """
    if not trans_rows_full:
        return

    from openpyxl.worksheet.properties import WorksheetProperties, Outline

    # Mapas por SKU desde all_rows
    cob_por_sku  = {r['sku']: r['cobert'] for r in all_rows}
    desc_por_sku = {r['sku']: r['desc']   for r in all_rows}

    # Agrupar tránsitos por PI, guardando lista completa de filas
    pi_rows_map = defaultdict(list)
    for r in trans_rows_full:
        pi = r['pi'] or '(sin PI)'
        pi_rows_map[pi].append(r)

    # Construir resumen por PI
    pi_list = []
    for pi, t_rows in pi_rows_map.items():
        skus_vistos = {}   # sku → {cantidad, valor_usd, desc}
        eta_bod = None
        mes_pi  = None
        marca_set = set()
        for r in t_rows:
            sk = r['sku']
            if sk not in skus_vistos:
                skus_vistos[sk] = {'cantidad': 0, 'valor_usd': 0,
                                   'desc': desc_por_sku.get(sk, r['desc'])}
            skus_vistos[sk]['cantidad']  += r['cantidad']
            skus_vistos[sk]['valor_usd'] += r['valor_usd']
            marca_set.add(r['marca'])
            eta = r['eta_bod']
            if isinstance(eta, _dt.datetime): eta = eta.date()
            if eta and (eta_bod is None or eta < eta_bod):
                eta_bod = eta
            if not mes_pi and r['mes']:
                mes_pi = r['mes']

        criticos_pi   = [s for s in skus_vistos if cob_por_sku.get(s, 99) < 1]
        inquiet_pi    = [s for s in skus_vistos if 1 <= cob_por_sku.get(s, 99) < 2]
        tot_usd_pi    = sum(v['valor_usd'] for v in skus_vistos.values())
        tot_qty_pi    = sum(v['cantidad']  for v in skus_vistos.values())

        # SKU detail: críticos primero, luego inquietantes, luego resto
        def _sku_order(sk):
            cob = cob_por_sku.get(sk, 99)
            if cob < 1:   return 0
            if cob < 2:   return 1
            return 2
        sku_detail = sorted(skus_vistos.keys(), key=_sku_order)

        pi_list.append({
            'pi':         pi,
            'eta':        eta_bod,
            'mes':        mes_pi,
            'marcas':     ', '.join(sorted(marca_set)),
            'n_skus':     len(skus_vistos),
            'n_criticos': len(criticos_pi),
            'n_inquiet':  len(inquiet_pi),
            'valor_usd':  tot_usd_pi,
            'cantidad':   tot_qty_pi,
            'sku_detail': sku_detail,
            'skus_vistos': skus_vistos,
        })

    pi_list.sort(key=lambda r: (-r['n_criticos'], -r['n_inquiet'], -r['valor_usd']))

    # ── Crear hoja ────────────────────────────────────────────────────
    ws = owb.create_sheet('Riesgo por Embarque')
    ws.sheet_properties = WorksheetProperties(
        outlinePr=Outline(summaryBelow=False, summaryRight=False)
    )
    ws.freeze_panes = 'A3'
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 36

    COLS = [
        ('PI / Embarque\n(+ para ver SKUs)', 24), ('ETA Bodega',    13),
        ('Mes\nLlegada',  10), ('Marcas',           24),
        ('SKUs\nTotales', 11), ('Críticos\n(<1m)',  11),
        ('Inquiet.\n(1-2m)',11),('Unidades\nTotal', 14),
        ('Valor\nUSD',    14), ('Nivel\nRiesgo',    14),
    ]
    NCOLS   = len(COLS)
    end_col = get_column_letter(NCOLS)

    tot_usd_all  = sum(r['valor_usd']  for r in pi_list)
    tot_crit_all = sum(r['n_criticos'] for r in pi_list)

    ws.merge_cells(f'A1:{end_col}1')
    t = ws.cell(row=1, column=1,
                value=f'CONCENTRACIÓN DE RIESGO POR EMBARQUE  |  '
                      f'{len(pi_list)} embarques  |  {tot_crit_all} SKUs críticos en tránsito  |  '
                      f'USD {tot_usd_all:,.0f} total')
    t.font      = Font(bold=True, color='FFFFFF', name='Calibri', size=12)
    t.fill      = PatternFill('solid', fgColor='1A2D5A')
    t.alignment = Alignment(horizontal='center', vertical='center')

    for col, (h, w) in enumerate(COLS, 1):
        hdr(ws, 2, col, h, w)

    def _riesgo(n_crit, n_inq):
        if n_crit >= 5: return ('FADBD8', 'CB4335', 'ALTO')
        if n_crit >= 2: return ('FDEBD0', 'E67E22', 'MEDIO')
        if n_crit >= 1: return ('FEF9E7', 'B7950B', 'BAJO')
        if n_inq  >= 3: return ('FEF9E7', 'B7950B', 'VIGILAR')
        return                 ('D5F5E3', '1E8449', 'OK')

    def _sku_estado(cob):
        if cob < 1:  return ('CRITICO',     C_ROJO_BG,  C_ROJO)
        if cob < 2:  return ('INQUIETANTE', C_NARAN_BG, C_NARAN)
        return              ('OK',          'F4F6F7',   '566573')

    rn = 3
    for pi_r in pi_list:
        bg = C_ALT if rn % 2 == 0 else 'FFFFFF'
        rb, rf, rlabel = _riesgo(pi_r['n_criticos'], pi_r['n_inquiet'])
        eta_str = pi_r['eta'].strftime('%d/%m/%Y') if pi_r['eta'] else '—'

        # ── Fila resumen PI (nivel 0) ─────────────────────────────────
        cel(ws, rn, 1,  pi_r['pi'],         bold=True, bg=bg)
        cel(ws, rn, 2,  eta_str,            align='center', bg=bg)
        cel(ws, rn, 3,  pi_r['mes'],        align='center', bg=bg, fg='566573')
        cel(ws, rn, 4,  pi_r['marcas'],     bg=bg, fg='566573')
        cel(ws, rn, 5,  pi_r['n_skus'],     align='center', fmt='0', bg=bg)

        crit_bg = C_ROJO_BG if pi_r['n_criticos'] > 0 else bg
        crit_fg = C_ROJO    if pi_r['n_criticos'] > 0 else '000000'
        cel(ws, rn, 6,  pi_r['n_criticos'], align='center', fmt='0',
            bg=crit_bg, fg=crit_fg, bold=pi_r['n_criticos'] > 0)

        inq_bg = C_NARAN_BG if pi_r['n_inquiet'] > 0 else bg
        inq_fg = C_NARAN    if pi_r['n_inquiet'] > 0 else '000000'
        cel(ws, rn, 7,  pi_r['n_inquiet'],  align='center', fmt='0', bg=inq_bg, fg=inq_fg)

        cel(ws, rn, 8,  pi_r['cantidad'],   align='right', fmt='#,##0', bg=bg)
        cel(ws, rn, 9,  pi_r['valor_usd'],  align='right', fmt='$ #,##0', bg=bg)
        cel(ws, rn, 10, rlabel,             align='center', bg=rb, fg=rf, bold=True)

        ws.row_dimensions[rn].height = 16
        rn += 1

        # ── Filas detalle SKU (nivel 1, colapsadas) ───────────────────
        for sk in pi_r['sku_detail']:
            cob     = cob_por_sku.get(sk, 99)
            det     = pi_r['skus_vistos'][sk]
            estado, sb, sf = _sku_estado(cob)
            desc_sk = det['desc'] or desc_por_sku.get(sk, '')

            cel(ws, rn, 1,  f'  └ {sk}',   bg=sb, fg=sf, bold=(cob < 1))
            cel(ws, rn, 2,  desc_sk,        bg='F8F9FA', fg='2C3E50')
            cel(ws, rn, 3,  '',             bg='F8F9FA')
            cel(ws, rn, 4,  '',             bg='F8F9FA')
            cel(ws, rn, 5,  '',             bg='F8F9FA')
            cel(ws, rn, 6,  estado,         align='center', bg=sb, fg=sf, bold=(cob < 1))
            cel(ws, rn, 7,  round(cob, 2),  align='center', fmt='0.00', bg=sb, fg=sf)
            cel(ws, rn, 8,  det['cantidad'],align='right',  fmt='#,##0', bg='F8F9FA')
            cel(ws, rn, 9,  det['valor_usd'],align='right', fmt='$ #,##0', bg='F8F9FA')
            cel(ws, rn, 10, '',             bg='F8F9FA')

            ws.row_dimensions[rn].outline_level = 1
            ws.row_dimensions[rn].hidden        = True
            ws.row_dimensions[rn].height        = 14
            rn += 1

    ws.auto_filter.ref = f'A2:{end_col}2'


# ══════════════════════════════════════════════════════════════════════
# HOJA CST MP+N — Resumen de costos por Marca
# ══════════════════════════════════════════════════════════════════════

def write_cst_mp_n_sku(owb, data_mp):
    """
    Hoja 'CST x Marca' — versión formateada de TD REPORTES CST MP+N SKU.
    Muestra resumen de costos por Marca con coberturas por mes (ABR→SEP).
    Números en millones de pesos ($M). Coberturas coloreadas.
    """
    if not data_mp:
        return

    # ── Detectar fila de encabezados y filas de datos ─────────────────
    hdr_row_idx = None
    for i, row in enumerate(data_mp):
        if row and str(row[0] or '').strip() == 'Marca':
            hdr_row_idx = i
            break
    if hdr_row_idx is None:
        return

    hdrs = [str(v or '').strip() for v in data_mp[hdr_row_idx]]

    # Mapear columnas por nombre
    def _ci(name):
        nl = name.lower()
        for j, h in enumerate(hdrs):
            if nl in h.lower():
                return j
        return None

    # Índices de columnas clave
    CI_MARCA      = 0
    CI_STK_HOY    = _ci('stock hoy cst')       or 4
    CI_TRANS_ABR  = _ci('transito abr')        or 5
    CI_STP_ABR    = _ci('stock + pedido csto abr') or 6
    CI_VTA_ABR    = _ci('venta cst abr')       or 7
    CI_COB_ACT    = _ci('cobert. act')         or 8
    CI_COB_ABR    = _ci('cobert. abr')         or 9
    # MAY
    CI_STI_MAY    = _ci('stock inicial cst may') or 10
    CI_CMP_MAY    = _ci('compra cst may')      or 11
    CI_LLG_MAY    = _ci('llegadas cst may')    or 12
    CI_STP_MAY    = _ci('stock + pedido csto may') or 13
    CI_VTA_MAY    = _ci('venta cst may')       or 14
    CI_COB_MAY    = _ci('cobert. may')         or 15
    # JUN
    CI_STI_JUN    = _ci('stock inicial cst jun') or 16
    CI_CMP_JUN    = _ci('compra cst jun')      or 17
    CI_LLG_JUN    = _ci('llegadas cst jun')    or 18
    CI_STP_JUN    = _ci('stock + pedido csto jun') or 19
    CI_VTA_JUN    = _ci('venta cst jun')       or 20
    CI_COB_JUN    = _ci('cobert. jun')         or 21
    # JUL
    CI_STI_JUL    = _ci('stock inicial cst jul') or 22
    CI_CMP_JUL    = _ci('compra cst jul')      or 23
    CI_LLG_JUL    = _ci('llegadas cst jul')    or 24
    CI_STP_JUL    = _ci('stock + pedido csto jul') or 25
    CI_VTA_JUL    = _ci('venta cst jul')       or 26
    CI_COB_JUL    = _ci('cobert. jul')         or 27
    # AGO
    CI_STI_AGO    = _ci('stock inicial cst ago') or 28
    CI_CMP_AGO    = _ci('compra cst ago')      or 29
    CI_LLG_AGO    = _ci('llegadas cst ago')    or 30
    CI_STP_AGO    = _ci('stock + pedido csto ago') or 31
    CI_VTA_AGO    = _ci('venta cst ago')       or 32
    CI_COB_AGO    = _ci('cobert. ago')         or 33
    # SEP
    CI_STI_SEP    = _ci('stock inicial cst sep') or 34
    CI_CMP_SEP    = _ci('compra cst sep')      or 35
    CI_LLG_SEP    = _ci('llegadas cst sep')    or 36
    CI_STP_SEP    = _ci('stock + pedido csto sep') or 37
    CI_VTA_SEP    = _ci('venta cst sep')       or 38
    CI_COB_SEP    = _ci('cobert. sep')         or 39

    # ── Leer filas de datos ───────────────────────────────────────────
    SKIP = {'total general', 'total empresa', 'proveedores nacionales',
            'total propia', ''}

    def _v(row, idx, default=0.0):
        try:
            v = row[idx]
            return float(v) if v is not None else default
        except (IndexError, TypeError, ValueError):
            return default

    rows_data   = []   # marcas reales
    rows_totals = []   # total general + prov.nacionales + total empresa

    for i in range(hdr_row_idx + 1, len(data_mp)):
        row = data_mp[i]
        if not row or row[0] is None:
            continue
        marca = str(row[0]).strip()
        if not marca:
            continue
        rec = {
            'marca':     marca,
            'stk_hoy':   _v(row, CI_STK_HOY),
            'trans_abr': _v(row, CI_TRANS_ABR),
            'stp_abr':   _v(row, CI_STP_ABR),
            'vta_abr':   _v(row, CI_VTA_ABR),
            'cob_act':   _v(row, CI_COB_ACT),
            'cob_abr':   _v(row, CI_COB_ABR),
            'sti_may':   _v(row, CI_STI_MAY),
            'cmp_may':   _v(row, CI_CMP_MAY),
            'llg_may':   _v(row, CI_LLG_MAY),
            'stp_may':   _v(row, CI_STP_MAY),
            'vta_may':   _v(row, CI_VTA_MAY),
            'cob_may':   _v(row, CI_COB_MAY),
            'sti_jun':   _v(row, CI_STI_JUN),
            'cmp_jun':   _v(row, CI_CMP_JUN),
            'llg_jun':   _v(row, CI_LLG_JUN),
            'stp_jun':   _v(row, CI_STP_JUN),
            'vta_jun':   _v(row, CI_VTA_JUN),
            'cob_jun':   _v(row, CI_COB_JUN),
            'sti_jul':   _v(row, CI_STI_JUL),
            'cmp_jul':   _v(row, CI_CMP_JUL),
            'llg_jul':   _v(row, CI_LLG_JUL),
            'stp_jul':   _v(row, CI_STP_JUL),
            'vta_jul':   _v(row, CI_VTA_JUL),
            'cob_jul':   _v(row, CI_COB_JUL),
            'sti_ago':   _v(row, CI_STI_AGO),
            'cmp_ago':   _v(row, CI_CMP_AGO),
            'llg_ago':   _v(row, CI_LLG_AGO),
            'stp_ago':   _v(row, CI_STP_AGO),
            'vta_ago':   _v(row, CI_VTA_AGO),
            'cob_ago':   _v(row, CI_COB_AGO),
            'sti_sep':   _v(row, CI_STI_SEP),
            'cmp_sep':   _v(row, CI_CMP_SEP),
            'llg_sep':   _v(row, CI_LLG_SEP),
            'stp_sep':   _v(row, CI_STP_SEP),
            'vta_sep':   _v(row, CI_VTA_SEP),
            'cob_sep':   _v(row, CI_COB_SEP),
        }
        if marca.lower() in SKIP:
            rows_totals.append(rec)
        else:
            rows_data.append(rec)

    if not rows_data:
        return

    # ── Paleta de colores por mes ─────────────────────────────────────
    # Cada mes tiene un color de fondo para el encabezado de grupo
    MES_COLORS = {
        'ABR': ('154360', '1A5276'),   # azul oscuro
        'MAY': ('145A32', '196F3D'),   # verde oscuro
        'JUN': ('5B2C6F', '76448A'),   # morado
        'JUL': ('6E2F1A', '935116'),   # naranja oscuro
        'AGO': ('1B4F72', '21618C'),   # azul acero
        'SEP': ('4D5656', '616A6B'),   # gris oscuro
    }

    # Color cobertura (igual que el resto del análisis)
    def _cob_style(meses):
        if meses is None or meses == 0:  return ('F2D7D5', 'E74C3C')  # rojo
        if meses < 1:                    return ('F2D7D5', 'E74C3C')
        if meses < 2:                    return ('FDEBD0', 'E67E22')  # naranja
        if meses < 4:                    return ('D5F5E3', '1E8449')  # verde
        if meses < 6:                    return ('D6EAF8', '1A5276')  # azul
        return                                  ('E8DAEF', '6C3483')  # morado

    FMT_M   = '$ #,##0.0,,'   # muestra en millones con 1 decimal
    FMT_COB = '0.00'           # meses de cobertura

    # ── Layout de columnas ────────────────────────────────────────────
    # Bloques: [Marca | StkHoy | CobACT] | [ABR: Trans|Stp|Vta|CobAbr]
    #          | [MAY-SEP: StkIni|Cmp|Llg|Stp|Vta|Cob]
    # Total: 1 + 2 + 4 + 6*5 = 37 columnas

    COL_WIDTHS = (
        [('Marca', 18), ('Stock Hoy\nCST ($M)', 13), ('Cobert.\nACT', 9)]
        + [('Tránsito\nABR ($M)', 13), ('Stk+Ped\nABR ($M)', 13),
           ('Venta\nABR ($M)', 13), ('Cobert.\nABR', 9)]
        + [('StkIni\n($M)', 11), ('Compra\n($M)', 11), ('Llegadas\n($M)', 11),
           ('Stk+Ped\n($M)', 12), ('Venta\n($M)', 11), ('Cobert.', 9)] * 5
    )
    NCOLS   = len(COL_WIDTHS)
    end_col = get_column_letter(NCOLS)

    ws = owb.create_sheet('CST x Marca')
    ws.freeze_panes = 'B5'

    today     = _dt.date.today()
    año       = today.year

    # ── Fila 1: Título ────────────────────────────────────────────────
    ws.merge_cells(f'A1:{end_col}1')
    t = ws.cell(row=1, column=1,
                value=f'REPORTE CST x MARCA — COBERTURA Y FLUJO {año}  (valores en $M CLP)')
    t.font      = Font(bold=True, color='FFFFFF', name='Calibri', size=13)
    t.fill      = PatternFill('solid', fgColor='1B2631')
    t.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    # ── Fila 2: Encabezados de grupo (mes) ───────────────────────────
    # Columnas: A=Marca, B=StkHoy, C=CobACT, D-G=ABR, H-M=MAY, N-S=JUN,
    #           T-Y=JUL, Z-AE=AGO, AF-AK=SEP
    grupo_merge = [
        (1, 1, ''),
        (2, 3, 'STOCK HOY'),
        (4, 7, 'ABRIL 26'),
        (8, 13, 'MAYO 26'),
        (14, 19, 'JUNIO 26'),
        (20, 25, 'JULIO 26'),
        (26, 31, 'AGOSTO 26'),
        (32, 37, 'SEPTIEMBRE 26'),
    ]
    mes_keys = ['', '', 'ABR', 'MAY', 'JUN', 'JUL', 'AGO', 'SEP']

    ws.row_dimensions[2].height = 20
    for (c1, c2, label), mkey in zip(grupo_merge, mes_keys):
        if c1 == c2:
            ws.cell(row=2, column=c1).fill = PatternFill('solid', fgColor='1B2631')
        else:
            ws.merge_cells(start_row=2, start_column=c1,
                           end_row=2,   end_column=c2)
            dark, _ = MES_COLORS.get(mkey, ('1B2631', '1B2631'))
            c = ws.cell(row=2, column=c1, value=label)
            c.font      = Font(bold=True, color='FFFFFF', name='Calibri', size=10)
            c.fill      = PatternFill('solid', fgColor=dark)
            c.alignment = Alignment(horizontal='center', vertical='center')
            # Borde lateral para separar grupos
            side = Side(style='medium', color='FFFFFF')
            for col_idx in range(c1, c2 + 1):
                cell = ws.cell(row=2, column=col_idx)
                left  = Side(style='medium', color='FFFFFF') if col_idx == c1  else Side()
                right = Side(style='medium', color='FFFFFF') if col_idx == c2  else Side()
                cell.border = Border(left=left, right=right)

    # ── Fila 3: Sub-encabezados de columna ────────────────────────────
    # Color de fondo: versión más clara del mes
    sub_mes = ['', '', '', 'ABR', 'ABR', 'ABR', 'ABR',
               'MAY','MAY','MAY','MAY','MAY','MAY',
               'JUN','JUN','JUN','JUN','JUN','JUN',
               'JUL','JUL','JUL','JUL','JUL','JUL',
               'AGO','AGO','AGO','AGO','AGO','AGO',
               'SEP','SEP','SEP','SEP','SEP','SEP']

    ws.row_dimensions[3].height = 36
    for col_idx, ((label, width), mkey) in enumerate(zip(COL_WIDTHS, sub_mes), 1):
        _, light = MES_COLORS.get(mkey, ('1B2631', '2C3E50'))
        bg = light if mkey else '2C3E50'
        c = ws.cell(row=3, column=col_idx, value=label)
        c.font      = Font(bold=True, color='FFFFFF', name='Calibri', size=9)
        c.fill      = PatternFill('solid', fgColor=bg)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border    = borde
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Filas de datos ────────────────────────────────────────────────
    def _write_row(rn, rec, is_total=False, total_label=None):
        bg = 'F8F9FA' if rn % 2 == 0 else 'FFFFFF'
        if is_total:
            bg = '1B2631'

        fg_default = '000000' if not is_total else 'FFFFFF'
        fnt = Font(name='Calibri', size=9, bold=is_total, color=fg_default)

        def _wr(col, val, fmt=None, bg_ov=None, fg_ov=None, bold_ov=None):
            c = ws.cell(row=rn, column=col, value=val)
            c.font = Font(name='Calibri', size=9,
                          bold=bold_ov if bold_ov is not None else is_total,
                          color=fg_ov or fg_default)
            c.fill = PatternFill('solid', fgColor=bg_ov or bg)
            c.alignment = Alignment(
                horizontal='left'  if col == 1 else 'right' if isinstance(val, float) else 'center',
                vertical='center')
            c.border = borde
            if fmt: c.number_format = fmt

        marca_label = total_label or rec['marca']
        _wr(1, marca_label, bold_ov=True)
        _wr(2, rec['stk_hoy'],  FMT_M)

        # Cobertura ACT (col 3)
        cb_bg, cb_fg = _cob_style(rec['cob_act'])
        _wr(3, rec['cob_act'], FMT_COB,
            bg_ov=cb_bg if not is_total else bg,
            fg_ov=cb_fg if not is_total else fg_default)

        # ABR (cols 4-7)
        _wr(4, rec['trans_abr'], FMT_M)
        _wr(5, rec['stp_abr'],   FMT_M)
        _wr(6, rec['vta_abr'],   FMT_M)
        cb_bg, cb_fg = _cob_style(rec['cob_abr'])
        _wr(7, rec['cob_abr'],  FMT_COB,
            bg_ov=cb_bg if not is_total else bg,
            fg_ov=cb_fg if not is_total else fg_default)

        # MAY (cols 8-13)
        _wr(8,  rec['sti_may'], FMT_M)
        _wr(9,  rec['cmp_may'], FMT_M)
        _wr(10, rec['llg_may'], FMT_M)
        _wr(11, rec['stp_may'], FMT_M)
        _wr(12, rec['vta_may'], FMT_M)
        cb_bg, cb_fg = _cob_style(rec['cob_may'])
        _wr(13, rec['cob_may'], FMT_COB,
            bg_ov=cb_bg if not is_total else bg,
            fg_ov=cb_fg if not is_total else fg_default)

        # JUN (cols 14-19)
        _wr(14, rec['sti_jun'], FMT_M)
        _wr(15, rec['cmp_jun'], FMT_M)
        _wr(16, rec['llg_jun'], FMT_M)
        _wr(17, rec['stp_jun'], FMT_M)
        _wr(18, rec['vta_jun'], FMT_M)
        cb_bg, cb_fg = _cob_style(rec['cob_jun'])
        _wr(19, rec['cob_jun'], FMT_COB,
            bg_ov=cb_bg if not is_total else bg,
            fg_ov=cb_fg if not is_total else fg_default)

        # JUL (cols 20-25)
        _wr(20, rec['sti_jul'], FMT_M)
        _wr(21, rec['cmp_jul'], FMT_M)
        _wr(22, rec['llg_jul'], FMT_M)
        _wr(23, rec['stp_jul'], FMT_M)
        _wr(24, rec['vta_jul'], FMT_M)
        cb_bg, cb_fg = _cob_style(rec['cob_jul'])
        _wr(25, rec['cob_jul'], FMT_COB,
            bg_ov=cb_bg if not is_total else bg,
            fg_ov=cb_fg if not is_total else fg_default)

        # AGO (cols 26-31)
        _wr(26, rec['sti_ago'], FMT_M)
        _wr(27, rec['cmp_ago'], FMT_M)
        _wr(28, rec['llg_ago'], FMT_M)
        _wr(29, rec['stp_ago'], FMT_M)
        _wr(30, rec['vta_ago'], FMT_M)
        cb_bg, cb_fg = _cob_style(rec['cob_ago'])
        _wr(31, rec['cob_ago'], FMT_COB,
            bg_ov=cb_bg if not is_total else bg,
            fg_ov=cb_fg if not is_total else fg_default)

        # SEP (cols 32-37)
        _wr(32, rec['sti_sep'], FMT_M)
        _wr(33, rec['cmp_sep'], FMT_M)
        _wr(34, rec['llg_sep'], FMT_M)
        _wr(35, rec['stp_sep'], FMT_M)
        _wr(36, rec['vta_sep'], FMT_M)
        cb_bg, cb_fg = _cob_style(rec['cob_sep'])
        _wr(37, rec['cob_sep'], FMT_COB,
            bg_ov=cb_bg if not is_total else bg,
            fg_ov=cb_fg if not is_total else fg_default)

        ws.row_dimensions[rn].height = 16

    rn = 4
    for rec in rows_data:
        _write_row(rn, rec)
        rn += 1

    # Separador
    ws.row_dimensions[rn].height = 4
    for c in range(1, NCOLS + 1):
        ws.cell(row=rn, column=c).fill = PatternFill('solid', fgColor='1B2631')
    rn += 1

    # Filas de totales
    TOTAL_LABELS = {
        'total general':           'TOTAL PROPIA',
        'proveedores nacionales':  'PROV. NACIONALES',
        'total empresa':           'TOTAL EMPRESA',
    }
    for rec in rows_totals:
        label = TOTAL_LABELS.get(rec['marca'].lower(), rec['marca'].upper())
        _write_row(rn, rec, is_total=True, total_label=label)
        rn += 1

    ws.auto_filter.ref = f'A3:{end_col}{rn - len(rows_totals) - 2}'

    # Nota al pie
    ws.row_dimensions[rn].height = 14
    ws.merge_cells(f'A{rn}:{end_col}{rn}')
    nota = ws.cell(row=rn, column=1,
                   value='Valores en millones de pesos CLP ($M).  '
                         'Cobert. = Stk+Ped / Venta del mes.  '
                         'Colores: 🔴 <1m  🟠 1-2m  🟢 2-4m  🔵 4-6m  🟣 >6m')
    nota.font      = Font(italic=True, size=8, color='7F8C8D', name='Calibri')
    nota.alignment = Alignment(horizontal='left', vertical='center')


# ══════════════════════════════════════════════════════════════════════
# ENTRY POINT PRINCIPAL
# ══════════════════════════════════════════════════════════════════════
def run_analisis(log_fn=print):
    """
    Genera analisis_planificacion_ABR26.xlsx con 4 hojas:
    críticos (<1 mes) y sobrestock (>6 meses), CST + unidades.
    """
    import xlwings as xw

    # ── Leer datos ────────────────────────────────────────────────────
    log_fn("   Leyendo datos del Excel...")
    app = xw.App(visible=False)
    try:
        wb = app.books.open(EXCEL_PATH)
        ws_cst   = wb.sheets['REPORTE CST FLAT']
        ws_unid  = wb.sheets['REPORTE UNID FLAT']
        ws_fcst  = wb.sheets['FCST BASE SKU MACRO']
        ws_trans = wb.sheets['BASE TRANSITOS']
        lr  = ws_cst.used_range.last_cell.row
        lrf = ws_fcst.used_range.last_cell.row
        lrt = ws_trans.used_range.last_cell.row
        data_cst   = ws_cst.range((2,1),(lr,44)).value
        data_unid  = ws_unid.range((2,1),(lr,47)).value
        data_fcst  = ws_fcst.range((4,1),(lrf,213)).value
        # Leer encabezados de fila 1 para descubrimiento dinámico de columnas
        _lc_trans   = ws_trans.used_range.last_cell.column
        _trans_hdrs = ws_trans.range((1,1),(1,_lc_trans)).value
        if not isinstance(_trans_hdrs, list): _trans_hdrs = [_trans_hdrs]
        # Expandir lectura a todas las columnas disponibles (máx 30)
        _ncols_trans = min(_lc_trans, 30)
        data_trans  = ws_trans.range((14,1),(lrt,_ncols_trans)).value
        # Leer VTA X marca meta para la hoja de análisis mensual
        ws_vta   = wb.sheets['VTA X marca meta']
        lr_vta   = ws_vta.used_range.last_cell.row
        lc_vta   = ws_vta.used_range.last_cell.column
        data_vta = ws_vta.range((1,1),(lr_vta,lc_vta)).value
        # Leer TD REPORTES CST MP+N SKU para hoja CST x Marca
        ws_mp    = wb.sheets['TD REPORTES CST MP+N SKU']
        lr_mp    = ws_mp.used_range.last_cell.row
        lc_mp    = ws_mp.used_range.last_cell.column
        data_mp  = ws_mp.range((1,1),(lr_mp, min(lc_mp, 46))).value
        wb.close()
    finally:
        app.quit()

    # ── Índices ───────────────────────────────────────────────────────
    fcst_por_sku = {}
    for row in data_fcst:
        if row is None or row[FCST_IDX_SKU] is None: continue
        fcst_por_sku[_fmt_sku(row[FCST_IDX_SKU])] = row

    # ── Mapeo dinámico de columnas BASE TRANSITOS (por nombre de encabezado) ─
    _tc_map = {}
    for _i, _h in enumerate(_trans_hdrs or []):
        if _h: _tc_map[str(_h).strip().lower()] = _i

    def _tcol(name, fallback):
        v = _tc_map.get(name.lower(), None)
        return fallback if v is None else v

    TC_SKU      = _tcol('sku',              0)
    # índice 1 = Descripción (confirmado en datos); índice 2 = PI; índice 6 = Cantidad
    # índice 12 = ETA Bodega; índice 13 = MES (fórmula); índice 17 = MARCA (fórmula)
    TC_DESC     = (_tcol('descripcion',     None) or
                   _tcol('descripción',     None) or 1)   # idx 1 = descripción real
    TC_PI       = _tcol('pi',               2)
    TC_TRANSPORTE = (_tcol('transporte',    None) or
                     _tcol('via',           None) or 4)   # idx 4 ≈ 'BARCO' / modo transporte
    TC_ETA_CHL  = (_tcol('fecha eta chile', None) or
                   _tcol('eta chile',       None) or 11)
    TC_CANTIDAD = _tcol('cantidad',         6)
    TC_ETA_BOD  = (_tcol('fecha eta bodega', None) or
                   _tcol('eta bodega',      None) or 12)
    TC_MES      = _tcol('mes',              13)
    TC_STK_ACT  = (_tcol('stock actual',    None) or
                   _tcol('stk actual',      None) or 14)
    TC_TIPO_CAT = (_tcol('tipo categoria',  None) or
                   _tcol('tipo categoría',  None) or 15)
    TC_VALOR    = (_tcol('valor usd total', None) or
                   _tcol('valor usd',       None) or 16)
    TC_MARCA_T  = _tcol('marca',            17)

    def _sr(row, idx, default=''):
        """Safe read: devuelve default si idx está fuera de rango o el valor es None."""
        try:
            v = row[idx]
            return v if v is not None else default
        except (IndexError, TypeError):
            return default

    trans_por_sku = defaultdict(list)
    trans_rows_full = []   # datos completos para hojas de tránsito

    for row in data_trans:
        if row is None or _sr(row, TC_SKU) in (None, ''): continue
        raw_sku = _sr(row, TC_SKU)
        sk = _fmt_sku(raw_sku)
        if not sk: continue

        pi  = str(_sr(row, TC_PI)).strip()
        eta = _sr(row, TC_ETA_BOD, None)
        mes = _sr(row, TC_MES,     None)
        qty = flt(_sr(row, TC_CANTIDAD, 0))

        if pi or eta:
            trans_por_sku[sk].append({'pi':pi,'eta':eta,'mes':mes,'qty':qty})
            # Solo agregar a trans_rows_full si tiene datos de tránsito real
            trans_rows_full.append({
                'sku':       sk,
                'desc':      str(_sr(row, TC_DESC)).strip(),
                'pi':        pi,
                'transp':    str(_sr(row, TC_TRANSPORTE)).strip(),
                'eta_chl':   _sr(row, TC_ETA_CHL, None),
                'cantidad':  qty,
                'eta_bod':   eta,
                'mes':       mes,
                'stk_act':   flt(_sr(row, TC_STK_ACT, 0)),
                'tipo_cat':  str(_sr(row, TC_TIPO_CAT)).strip(),
                'valor_usd': flt(_sr(row, TC_VALOR,   0)),
                'marca':     str(_sr(row, TC_MARCA_T)).strip(),
            })

    # ── Construir dataset ─────────────────────────────────────────────
    CATS_EXCLUIR = {'(sin clasificar)', ''}
    all_rows = []
    for rc, ru in zip(data_cst, data_unid):
        if rc is None or rc[0] is None: continue
        marca = str(rc[0]).strip()
        if marca.lower().startswith('total '): continue
        cat_com = str(rc[3]).strip() if rc[3] else ''
        if cat_com in CATS_EXCLUIR or 'descontinua' in cat_com.lower(): continue

        cat_padre = str(rc[1]).strip() if rc[1] else ''
        cat_hijo  = str(rc[2]).strip() if rc[2] else ''
        sku  = _fmt_sku(rc[4])
        desc = str(rc[5]).strip() if rc[5] else ''
        cobert = flt(rc[CST_COBERT], default=None)
        if cobert is None: continue

        sku_norm = str(int(float(sku))) if sku.replace('.','',1).isdigit() else sku
        fcst_row = fcst_por_sku.get(sku) or fcst_por_sku.get(sku_norm)
        ranking   = fcst_row[FCST_IDX_RANKING] if fcst_row and fcst_row[FCST_IDX_RANKING] is not None else ''
        puerto    = str(fcst_row[FCST_IDX_PUERTO]).strip() if fcst_row and fcst_row[FCST_IDX_PUERTO] is not None else ''
        _sp_raw   = fcst_row[FCST_IDX_SKU_PADRE] if fcst_row else None
        if _sp_raw is not None:
            # Normalizar: si es número entero, quitar decimales (ej. 1566335721736.0 → '1566335721736')
            try:
                _sp_f = float(_sp_raw)
                sku_padre = str(int(_sp_f)) if _sp_f == int(_sp_f) else str(_sp_raw).strip()
            except (ValueError, TypeError):
                sku_padre = str(_sp_raw).strip()
        else:
            sku_padre = ''
        llegadas = {m: flt(fcst_row[TRANSITO_FCST[m]]) for m in MESES_ORDEN} \
                   if fcst_row else {m: 0 for m in MESES_ORDEN}

        mm_name = {4:'ABR26',5:'MAY26',6:'JUN26',7:'JUL26',8:'AGO26',9:'SEP26',10:'OCT26'}
        emb = trans_por_sku.get(sku_norm, [])

        # Si FCST BASE tiene 0 para un mes pero BASE TRANSITOS tiene qty → usar qty de BASE TRANSITOS
        if emb:
            trans_qty = defaultdict(float)
            for e in emb:
                try:
                    mn = mm_name.get(int(e['mes']))
                    if mn and e['qty'] > 0:
                        trans_qty[mn] += e['qty']
                except: pass
            for mn, qty in trans_qty.items():
                if llegadas.get(mn, 0) == 0:
                    llegadas[mn] = qty

        # prox desde FCST BASE: solo meses cuyo cutoff (día 10) aún no pasó
        prox = next((m for m in MESES_ORDEN if llegadas[m] > 0 and _is_valid_prox(m)), 'Sin embarque')

        if emb:
            def _s(e): return e['eta'] if isinstance(e['eta'],(_dt.datetime,_dt.date)) else _dt.datetime(2099,1,1)
            sorted_emb = sorted(emb, key=_s)
            # Buscar el embarque más próximo cuyo mes aún sea válido (cutoff no pasado)
            best = None
            for e in sorted_emb:
                try:
                    ep = mm_name.get(int(e['mes']), None)
                    if ep and _is_valid_prox(ep):
                        best = e
                        break
                except: pass
            # Si no hay embarque futuro válido, usar el primero de igual forma para mostrar PI/ETA
            if best is None:
                best = sorted_emb[0]
            pi_emb = best['pi']; eta_bod = best['eta']
            try:
                ep = mm_name.get(int(best['mes']), None)
                if ep and _is_valid_prox(ep):
                    prox = ep
            except: pass
        else:
            pi_emb = ''; eta_bod = None

        all_rows.append({
            'marca': marca, 'cat_com': cat_com,
            'cat_padre': cat_padre, 'cat_hijo': cat_hijo, 'ranking': ranking,
            'sku': sku, 'sku_padre': sku_padre, 'desc': desc,
            'cobert': round(cobert, 3),
            'stock_cst':  flt(rc[CST_STOCK]),  'venta_cst':  flt(rc[CST_VENTA]),
            'stock_unid': flt(ru[UNID_STOCK]), 'venta_unid': flt(ru[UNID_VENTA]),
            'llegadas': llegadas, 'prox': prox, 'pi': pi_emb, 'eta': eta_bod,
            'puerto': puerto,     # Puerto Origen desde FCST BASE col 19 (idx 18)
            'has_emb': bool(emb), # True si hay OC registrada en BASE TRANSITOS
            # VentaCst por mes para proyección de fecha de quiebre
            'venta_mes': {
                4: flt(rc[10]), 5: flt(rc[16]), 6: flt(rc[22]),
                7: flt(rc[28]), 8: flt(rc[34]),
            },
            # Stk+Ped por mes (StkIni + Llegadas confirmadas) — columna correcta para proyección
            # rc[9]=ABR, rc[15]=MAY, rc[21]=JUN, rc[27]=JUL, rc[33]=AGO
            'stk_ini_mes': {
                4: flt(rc[9]),  5: flt(rc[15]), 6: flt(rc[21]),
                7: flt(rc[27]), 8: flt(rc[33]),
            },
        })

    criticos      = [r for r in all_rows if r['cobert'] < 1]
    inquietantes  = [r for r in all_rows if r['cobert'] < 2]
    sobrestock    = [r for r in all_rows if r['cobert'] > 6]
    log_fn(f"   Criticos (<1 mes): {len(criticos)} SKUs | "
           f"Inquietantes (<2 mes): {len(inquietantes)} SKUs | "
           f"Sobrestock (>6 mes): {len(sobrestock)} SKUs")

    # ── Generar Excel ─────────────────────────────────────────────────
    res_c, _ = agrupar(criticos)
    res_s, _ = agrupar(sobrestock)

    owb = openpyxl.Workbook()
    owb.remove(owb.active)

    # ── Mapa cobertura amplio (todos los SKUs de CST, sin filtros) ───────
    # Incluye descontinuados, sin clasificar, sin venta — para lookup en tránsitos
    # Estructura: {sku: {'tipo': 'cobert'|'sin_venta', 'actual': val,
    #                    4: cob_abr, 5: cob_may, 6: cob_jun, 7: cob_jul, 8: cob_ago}}
    _CST_COB_MES    = {4: 11, 5: 17, 6: 23, 7: 29, 8: 35}   # mes_num → col_idx cobertura
    _CST_VENTA_IDXS = [10, 16, 22, 28, 34]                   # VentaCst ABR-MAY-JUN-JUL-AGO
    cob_map_full = {}
    for rc in data_cst:
        if rc is None or rc[0] is None: continue
        sk = _fmt_sku(rc[4])
        if not sk or sk == '0': continue
        # sin_venta solo si NO hay ventas en ningún mes (ABR–AGO)
        any_vta = any(flt(rc[i]) > 0 for i in _CST_VENTA_IDXS if len(rc) > i)
        tipo    = 'cobert' if any_vta else 'sin_venta'
        cob_act = rc[CST_COBERT]
        entry = {
            'tipo':   tipo,
            'actual': round(flt(cob_act), 2) if cob_act is not None else 0.0,
        }
        for mes_n, col_i in _CST_COB_MES.items():
            c = rc[col_i] if len(rc) > col_i else None
            entry[mes_n] = round(flt(c), 2) if c is not None else 0.0
        cob_map_full[sk] = entry

    # ── Alertar SKUs en tránsito sin cobertura (excluir descontinuados) ──
    skus_en_transito = {_fmt_sku(r['sku']) for r in trans_rows_full if r.get('sku')}
    for sk in sorted(skus_en_transito):
        if sk and sk not in cob_map_full:
            tipo = next((r.get('tipo_cat', '') or '' for r in trans_rows_full
                         if _fmt_sku(r.get('sku','')) == sk), '')
            if 'descontinua' not in tipo.lower():
                log_fn(f"   ALERTA cobertura: SKU {sk} en tránsito sin cobertura "
                       f"en REPORTE CST FLAT (tipo_cat: {tipo or '—'})")

    # ── HOJAS en orden definitivo ──────────────────────────────────────
    log_fn("   Generando hojas...")
    # 1. VTA x Marca MAY
    write_vta_marca_mes(owb, data_vta)
    # 2. CST x Marca
    write_cst_mp_n_sku(owb, data_mp)
    # 3. Critico x Marca
    write_resumen_marca(owb, criticos,
        f'CRÍTICO x MARCA — Cobertura < 1 mes | {_CUR_MES_YEAR}',
        C_ROJO, 'Critico x Marca',
        cob_color_critico,
        'Sin\nStock',
        lambda rows: sum(1 for r in rows if r['stock_cst'] <= 0),
        lambda v, bg: (C_ROJO_BG, C_ROJO, True) if v > 0 else (bg, '000000', False),
    )
    # 4. Detalle Critico
    write_detalle(owb, criticos,
        f'DETALLE CRITICOS — Cobertura < 1 mes | Costo + Unidades | {_CUR_MES_YEAR}',
        C_ROJO, 'Detalle Critico', cob_color_critico, prox_color_critico,
        lambda x: (x['marca'],x['cat_com'],x['prox']=='Sin embarque',x['cobert']),
        show_carga=True,
    )
    # 5. Capital Inmovilizado
    write_capital_inmovilizado(owb, sobrestock)
    # 6. Sobrestock x SKU Padre
    write_sobrestock_sku_padre(owb, sobrestock)
    # 7. Sobrestock c-Llegada
    write_sobrestock_con_llegada(owb, sobrestock)
    # 8. Tránsitos por Embarque
    write_transitos_embarque(owb, trans_rows_full, cob_map_full)
    # 9. Nuevos en Tránsito
    write_nuevos_transito(owb, trans_rows_full)
    _año_hoy = _dt.date.today().year
    log_fn(f"   Hojas generadas: VTA x Marca MAY {_año_hoy} | CST x Marca | "
           f"Critico x Marca | Detalle Critico | Capital Inmovilizado | "
           f"Sobrestock x SKU Padre | Sobrestock c-Llegada | "
           f"Tránsitos por Embarque | Nuevos en Tránsito")

    owb.save(OUT_PATH)
    log_fn(f"   Guardado: {OUT_PATH}")


# ── Ejecución directa ─────────────────────────────────────────────────
if __name__ == '__main__':
    sys.stdout.reconfigure(encoding='utf-8')
    run_analisis()
