"""
Crea / actualiza dos hojas de reporte CST en FORECAST FINAL SKU 26-27.xlsx:

  REPORTE CST FLAT  — tabla plana jerarquía completa, solo columnas COSTO, AutoFilter
  REPORTE CST PIVOT — tabla dinámica Excel, solo columnas COSTO, sin totales

Columnas replicadas del TD REPORTES CST MP+N SKU (solo costo, sin unidades):
  ABR26 : Llegadas Cst, Stock+Ped Csto, Venta Cst, Cobert.           (4 cols)
  MAY-AGO: Stock Ini Cst, Compra Cst, Llegadas Cst, Stk+Ped, Venta Cst, Cobert. (6 cols c/u)
  SEP-OCT: Stock Ini Cst, Llegadas Cst, Stk+Ped, Venta Cst, Cobert. (5 cols c/u, sin Compra)
"""

import sys
sys.stdout.reconfigure(encoding='utf-8')

import datetime as _fdt
import xlwings as xw
from openpyxl.utils import get_column_letter, column_index_from_string

# ── Ventana dinámica Cobert. ACT 26 ────────────────────────────────────────────
# Usa los 3 meses contiguos a partir del mes actual (ej: MAY→MAY+JUN+JUL)
_TODAY_MONTH = _fdt.date.today().month
_ALL_MONTH_KEYS = ['ABR26', 'MAY26', 'JUN26', 'JUL26', 'AGO26', 'SEP26', 'OCT26']
_BASE_IDX = min(max(0, _TODAY_MONTH - 4), len(_ALL_MONTH_KEYS) - 3)
_ACT_M1, _ACT_M2, _ACT_M3 = (_ALL_MONTH_KEYS[_BASE_IDX],
                               _ALL_MONTH_KEYS[_BASE_IDX + 1],
                               _ALL_MONTH_KEYS[_BASE_IDX + 2])

# ── Configuración ──────────────────────────────────────────────────────────────
EXCEL_PATH  = r'C:\Users\felip\Desktop\UNIONX\FORECAST FINAL SKU\FORECAST FINAL SKU 26-27 V2.xlsx'
TD_SHEET    = 'TD REPORTES CST MP+N SKU'
BASE_SHEET  = 'FCST BASE SKU MACRO'
FLAT_SHEET  = 'REPORTE CST FLAT'
PIVOT_SHEET = 'REPORTE CST PIVOT'
OLD_SHEET   = 'REPORTE CST FORMATO'

NEW_SHEET   = FLAT_SHEET   # alias para actualizar_reportes.py

BASE_HDR_ROW  = 3
BASE_DATA_ROW = 4
NEW_HDR_ROW   = 1
NEW_DATA_START = 2

N_FIXED = 8   # Marca, Cat Padre, Cat Hijo, Cat Comercial, SKU, Desc, Stock Hoy CST, Cobert. ACT 26

# ── Definición de bloques mensuales (solo costo) ───────────────────────────────
# Columna de Venta Cst ABR26 en el FLAT = col 11 (I=9 Llegadas, J=10 Stk+Ped, K=11 VentaCst)
# Cobert. ACT 26 (col 8) = G(StockHoy) / K(VentaCstABR)

# ABR26: (llegadas_ltr, stk_ped_ltr, venta_cst_ltr)
# V2 headers verificados (max col HE=213): DA=Llegadas Cst ABR26, DC=Stk+Ped Csto ABR26, CY=Venta Cst ABR26
ABR_COLS = ('DA', 'DC', 'CY')

# MAY-AGO: (mes, stk_ini, compra, llegadas, stk_ped, venta_cst)
# Headers verificados del FCST BASE V2 fila 3 (cols 1-213):
MAYO_AGO_BLOCKS = [
    ('MAY26', 'DK', 'DM', 'DN', 'DP', 'DL'),   # DK=StkIni Cst, DM=Compra Cst, DN=Llegadas Cst, DP=Stk+Ped Cst, DL=Venta Cst
    ('JUN26', 'DW', 'DY', 'DZ', 'EB', 'DX'),   # DW=StkIni Cst, DY=Compra Cst, DZ=Llegadas Cst, EB=Stk+Ped Cst, DX=Venta Cst
    ('JUL26', 'EI', 'EK', 'EL', 'EN', 'EJ'),   # EI=StkIni Cst, EK=Compra Cst, EL=Llegadas Cst, EN=Stk+Ped Cst, EJ=Venta Cst
    ('AGO26', 'EU', 'EW', 'EX', 'EZ', 'EV'),   # EU=StkIni Cst, EW=Compra Cst, EX=Llegadas Cst, EZ=Stk+Ped Cst, EV=Venta Cst
]

# SEP-OCT: (mes, stk_ini, llegadas, stk_ped, venta_cst)   ← sin Compra
SEP_OCT_BLOCKS = [
    ('SEP26', 'FG', 'FJ', 'FL', 'FH'),   # FG=StkIni Cst, FJ=Llegadas Cst, FL=Stk+Ped Cst, FH=Venta Cst
    ('OCT26', 'FS', 'FU', 'FW', 'FT'),   # FS=StkIni Cst, FU=Llegadas Cst, FW=Stk+Ped Cst, FT=Venta Cst
]

# Tamaños de bloque (para formateo de colores alternados)
# ABR=4, MAY-AGO=6 c/u, SEP-OCT=4 c/u (sin Cobert.)
BLOCK_SIZES  = [4, 6, 6, 6, 6, 4, 4]
BLOCK_LABELS = ['ABR26', 'MAY26', 'JUN26', 'JUL26', 'AGO26', 'SEP26', 'OCT26']

# Posiciones fijas (col 1-based) de VentaCst y Stk+Ped en el FLAT
# Usadas para construir fórmulas de cobertura con promedio 3 meses
# Layout: 8 fijas + ABR(4) + MAY(6) + JUN(6) + JUL(6) + AGO(6) + SEP(4) + OCT(4) = 44
_VC = {   # VentaCst por mes
    'ABR26': 11,   # K
    'MAY26': 17,   # Q
    'JUN26': 23,   # W
    'JUL26': 29,   # AC
    'AGO26': 35,   # AI
    'SEP26': 40,   # AN
    'OCT26': 44,   # AR
}
_SP = {   # Stock+Pedido Csto por mes
    'ABR26': 10,   # J
    'MAY26': 16,   # P
    'JUN26': 22,   # V
    'JUL26': 28,   # AB
    'AGO26': 34,   # AH
}

# Campos de valor para el PIVOT (col_letra_base, caption_display)
# ── Definición de bloques mensuales (unidades) ────────────────────────────────
UNID_SHEET = 'REPORTE UNID FLAT'

# Para cada mes: (mes, ppto_ltr, compra_ltr, llegadas_ltr, stk_ped_unid_ltr)
# Stock Inicial Unid: ABR usa Stock Hoy Unid (col K en base); MAY+ tienen col propia
# Headers verificados del FCST BASE fila 3:
UNID_ABR = ('ABR26', 'CU', 'CV', 'CW', 'DB')   # ppto=CU(Venta PPTO ABR26), compra=CV(Pedido/Llegadas ABR26), llegadas=CW(Transito ABR 26), stk_ped_unid=DB(Stk+Ped Unid ABR26)
UNID_MAY_AGO_BLOCKS = [
    # (mes, stk_ini_unid, ppto, compra, llegadas, stk_ped_unid)
    ('MAY26', 'DF', 'DG', 'DH', 'DJ', 'DO'),   # DF=StkIni, DG=Ppto(Venta PPTO MAY26), DH=Pedido/Llegadas MAY26, DJ=Embarcado Mayo, DO=Stk+Ped Unid
    ('JUN26', 'DS', 'DT', 'DU', 'DV', 'EA'),   # DS=StkIni, DT=Ppto(Venta PPTO JUN26), DU=Pedido/Llegadas JUN26, DV=Transito JUN 26, EA=Stk+Ped Unid
    ('JUL26', 'EE', 'EF', 'EG', 'EH', 'EM'),   # EE=StkIni, EF=Ppto(Venta PPTO JUL26), EG=Compra JUL 26, EH=Transito JUL 26, EM=Stk+Ped Unid
    ('AGO26', 'EQ', 'ER', 'ES', 'ET', 'EY'),   # EQ=StkIni, ER=Ppto(Venta PPTO AGO26), ES=Pedido/Llegadas AGO26, ET=Transito AGO 26, EY=Stk+Ped Unid
]
UNID_SEP_OCT_BLOCKS = [
    # (mes, stk_ini_unid, ppto, compra, llegadas, stk_ped_unid)
    ('SEP26', 'FC', 'FD', 'FE', 'FF', 'FK'),   # FC=StkIni, FD=Ppto(Venta PPTO SEP26), FE=Pedido/Llegadas SEP26, FF=Transito SEP 26, FK=Stk+Ped Unid
    ('OCT26', 'FO', 'FP', 'FQ', 'FR', 'FV'),   # FO=StkIni, FP=Ppto(Venta PPTO OCT26), FQ=Pedido/Llegadas OCT26, FR=Transito OCT 26, FV=Stk+Ped Unid
]
UNID_STK_HOY = 'K'   # col 11 en FCST BASE V2 = Stock HOY (unidades)

# Layout: 8 fijas + ABR(5) + MAY-AGO(6×4) + SEP-OCT(5×2) = 47 cols
# VentaPPTO positions (col 1-based) en REPORTE UNID FLAT
_VC_UNID = {
    'ABR26': 12,   # L  — col 8+1+1+1+1  = 12
    'MAY26': 18,   # R  — col 13+1+1+1+1+1+1 = wait, 12+1(cobert)+5 = 18? Let me recalc:
    # 8 fixed + 5 ABR(9..13) → MAY starts col 14: StkIni(14),Compra(15),Llegadas(16),StkPed(17),Venta(18)
    'JUN26': 24,   # X  — 19(cobert)+5JUN = 20..25, Venta=24
    'JUL26': 30,   # AD — 26..31, Venta=30
    'AGO26': 36,   # AJ — 32..37, Venta=36
    'SEP26': 42,   # AP — 38..42, Venta=42
    'OCT26': 47,   # AU — 43..47, Venta=47
}
_SP_UNID = {   # Stock+Pedido Unid positions
    'ABR26': 11,   # K
    'MAY26': 17,   # Q
    'JUN26': 23,   # W
    'JUL26': 29,   # AC
    'AGO26': 35,   # AI
}
BLOCK_SIZES_UNID  = [5, 6, 6, 6, 6, 5, 5]   # ABR=5(con Cobert), MAY-AGO=6, SEP-OCT=5(sin Cobert)

PIVOT_VAL_DEFS = [
    # ABR26 — sin StockIni (StockHoy es global), igual que TD
    ('MJ', 'Llegadas Cst ABR26'),
    ('ML', 'Stk+Ped Csto ABR26'),
    ('MH', 'Venta Cst ABR26'),
    # MAY26
    ('MU', 'Stk Ini Cst MAY26'),  ('MW', 'Compra Cst MAY26'),
    ('MX', 'Llegadas Cst MAY26'), ('MZ', 'Stk+Ped Csto MAY26'), ('MV', 'Venta Cst MAY26'),
    # JUN26
    ('NI', 'Stk Ini Cst JUN26'),  ('NK', 'Compra Cst JUN26'),
    ('NL', 'Llegadas Cst JUN26'), ('NN', 'Stk+Ped Csto JUN26'), ('NJ', 'Venta Cst JUN26'),
    # JUL26
    ('NW', 'Stk Ini Cst JUL26'),  ('NY', 'Compra Cst JUL26'),
    ('NZ', 'Llegadas Cst JUL26'), ('OB', 'Stk+Ped Csto JUL26'), ('NX', 'Venta Cst JUL26'),
    # AGO26
    ('OK', 'Stk Ini Cst AGO26'),  ('OM', 'Compra Cst AGO26'),
    ('ON', 'Llegadas Cst AGO26'), ('OP', 'Stk+Ped Csto AGO26'), ('OL', 'Venta Cst AGO26'),
    # SEP26 (sin Compra)
    ('OY', 'Stk Ini Cst SEP26'),
    ('PA', 'Llegadas Cst SEP26'), ('PC', 'Stk+Ped Csto SEP26'), ('OZ', 'Venta Cst SEP26'),
    # OCT26 (sin Compra)
    ('PL', 'Stk Ini Cst OCT26'),
    ('PN', 'Llegadas Cst OCT26'), ('PP', 'Stk+Ped Csto OCT26'), ('PM', 'Venta Cst OCT26'),
]

# ── Paleta ─────────────────────────────────────────────────────────────────────
def rgb(r, g, b):
    return int(r) + int(g) * 256 + int(b) * 65536

C_FLAT_HDR    = rgb(31,  78, 121)    # azul oscuro encabezado
C_FLAT_HDR_FG = rgb(255, 255, 255)   # blanco
C_DIM_DATA    = rgb(221, 235, 247)   # azul pálido — columnas dimensión
C_MONTH_1     = rgb(189, 215, 238)   # azul claro  — bloques mes pares
C_MONTH_2     = rgb(222, 235, 246)   # azul lavanda — bloques mes impares

CF_RED_BG    = rgb(255, 199, 206);  CF_RED_FG    = rgb(156,   0,   6)
CF_YELLOW_BG = rgb(255, 235, 156);  CF_YELLOW_FG = rgb(156,  87,   0)
CF_GREEN_BG  = rgb(198, 239, 206);  CF_GREEN_FG  = rgb(  0,  97,   0)
CF_ORANGE_BG = rgb(255, 192,   0);  CF_ORANGE_FG = rgb(120,  50,   0)

def log(msg):
    print(msg, flush=True)


# ══════════════════════════════════════════════════════════════════════════════
# FLAT — tabla plana (solo costo, espeja TD)
# ══════════════════════════════════════════════════════════════════════════════

def _cobert_formula(sp_col_num, vc1_col_num, vc2_col_num, vc3_col_num, r):
    """
    Cobertura = Stk+Ped / promedio(VentaCst mes, mes+1, mes+2)
    Si la suma de ventas es 0 devuelve "".
    """
    sp  = get_column_letter(sp_col_num)
    v1  = get_column_letter(vc1_col_num)
    v2  = get_column_letter(vc2_col_num)
    v3  = get_column_letter(vc3_col_num)
    return (f'=IFERROR(IF(({v1}{r}+{v2}{r}+{v3}{r})=0,"",'
            f'{sp}{r}/(({v1}{r}+{v2}{r}+{v3}{r})/3)),"")' )


def build_flat_columns():
    """
    Layout 44 columnas (solo costo):
      A-H   8 fijas: Marca, CatPadre, CatHijo, CatCom, SKU, Desc, StockHoy, Cobert.ACT
      I-L   4  ABR26: Llegadas, Stk+Ped, VentaCst, Cobert(prom ABR+MAY+JUN)
      M-R   6  MAY26: StklIni, Compra, Llegadas, Stk+Ped, VentaCst, Cobert(prom MAY+JUN+JUL)
      S-X   6  JUN26: StkIni, Compra, Llegadas, Stk+Ped, VentaCst, Cobert(prom JUN+JUL+AGO)
      Y-AD  6  JUL26: StkIni, Compra, Llegadas, Stk+Ped, VentaCst, Cobert(prom JUL+AGO+SEP)
      AE-AJ 6  AGO26: StkIni, Compra, Llegadas, Stk+Ped, VentaCst, Cobert(prom AGO+SEP+OCT)
      AK-AN 4  SEP26: StkIni, Llegadas, Stk+Ped, VentaCst  (sin Cobert.)
      AO-AR 4  OCT26: StkIni, Llegadas, Stk+Ped, VentaCst  (sin Cobert.)
    """
    BASE = f"'{BASE_SHEET}'"
    R    = BASE_DATA_ROW
    r    = NEW_DATA_START

    def bref(ltr):
        return f'=IFERROR({BASE}!{ltr}{R},"")'

    cols = []

    # ── Columnas fijas ────────────────────────────────────────────────────────
    cols.append(('Marca',               bref('B')))   # 1=A
    cols.append(('Categoria Padre',     bref('C')))   # 2=B
    cols.append(('Categoria Hijo',      bref('D')))   # 3=C
    cols.append(('Categoria Comercial', bref('H')))   # 4=D
    cols.append(('SKU',                 bref('E')))   # 5=E
    cols.append(('Descripcion',         bref('F')))   # 6=F
    cols.append(('Stock Hoy CST',       bref('O')))   # 7=G  (V2: O=Stock Hoy CST)

    # Cobert. ACT 26 = StockHoy(G=7) / prom(VentaCst mes_actual + mes+1 + mes+2)
    cols.append(('Cobert. ACT 26',
                 _cobert_formula(7, _VC[_ACT_M1], _VC[_ACT_M2], _VC[_ACT_M3], r)))  # 8=H

    # ── ABR26 (4 cols: I J K L) ───────────────────────────────────────────────
    llegadas_ltr, stk_ped_ltr_base, venta_cst_ltr_base = ABR_COLS
    cols.append(('Llegadas Cst ABR26',        bref(llegadas_ltr)))       # 9=I
    cols.append(('Stock + Pedido Csto ABR26', bref(stk_ped_ltr_base)))   # 10=J
    cols.append(('Venta Cst ABR26',           bref(venta_cst_ltr_base))) # 11=K
    cols.append(('Cobert. ABR26',
                 _cobert_formula(_SP['ABR26'], _VC['ABR26'], _VC['MAY26'], _VC['JUN26'], r)))  # 12=L

    # ── MAY26 (6 cols: M N O P Q R) ──────────────────────────────────────────
    mes, stk_ini, compra, llegadas, stk_ped, venta_cst = MAYO_AGO_BLOCKS[0]
    cols.append((f'Stock Inicial Cst {mes}',   bref(stk_ini)))   # 13=M
    cols.append((f'Compra Cst {mes}',          bref(compra)))    # 14=N
    cols.append((f'Llegadas Cst {mes}',        bref(llegadas)))  # 15=O
    cols.append((f'Stock + Pedido Csto {mes}', bref(stk_ped)))   # 16=P
    cols.append((f'Venta Cst {mes}',           bref(venta_cst))) # 17=Q
    cols.append((f'Cobert. {mes}',
                 _cobert_formula(_SP['MAY26'], _VC['MAY26'], _VC['JUN26'], _VC['JUL26'], r)))  # 18=R

    # ── JUN26 (6 cols: S T U V W X) ──────────────────────────────────────────
    mes, stk_ini, compra, llegadas, stk_ped, venta_cst = MAYO_AGO_BLOCKS[1]
    cols.append((f'Stock Inicial Cst {mes}',   bref(stk_ini)))   # 19=S
    cols.append((f'Compra Cst {mes}',          bref(compra)))    # 20=T
    cols.append((f'Llegadas Cst {mes}',        bref(llegadas)))  # 21=U
    cols.append((f'Stock + Pedido Csto {mes}', bref(stk_ped)))   # 22=V
    cols.append((f'Venta Cst {mes}',           bref(venta_cst))) # 23=W
    cols.append((f'Cobert. {mes}',
                 _cobert_formula(_SP['JUN26'], _VC['JUN26'], _VC['JUL26'], _VC['AGO26'], r)))  # 24=X

    # ── JUL26 (6 cols: Y Z AA AB AC AD) ──────────────────────────────────────
    mes, stk_ini, compra, llegadas, stk_ped, venta_cst = MAYO_AGO_BLOCKS[2]
    cols.append((f'Stock Inicial Cst {mes}',   bref(stk_ini)))   # 25=Y
    cols.append((f'Compra Cst {mes}',          bref(compra)))    # 26=Z
    cols.append((f'Llegadas Cst {mes}',        bref(llegadas)))  # 27=AA
    cols.append((f'Stock + Pedido Csto {mes}', bref(stk_ped)))   # 28=AB
    cols.append((f'Venta Cst {mes}',           bref(venta_cst))) # 29=AC
    cols.append((f'Cobert. {mes}',
                 _cobert_formula(_SP['JUL26'], _VC['JUL26'], _VC['AGO26'], _VC['SEP26'], r)))  # 30=AD

    # ── AGO26 (6 cols: AE AF AG AH AI AJ) ────────────────────────────────────
    mes, stk_ini, compra, llegadas, stk_ped, venta_cst = MAYO_AGO_BLOCKS[3]
    cols.append((f'Stock Inicial Cst {mes}',   bref(stk_ini)))   # 31=AE
    cols.append((f'Compra Cst {mes}',          bref(compra)))    # 32=AF
    cols.append((f'Llegadas Cst {mes}',        bref(llegadas)))  # 33=AG
    cols.append((f'Stock + Pedido Csto {mes}', bref(stk_ped)))   # 34=AH
    cols.append((f'Venta Cst {mes}',           bref(venta_cst))) # 35=AI
    cols.append((f'Cobert. {mes}',
                 _cobert_formula(_SP['AGO26'], _VC['AGO26'], _VC['SEP26'], _VC['OCT26'], r)))  # 36=AJ

    # ── SEP26 (4 cols: AK AL AM AN) — sin Cobert. ────────────────────────────
    mes, stk_ini, llegadas, stk_ped, venta_cst = SEP_OCT_BLOCKS[0]
    cols.append((f'Stock Inicial Cst {mes}',   bref(stk_ini)))   # 37=AK
    cols.append((f'Llegadas Cst {mes}',        bref(llegadas)))  # 38=AL
    cols.append((f'Stock + Pedido Csto {mes}', bref(stk_ped)))   # 39=AM
    cols.append((f'Venta Cst {mes}',           bref(venta_cst))) # 40=AN

    # ── OCT26 (4 cols: AO AP AQ AR) — sin Cobert. ────────────────────────────
    mes, stk_ini, llegadas, stk_ped, venta_cst = SEP_OCT_BLOCKS[1]
    cols.append((f'Stock Inicial Cst {mes}',   bref(stk_ini)))   # 41=AO
    cols.append((f'Llegadas Cst {mes}',        bref(llegadas)))  # 42=AP
    cols.append((f'Stock + Pedido Csto {mes}', bref(stk_ped)))   # 43=AQ
    cols.append((f'Venta Cst {mes}',           bref(venta_cst))) # 44=AR

    return cols


def create_reporte_cst_flat(wb_xw):
    ws_base = wb_xw.sheets[BASE_SHEET]
    base_last_row = ws_base.used_range.last_cell.row
    num_data_rows = base_last_row - BASE_DATA_ROW + 1
    log(f"[FLAT] FCST BASE: {num_data_rows} filas (fila {BASE_DATA_ROW}→{base_last_row})")

    cols        = build_flat_columns()
    headers     = [c[0] for c in cols]
    formulas_r2 = [c[1] for c in cols]
    total_cols  = len(cols)
    data_end_row = NEW_DATA_START + num_data_rows - 1
    log(f"[FLAT] {total_cols} columnas | filas datos hasta {data_end_row}")

    # Verificar posiciones clave
    assert len(cols) == 44, f"Esperaba 44 columnas, hay {len(cols)}"
    for mes, expected in _VC.items():
        hdr = f'Venta Cst {mes}'
        idx = next((i+1 for i,h in enumerate(headers) if h==hdr), None)
        assert idx == expected, f"'{hdr}' debería estar en col {expected}, está en {idx}"

    # ── Crear/limpiar hoja ────────────────────────────────────────────────────
    sheet_names = [s.name for s in wb_xw.sheets]
    if FLAT_SHEET in sheet_names:
        ws = wb_xw.sheets[FLAT_SHEET]
        ws.clear()
        log(f"[FLAT] Hoja '{FLAT_SHEET}' limpiada.")
    else:
        after = OLD_SHEET if OLD_SHEET in sheet_names else TD_SHEET
        wb_xw.sheets.add(FLAT_SHEET, after=wb_xw.sheets[after])
        ws = wb_xw.sheets[FLAT_SHEET]
        log(f"[FLAT] Hoja '{FLAT_SHEET}' creada.")

    # ── Encabezados + fórmulas ───────────────────────────────────────────────
    ws.range((NEW_HDR_ROW, 1)).value = [headers]
    ws.range((NEW_DATA_START, 1)).value = [formulas_r2]
    src = ws.range((NEW_DATA_START, 1), (NEW_DATA_START, total_cols))
    dst = ws.range((NEW_DATA_START, 1), (data_end_row,   total_cols))
    src.api.AutoFill(dst.api, 0)
    log("[FLAT] Fórmulas generadas.")

    # ── Anchos de columna ────────────────────────────────────────────────────
    # Fijas (A-H)
    fixed_widths = [22, 18, 16, 16, 12, 30, 14, 10]
    for i, w in enumerate(fixed_widths):
        try: ws.api.Columns(i + 1).ColumnWidth = w
        except: pass
    # ABR26 (4 cols): Llegadas, Stk+Ped, VentaCst, Cobert
    abr_widths = [14, 20, 14, 10]
    for j, w in enumerate(abr_widths):
        try: ws.api.Columns(N_FIXED + j + 1).ColumnWidth = w
        except: pass
    # MAY-AGO (6 cols): StokIni, Compra, Llegadas, Stk+Ped, VentaCst, Cobert
    mayo_ago_widths = [18, 16, 14, 20, 14, 10]
    for m_idx in range(4):   # 4 meses MAY-AGO
        start_col = N_FIXED + 4 + m_idx * 6 + 1   # ABR=4, luego 6 por mes
        for j, w in enumerate(mayo_ago_widths):
            try: ws.api.Columns(start_col + j).ColumnWidth = w
            except: pass
    # SEP-OCT (4 cols c/u, sin Cobert): StkIni, Llegadas, Stk+Ped, VentaCst
    sep_oct_widths = [18, 14, 20, 14]
    for m_idx in range(2):
        start_col = N_FIXED + 4 + 4 * 6 + m_idx * 4 + 1
        for j, w in enumerate(sep_oct_widths):
            try: ws.api.Columns(start_col + j).ColumnWidth = w
            except: pass

    # ── Congelar paneles en G2 ───────────────────────────────────────────────
    try:
        ws.activate()
        ws.range((NEW_DATA_START, 7)).select()
        wb_xw.api.Application.ActiveWindow.FreezePanes = True
    except Exception as e:
        log(f"   ADVERTENCIA freeze: {e}")

    # ── Formato encabezado (fila 1) ──────────────────────────────────────────
    log("[FLAT] Formateando encabezados y datos...")
    hdr_api = ws.range((NEW_HDR_ROW, 1), (NEW_HDR_ROW, total_cols)).api
    hdr_api.Interior.Color      = C_FLAT_HDR
    hdr_api.Font.Bold           = True
    hdr_api.Font.Color          = C_FLAT_HDR_FG
    hdr_api.Font.Size           = 10
    hdr_api.RowHeight           = 34
    hdr_api.VerticalAlignment   = -4108   # xlCenter
    hdr_api.HorizontalAlignment = -4108
    hdr_api.WrapText            = True

    # ── Color datos ───────────────────────────────────────────────────────────
    # Columnas fijas: azul pálido
    ws.range((NEW_DATA_START, 1), (data_end_row, N_FIXED)).api.Interior.Color = C_DIM_DATA

    # Columnas mensuales: alternar por bloque, con borde izquierdo
    cur_col = N_FIXED + 1
    for m_idx, (label, size) in enumerate(zip(BLOCK_LABELS, BLOCK_SIZES)):
        block_end = cur_col + size - 1
        color = C_MONTH_1 if m_idx % 2 == 0 else C_MONTH_2
        ws.range((NEW_DATA_START, cur_col), (data_end_row, block_end)).api.Interior.Color = color
        try:
            ws.range((NEW_HDR_ROW, cur_col), (data_end_row, cur_col)).api.Borders(7).Weight = 2
        except: pass
        cur_col = block_end + 1

    # ── Formato numérico ─────────────────────────────────────────────────────
    cobert_cols = []
    for i, hdr in enumerate(headers):
        col = i + 1
        if col <= N_FIXED:
            if hdr == 'Stock Hoy CST':
                ws.range((NEW_DATA_START, col), (data_end_row, col)).api.NumberFormat = '#.##0'
            elif 'COBERT' in hdr.upper():
                ws.range((NEW_DATA_START, col), (data_end_row, col)).api.NumberFormat = '#.##0,0'
                cobert_cols.append(col)
            continue
        rng = ws.range((NEW_DATA_START, col), (data_end_row, col)).api
        if 'COBERT' in hdr.upper():
            rng.NumberFormat = '#.##0,0'
            cobert_cols.append(col)
        else:
            rng.NumberFormat = '#.##0'
    log(f"[FLAT] Cobert. cols: {[get_column_letter(c) for c in cobert_cols]}")

    # ── Formato condicional en Cobert. ───────────────────────────────────────
    log(f"[FLAT] Aplicando CF en {len(cobert_cols)} columnas...")
    union_rng = None
    for c in cobert_cols:
        rng = ws.range((NEW_DATA_START, c), (data_end_row, c)).api
        union_rng = rng if union_rng is None else ws.api.Application.Union(union_rng, rng)

    if union_rng is not None:
        union_rng.FormatConditions.Delete()
        fc0 = union_rng.FormatConditions.Add(Type=1, Operator=3, Formula1='""')
        fc0.StopIfTrue = True
        fc = union_rng.FormatConditions.Add(Type=1, Operator=5, Formula1="6")
        fc.Interior.Color = CF_RED_BG;    fc.Font.Color = CF_RED_FG;    fc.StopIfTrue = False
        fc = union_rng.FormatConditions.Add(Type=1, Operator=1, Formula1="4", Formula2="5.9999")
        fc.Interior.Color = CF_YELLOW_BG; fc.Font.Color = CF_YELLOW_FG; fc.StopIfTrue = False
        fc = union_rng.FormatConditions.Add(Type=1, Operator=1, Formula1="2", Formula2="3.9999")
        fc.Interior.Color = CF_GREEN_BG;  fc.Font.Color = CF_GREEN_FG;  fc.StopIfTrue = False
        fc = union_rng.FormatConditions.Add(Type=1, Operator=6, Formula1="2")
        fc.Interior.Color = CF_ORANGE_BG; fc.Font.Color = CF_ORANGE_FG; fc.StopIfTrue = False
        log("[FLAT] CF aplicado.")

    # ── ListObject (solo AutoFilter, sin slicers ni totales) ─────────────────
    log("[FLAT] Creando ListObject...")
    try:
        ws.api.ListObjects("TblReporteCstFlat").Delete()
    except: pass
    tbl_rng = ws.range((NEW_HDR_ROW, 1), (data_end_row, total_cols))
    tbl = ws.api.ListObjects.Add(SourceType=1, Source=tbl_rng.api, XlListObjectHasHeaders=1)
    tbl.Name                        = "TblReporteCstFlat"
    tbl.TableStyle                  = "TableStyleLight1"
    tbl.ShowTotals                  = False   # sin fila de totales
    tbl.ShowTableStyleRowStripes    = False
    tbl.ShowTableStyleColumnStripes = False
    tbl.ShowTableStyleFirstColumn   = False
    tbl.ShowTableStyleLastColumn    = False
    log(f"[FLAT] '{FLAT_SHEET}' completado. {total_cols} cols × {num_data_rows} filas.")


# ══════════════════════════════════════════════════════════════════════════════
# PIVOT — tabla dinámica (solo costo, sin totales)
# ══════════════════════════════════════════════════════════════════════════════

def _read_base_headers(wb_xw):
    ws_base  = wb_xw.sheets[BASE_SHEET]
    MAX_SCAN = 500
    vals = ws_base.range((BASE_HDR_ROW, 1), (BASE_HDR_ROW, MAX_SCAN)).value
    if vals and isinstance(vals[0], list):
        vals = vals[0]
    hdr_map      = {}
    last_hdr_col = 0
    for i, v in enumerate(vals):
        if v is not None and str(v).strip() != '':
            hdr_map[get_column_letter(i + 1)] = str(v)   # SIN strip — preserva espacios para PivotFields
            last_hdr_col = i + 1
    last_data_row = ws_base.used_range.last_cell.row
    return hdr_map, last_data_row, last_hdr_col


def create_reporte_cst_pivot(wb_xw):
    """
    PivotTable desde REPORTE CST FLAT (tabla TblReporteCstFlat).
    Al sourcer desde el FLAT evitamos los nombres duplicados de FCST BASE que
    corrompen el PivotCache.
    Jerarquía: Marca > Categoria Padre > Categoria Hijo > Categoria Comercial > SKU > Descripcion
    Valores: solo costo (todos los campos del FLAT menos Cobert. y Stock Hoy CST).
    Sin RowGrand ni ColumnGrand.
    """
    # Leer encabezados del FLAT para saber qué campos hay
    ws_flat = wb_xw.sheets[FLAT_SHEET]
    flat_last_row = ws_flat.used_range.last_cell.row
    flat_last_col = ws_flat.used_range.last_cell.column
    flat_last_col_ltr = get_column_letter(flat_last_col)
    src_str = f"'{FLAT_SHEET}'!A{NEW_HDR_ROW}:{flat_last_col_ltr}{flat_last_row}"
    log(f"[PIVOT] Fuente: {src_str}")

    # Leer encabezados del FLAT (fila 1)
    flat_headers = ws_flat.range((NEW_HDR_ROW, 1), (NEW_HDR_ROW, flat_last_col)).value
    if flat_headers and isinstance(flat_headers[0], list):
        flat_headers = flat_headers[0]
    log(f"[PIVOT] {len(flat_headers)} columnas en FLAT.")

    # Dimensiones (jerarquía)
    dim_names = ['Marca', 'Categoria Padre', 'Categoria Hijo',
                 'Categoria Comercial', 'SKU', 'Descripcion']

    # Campos de valor: todas las cols del FLAT que NO son dimensión ni Cobert. ni Stock Hoy
    skip_set = set(dim_names) | {'Cobert. ACT 26', 'Stock Hoy CST'}
    val_names = [h for h in flat_headers
                 if h and h not in skip_set and 'COBERT' not in str(h).upper()]
    log(f"[PIVOT] {len(dim_names)} dimensiones | {len(val_names)} campos valor.")

    # ── Crear/recrear hoja ────────────────────────────────────────────────────
    sheet_names = [s.name for s in wb_xw.sheets]
    if PIVOT_SHEET in sheet_names:
        wb_xw.sheets[PIVOT_SHEET].delete()
    after = FLAT_SHEET if FLAT_SHEET in [s.name for s in wb_xw.sheets] else TD_SHEET
    wb_xw.sheets.add(PIVOT_SHEET, after=wb_xw.sheets[after])
    ws_piv = wb_xw.sheets[PIVOT_SHEET]
    log(f"[PIVOT] Hoja '{PIVOT_SHEET}' creada.")

    # ── PivotCache desde FLAT ─────────────────────────────────────────────────
    try:
        pc = wb_xw.api.PivotCaches().Create(SourceType=1, SourceData=src_str)
    except Exception as e:
        log(f"[PIVOT] ERROR PivotCache: {e}"); raise

    # ── PivotTable en A3 ─────────────────────────────────────────────────────
    try:
        pt = pc.CreatePivotTable(
            TableDestination=ws_piv.range('A3').api,
            TableName='PivotReporteCst'
        )
    except Exception as e:
        log(f"[PIVOT] ERROR CreatePivotTable: {e}"); raise
    log("[PIVOT] PivotTable 'PivotReporteCst' creada.")

    xlRowField = 1
    xlSum      = -4157

    # Campos de fila (jerarquía)
    for fname in dim_names:
        try:
            pt.PivotFields(fname).Orientation = xlRowField
            log(f"   Fila: '{fname}'")
        except Exception as e:
            log(f"   ADVERTENCIA campo fila '{fname}': {e}")

    # Campos de valor (solo costo, sin Cobert.)
    # Usamos Orientation + Function directamente (más robusto que AddDataField con PivotCache de fórmulas)
    xlDataField = 4
    for fname in val_names:
        try:
            pf = pt.PivotFields(fname)
            pf.Orientation = xlDataField
            pf.Function    = xlSum
            pf.NumberFormat = '#.##0'
            log(f"   Valor: '{fname}'")
        except Exception as e:
            log(f"   ADVERTENCIA campo valor '{fname}': {e}")

    # ── Estilo y opciones — SIN totales ──────────────────────────────────────
    try: pt.TableStyle2 = 'PivotStyleMedium4'
    except: pass
    try: pt.ShowDrillIndicators = True
    except: pass
    try: pt.ColumnGrand = False   # sin totales de columna
    except: pass
    try: pt.RowGrand    = False   # sin totales de fila
    except: pass
    try: pt.RepeatAllLabels(2)    # xlRepeatLabels
    except: pass
    try: pt.NullString = "0"
    except: pass

    # Título
    try:
        ws_piv.range('A1').value = 'REPORTE CST — Tabla Dinámica por Marca / Categoría / SKU (solo costo)'
        ws_piv.range('A1').api.Font.Bold  = True
        ws_piv.range('A1').api.Font.Size  = 12
        ws_piv.range('A1').api.Font.Color = rgb(31, 78, 121)
    except: pass

    log(f"[PIVOT] '{PIVOT_SHEET}' completado.")


# ══════════════════════════════════════════════════════════════════════════════
# UNID — tabla plana en unidades
# ══════════════════════════════════════════════════════════════════════════════

def build_unid_columns():
    """
    47 columnas — misma estructura que CST pero en unidades:
      A-H   8 fijas: Marca, CatPadre, CatHijo, CatCom, SKU, Desc, Stock Hoy Unid, Cobert.ACT
      I-M   5  ABR26 (con Cobert): Compra, Llegadas, Stk+Ped Unid, Venta PPTO, Cobert
      N-S   6  MAY26 (con Cobert): StkIni, Compra, Llegadas, Stk+Ped Unid, Venta PPTO, Cobert
      T-Y   6  JUN26 ...
      Z-AE  6  JUL26 ...
      AF-AK 6  AGO26 ...
      AL-AP 5  SEP26 (sin Cobert): StkIni, Compra, Llegadas, Stk+Ped Unid, Venta PPTO
      AQ-AU 5  OCT26 (sin Cobert)
    Cobert = Stk+Ped Unid / prom(VentaPPTO mes, mes+1, mes+2) — solo ABR-AGO
    Cobert ACT 26 = StockHoyUnid / prom(VentaPPTO ABR+MAY+JUN)
    """
    BASE = f"'{BASE_SHEET}'"
    R    = BASE_DATA_ROW
    r    = NEW_DATA_START

    def bref(ltr):
        return f'=IFERROR({BASE}!{ltr}{R},"")'

    cols = []

    # ── Columnas fijas ────────────────────────────────────────────────────────
    cols.append(('Marca',               bref('B')))          # 1=A
    cols.append(('Categoria Padre',     bref('C')))          # 2=B
    cols.append(('Categoria Hijo',      bref('D')))          # 3=C
    cols.append(('Categoria Comercial', bref('H')))          # 4=D
    cols.append(('SKU',                 bref('E')))          # 5=E
    cols.append(('Descripcion',         bref('F')))          # 6=F
    cols.append(('Stock Hoy Unid',      bref(UNID_STK_HOY))) # 7=G

    # Cobert. ACT 26 = StockHoyUnid(G=7) / prom(VentaPPTO mes_actual + mes+1 + mes+2)
    cols.append(('Cobert. ACT 26',
                 _cobert_formula(7, _VC_UNID[_ACT_M1], _VC_UNID[_ACT_M2], _VC_UNID[_ACT_M3], r)))  # 8=H

    # ── ABR26 (5 cols: I J K L M) ────────────────────────────────────────────
    mes, ppto, compra, llegadas, stk_ped = UNID_ABR
    cols.append((f'Compra Unid {mes}',           bref(compra)))    # 9=I
    cols.append((f'Llegadas Unid {mes}',         bref(llegadas)))  # 10=J
    cols.append((f'Stock + Pedido Unid {mes}',   bref(stk_ped)))   # 11=K
    cols.append((f'Venta PPTO {mes}',            bref(ppto)))      # 12=L
    cols.append(('Cobert. ABR26',
                 _cobert_formula(_SP_UNID['ABR26'], _VC_UNID['ABR26'], _VC_UNID['MAY26'], _VC_UNID['JUN26'], r)))  # 13=M

    # ── MAY26 (6 cols: N..S) ─────────────────────────────────────────────────
    mes, stk_ini, ppto, compra, llegadas, stk_ped = UNID_MAY_AGO_BLOCKS[0]
    cols.append((f'Stock Inicial Unid {mes}',    bref(stk_ini)))   # 14=N
    cols.append((f'Compra Unid {mes}',           bref(compra)))    # 15=O
    cols.append((f'Llegadas Unid {mes}',         bref(llegadas)))  # 16=P
    cols.append((f'Stock + Pedido Unid {mes}',   bref(stk_ped)))   # 17=Q
    cols.append((f'Venta PPTO {mes}',            bref(ppto)))      # 18=R
    cols.append((f'Cobert. {mes}',
                 _cobert_formula(_SP_UNID['MAY26'], _VC_UNID['MAY26'], _VC_UNID['JUN26'], _VC_UNID['JUL26'], r)))  # 19=S

    # ── JUN26 (6 cols: T..Y) ─────────────────────────────────────────────────
    mes, stk_ini, ppto, compra, llegadas, stk_ped = UNID_MAY_AGO_BLOCKS[1]
    cols.append((f'Stock Inicial Unid {mes}',    bref(stk_ini)))   # 20=T
    cols.append((f'Compra Unid {mes}',           bref(compra)))    # 21=U
    cols.append((f'Llegadas Unid {mes}',         bref(llegadas)))  # 22=V
    cols.append((f'Stock + Pedido Unid {mes}',   bref(stk_ped)))   # 23=W
    cols.append((f'Venta PPTO {mes}',            bref(ppto)))      # 24=X
    cols.append((f'Cobert. {mes}',
                 _cobert_formula(_SP_UNID['JUN26'], _VC_UNID['JUN26'], _VC_UNID['JUL26'], _VC_UNID['AGO26'], r)))  # 25=Y

    # ── JUL26 (6 cols: Z..AE) ────────────────────────────────────────────────
    mes, stk_ini, ppto, compra, llegadas, stk_ped = UNID_MAY_AGO_BLOCKS[2]
    cols.append((f'Stock Inicial Unid {mes}',    bref(stk_ini)))   # 26=Z
    cols.append((f'Compra Unid {mes}',           bref(compra)))    # 27=AA
    cols.append((f'Llegadas Unid {mes}',         bref(llegadas)))  # 28=AB
    cols.append((f'Stock + Pedido Unid {mes}',   bref(stk_ped)))   # 29=AC
    cols.append((f'Venta PPTO {mes}',            bref(ppto)))      # 30=AD
    cols.append((f'Cobert. {mes}',
                 _cobert_formula(_SP_UNID['JUL26'], _VC_UNID['JUL26'], _VC_UNID['AGO26'], _VC_UNID['SEP26'], r)))  # 31=AE

    # ── AGO26 (6 cols: AF..AK) ───────────────────────────────────────────────
    mes, stk_ini, ppto, compra, llegadas, stk_ped = UNID_MAY_AGO_BLOCKS[3]
    cols.append((f'Stock Inicial Unid {mes}',    bref(stk_ini)))   # 32=AF
    cols.append((f'Compra Unid {mes}',           bref(compra)))    # 33=AG
    cols.append((f'Llegadas Unid {mes}',         bref(llegadas)))  # 34=AH
    cols.append((f'Stock + Pedido Unid {mes}',   bref(stk_ped)))   # 35=AI
    cols.append((f'Venta PPTO {mes}',            bref(ppto)))      # 36=AJ
    cols.append((f'Cobert. {mes}',
                 _cobert_formula(_SP_UNID['AGO26'], _VC_UNID['AGO26'], _VC_UNID['SEP26'], _VC_UNID['OCT26'], r)))  # 37=AK

    # ── SEP26 (5 cols: AL..AP, sin Cobert) ───────────────────────────────────
    mes, stk_ini, ppto, compra, llegadas, stk_ped = UNID_SEP_OCT_BLOCKS[0]
    cols.append((f'Stock Inicial Unid {mes}',    bref(stk_ini)))   # 38=AL
    cols.append((f'Compra Unid {mes}',           bref(compra)))    # 39=AM
    cols.append((f'Llegadas Unid {mes}',         bref(llegadas)))  # 40=AN
    cols.append((f'Stock + Pedido Unid {mes}',   bref(stk_ped)))   # 41=AO
    cols.append((f'Venta PPTO {mes}',            bref(ppto)))      # 42=AP

    # ── OCT26 (5 cols: AQ..AU, sin Cobert) ───────────────────────────────────
    mes, stk_ini, ppto, compra, llegadas, stk_ped = UNID_SEP_OCT_BLOCKS[1]
    cols.append((f'Stock Inicial Unid {mes}',    bref(stk_ini)))   # 43=AQ
    cols.append((f'Compra Unid {mes}',           bref(compra)))    # 44=AR
    cols.append((f'Llegadas Unid {mes}',         bref(llegadas)))  # 45=AS
    cols.append((f'Stock + Pedido Unid {mes}',   bref(stk_ped)))   # 46=AT
    cols.append((f'Venta PPTO {mes}',            bref(ppto)))      # 47=AU

    return cols


def create_reporte_unid_flat(wb_xw):
    ws_base = wb_xw.sheets[BASE_SHEET]
    base_last_row = ws_base.used_range.last_cell.row
    num_data_rows = base_last_row - BASE_DATA_ROW + 1

    cols        = build_unid_columns()
    headers     = [c[0] for c in cols]
    formulas_r2 = [c[1] for c in cols]
    total_cols  = len(cols)
    data_end_row = NEW_DATA_START + num_data_rows - 1

    # Verificar posiciones clave
    assert total_cols == 47, f"Esperaba 47 cols, hay {total_cols}"
    for mes, expected in _VC_UNID.items():
        hdr = f'Venta PPTO {mes}'
        idx = next((i+1 for i, h in enumerate(headers) if h == hdr), None)
        assert idx == expected, f"'{hdr}' debería estar en col {expected}, está en {idx}"
    log(f"[UNID] {total_cols} columnas | {num_data_rows} filas | hasta col {get_column_letter(total_cols)}")

    # ── Crear/limpiar hoja ────────────────────────────────────────────────────
    sheet_names = [s.name for s in wb_xw.sheets]
    if UNID_SHEET in sheet_names:
        ws = wb_xw.sheets[UNID_SHEET]
        ws.clear()
        log(f"[UNID] Hoja '{UNID_SHEET}' limpiada.")
    else:
        after_sheet = FLAT_SHEET if FLAT_SHEET in sheet_names else TD_SHEET
        wb_xw.sheets.add(UNID_SHEET, after=wb_xw.sheets[after_sheet])
        ws = wb_xw.sheets[UNID_SHEET]
        log(f"[UNID] Hoja '{UNID_SHEET}' creada.")

    # ── Encabezados + fórmulas ───────────────────────────────────────────────
    ws.range((NEW_HDR_ROW, 1)).value = [headers]
    ws.range((NEW_DATA_START, 1)).value = [formulas_r2]
    src = ws.range((NEW_DATA_START, 1), (NEW_DATA_START, total_cols))
    dst = ws.range((NEW_DATA_START, 1), (data_end_row,   total_cols))
    src.api.AutoFill(dst.api, 0)
    log("[UNID] Fórmulas generadas.")

    # ── Anchos de columna ────────────────────────────────────────────────────
    fixed_widths = [22, 18, 16, 16, 12, 30, 14, 10]
    for i, w in enumerate(fixed_widths):
        try: ws.api.Columns(i + 1).ColumnWidth = w
        except: pass
    # ABR (5 cols): Compra, Llegadas, Stk+Ped, VentaPPTO, Cobert
    abr_w = [14, 14, 18, 14, 10]
    for j, w in enumerate(abr_w):
        try: ws.api.Columns(N_FIXED + j + 1).ColumnWidth = w
        except: pass
    # MAY-AGO (6 cols c/u): StkIni, Compra, Llegadas, Stk+Ped, VentaPPTO, Cobert
    mayo_ago_w = [18, 14, 14, 18, 14, 10]
    for m_idx in range(4):
        sc = N_FIXED + 5 + m_idx * 6 + 1
        for j, w in enumerate(mayo_ago_w):
            try: ws.api.Columns(sc + j).ColumnWidth = w
            except: pass
    # SEP-OCT (5 cols c/u, sin Cobert): StkIni, Compra, Llegadas, Stk+Ped, VentaPPTO
    sep_oct_w = [18, 14, 14, 18, 14]
    for m_idx in range(2):
        sc = N_FIXED + 5 + 4 * 6 + m_idx * 5 + 1
        for j, w in enumerate(sep_oct_w):
            try: ws.api.Columns(sc + j).ColumnWidth = w
            except: pass

    # ── Congelar paneles en G2 ───────────────────────────────────────────────
    try:
        ws.activate()
        ws.range((NEW_DATA_START, 7)).select()
        wb_xw.api.Application.ActiveWindow.FreezePanes = True
    except: pass

    # ── Formato encabezado ───────────────────────────────────────────────────
    log("[UNID] Formateando...")
    hdr_api = ws.range((NEW_HDR_ROW, 1), (NEW_HDR_ROW, total_cols)).api
    hdr_api.Interior.Color      = C_FLAT_HDR
    hdr_api.Font.Bold           = True
    hdr_api.Font.Color          = C_FLAT_HDR_FG
    hdr_api.Font.Size           = 10
    hdr_api.RowHeight           = 34
    hdr_api.VerticalAlignment   = -4108
    hdr_api.HorizontalAlignment = -4108
    hdr_api.WrapText            = True

    # Fijas: azul pálido
    ws.range((NEW_DATA_START, 1), (data_end_row, N_FIXED)).api.Interior.Color = C_DIM_DATA

    # Mensuales: alternar colores por bloque
    cur_col = N_FIXED + 1
    for m_idx, (label, size) in enumerate(zip(BLOCK_LABELS, BLOCK_SIZES_UNID)):
        block_end = cur_col + size - 1
        color = C_MONTH_1 if m_idx % 2 == 0 else C_MONTH_2
        ws.range((NEW_DATA_START, cur_col), (data_end_row, block_end)).api.Interior.Color = color
        try:
            ws.range((NEW_HDR_ROW, cur_col), (data_end_row, cur_col)).api.Borders(7).Weight = 2
        except: pass
        cur_col = block_end + 1

    # ── Formato numérico ─────────────────────────────────────────────────────
    cobert_cols = []
    for i, hdr in enumerate(headers):
        col = i + 1
        rng = ws.range((NEW_DATA_START, col), (data_end_row, col)).api
        if 'COBERT' in hdr.upper():
            rng.NumberFormat = '#.##0,0'
            cobert_cols.append(col)
        elif col > N_FIXED:
            rng.NumberFormat = '#.##0'
        elif hdr in ('Stock Hoy Unid',):
            rng.NumberFormat = '#.##0'
    log(f"[UNID] Cobert. cols: {[get_column_letter(c) for c in cobert_cols]}")

    # ── Formato condicional en Cobert. ───────────────────────────────────────
    union_rng = None
    for c in cobert_cols:
        rng = ws.range((NEW_DATA_START, c), (data_end_row, c)).api
        union_rng = rng if union_rng is None else ws.api.Application.Union(union_rng, rng)
    if union_rng is not None:
        union_rng.FormatConditions.Delete()
        fc0 = union_rng.FormatConditions.Add(Type=1, Operator=3, Formula1='""')
        fc0.StopIfTrue = True
        fc = union_rng.FormatConditions.Add(Type=1, Operator=5, Formula1="6")
        fc.Interior.Color = CF_RED_BG;    fc.Font.Color = CF_RED_FG;    fc.StopIfTrue = False
        fc = union_rng.FormatConditions.Add(Type=1, Operator=1, Formula1="4", Formula2="5.9999")
        fc.Interior.Color = CF_YELLOW_BG; fc.Font.Color = CF_YELLOW_FG; fc.StopIfTrue = False
        fc = union_rng.FormatConditions.Add(Type=1, Operator=1, Formula1="2", Formula2="3.9999")
        fc.Interior.Color = CF_GREEN_BG;  fc.Font.Color = CF_GREEN_FG;  fc.StopIfTrue = False
        fc = union_rng.FormatConditions.Add(Type=1, Operator=6, Formula1="2")
        fc.Interior.Color = CF_ORANGE_BG; fc.Font.Color = CF_ORANGE_FG; fc.StopIfTrue = False
        log("[UNID] CF aplicado.")

    # ── ListObject (AutoFilter, sin slicers ni totales) ───────────────────────
    try:
        ws.api.ListObjects("TblReporteUnidFlat").Delete()
    except: pass
    tbl_rng = ws.range((NEW_HDR_ROW, 1), (data_end_row, total_cols))
    tbl = ws.api.ListObjects.Add(SourceType=1, Source=tbl_rng.api, XlListObjectHasHeaders=1)
    tbl.Name                        = "TblReporteUnidFlat"
    tbl.TableStyle                  = "TableStyleLight1"
    tbl.ShowTotals                  = False
    tbl.ShowTableStyleRowStripes    = False
    tbl.ShowTableStyleColumnStripes = False
    tbl.ShowTableStyleFirstColumn   = False
    tbl.ShowTableStyleLastColumn    = False
    log(f"[UNID] '{UNID_SHEET}' completado. {total_cols} cols × {num_data_rows} filas.")


# ══════════════════════════════════════════════════════════════════════════════
# Entry point
# ══════════════════════════════════════════════════════════════════════════════

def create_reporte_cst_formato(wb_xw):
    """Llamado desde actualizar_reportes.py."""
    sheet_names = [s.name for s in wb_xw.sheets]
    if OLD_SHEET in sheet_names:
        try:
            wb_xw.sheets[OLD_SHEET].delete()
            log(f"Hoja antigua '{OLD_SHEET}' eliminada.")
        except Exception as e:
            log(f"   ADVERTENCIA al eliminar '{OLD_SHEET}': {e}")

    create_reporte_cst_flat(wb_xw)
    create_reporte_unid_flat(wb_xw)

    # Eliminar hoja PIVOT si quedó de versiones anteriores
    sheet_names = [s.name for s in wb_xw.sheets]
    if PIVOT_SHEET in sheet_names:
        try:
            wb_xw.sheets[PIVOT_SHEET].delete()
            log(f"Hoja '{PIVOT_SHEET}' eliminada.")
        except Exception as e:
            log(f"   ADVERTENCIA al eliminar '{PIVOT_SHEET}': {e}")


def main():
    log("=" * 60)
    log("Abriendo archivo con xlwings...")
    app = xw.App(visible=False)
    app.display_alerts = False

    try:
        wb = app.books.open(EXCEL_PATH)

        create_reporte_cst_formato(wb)

        log("Refrescando pivot TD fuente...")
        try:
            pts = wb.sheets[TD_SHEET].api.PivotTables()
            for i in range(1, pts.Count + 1):
                try: pts.Item(i).RefreshTable()
                except: pass
        except: pass
        try:
            wb.api.Application.CalculateUntilAsyncQueriesDone()
        except: pass

        log("Guardando...")
        wb.save()
        log("=" * 60)
        log("OK. REPORTE CST FLAT y REPORTE CST PIVOT listos.")

    except Exception as e:
        log(f"ERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
    finally:
        try: wb.close()
        except: pass
        try: app.quit()
        except: pass


if __name__ == '__main__':
    main()
