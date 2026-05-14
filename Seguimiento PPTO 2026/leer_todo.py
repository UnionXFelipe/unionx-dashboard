"""Lee todos los valores necesarios del archivo nuevo y los guarda en un JSON."""
import json
from openpyxl import load_workbook

FILE = r'C:\Users\felip\Desktop\UNIONX\PPTO 2026\Metas oficiales 1SEM Nuevo.xlsx'
wb = load_workbook(FILE, data_only=True)
print("Hojas:", wb.sheetnames)

# ── Hoja 1: PPTO 2026 REAL PP+M ──────────────────────────────────────────────
ws1 = wb['PPTO 2026 REAL PP+M']

# Columnas 2026 VNT (1-based): S=19,U=21,W=23,Y=25,AA=27,AC=29,AE=31,AG=33,AI=35,AK=37,AM=39,AO=41
MONTH_COLS_S1 = [19,21,23,25,27,29,31,33,35,37,39,41]

# Buscar canales dinámicamente
canal_vnt = {}
canal_rows = {}
print("\n=== PPTO 2026 REAL PP+M: filas con Venta Neta Total ===")
for i, row in enumerate(ws1.iter_rows(min_row=55, max_row=170, values_only=True), start=55):
    if row[0] and row[1] == 'Venta Neta Total':
        canal = str(row[0])
        vals = []
        for col in MONTH_COLS_S1:
            v = ws1.cell(i, col).value
            vals.append(float(v) if v is not None else 0.0)
        canal_vnt[canal] = vals
        canal_rows[canal] = i
        print(f"  Row {i}: {canal} -> Ene={vals[0]:,.0f}, Feb={vals[1]:,.0f}")

# ── Hoja 2: PPTO MARCA 2026 ───────────────────────────────────────────────────
ws2 = wb['PPTO MARCA 2026']

# Venta neta por canal-marca (cols D-O = 4-15, filas 69-117)
# Contrib frontal por canal-marca (cols BQ-CB = 69-80, filas 69-117)
brand_canal_venta   = []   # [{canal, marca, meses:[12]}]
brand_canal_contrib = []

print("\n=== PPTO MARCA 2026: filas 67-120 ===")
for i in range(67, 120):
    canal_v = ws2.cell(i, 2).value
    marca_v = ws2.cell(i, 3).value
    if canal_v is None:
        continue
    if 'Total' in str(canal_v) or 'TOTAL' in str(canal_v) or 'SUPUESTO' in str(canal_v):
        continue
    if marca_v is None:
        continue
    # Venta: cols 4-15 (D-O)
    venta_vals = [float(ws2.cell(i, c).value or 0) for c in range(4, 16)]
    # Contrib: cols 69-80 (BQ-CB)
    contrib_vals = [float(ws2.cell(i, c).value or 0) for c in range(69, 81)]
    brand_canal_venta.append({'canal': str(canal_v), 'marca': str(marca_v), 'meses': venta_vals})
    brand_canal_contrib.append({'canal': str(canal_v), 'marca': str(marca_v), 'meses': contrib_vals})
    print(f"  Row {i}: {canal_v} | {marca_v} | VentaEne={venta_vals[0]:,.0f} | ContribEne={contrib_vals[0]:,.0f}")

# ── Hoja MATTEL: meta por canal (Distribución y Marketplace) ─────────────────
# Columnas: Ene=col5, Feb=col7, Mar=col9 ... (col = 5 + mes_idx*2)
# Fila 2 = Dist venta, Fila 3 = Mkt venta, Fila 4 = Contrib total
ws_m = wb['MATTEL']
mattel_dist_v  = [float(ws_m.cell(2, 5 + i*2).value or 0) for i in range(12)]
mattel_mkt_v   = [float(ws_m.cell(3, 5 + i*2).value or 0) for i in range(12)]
mattel_total_c = [float(ws_m.cell(4, 5 + i*2).value or 0) for i in range(12)]
print(f"\nMATTEL dist venta (Mar-Sep): {[round(v) for v in mattel_dist_v[2:9]]}")
print(f"MATTEL mkt  venta (Mar-Sep): {[round(v) for v in mattel_mkt_v[2:9]]}")
print(f"MATTEL contrib total (Mar-Sep): {[round(v) for v in mattel_total_c[2:9]]}")

wb.close()

# ── Guardar en JSON ───────────────────────────────────────────────────────────
data = {
    'canal_vnt': canal_vnt,
    'canal_rows': canal_rows,
    'brand_canal_venta': brand_canal_venta,
    'brand_canal_contrib': brand_canal_contrib,
    'mattel_dist_v': mattel_dist_v,
    'mattel_mkt_v': mattel_mkt_v,
    'mattel_total_c': mattel_total_c,
}
with open(r'C:\Users\felip\Desktop\UnionX Cloude\Seguimiento PPTO 2026\datos_leidos.json', 'w', encoding='utf-8') as f:
    json.dump(data, f, ensure_ascii=False, indent=2)

print("\nDatos guardados en datos_leidos.json")
print(f"Canales encontrados: {list(canal_vnt.keys())}")
print(f"Filas brand-canal venta: {len(brand_canal_venta)}")
print(f"Filas brand-canal contrib: {len(brand_canal_contrib)}")
