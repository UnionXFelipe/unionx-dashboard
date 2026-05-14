"""
Crea las 4 hojas de seguimiento en Metas oficiales 1SEM Nuevo.xlsx
con valores 100% hardcodeados. Sin dependencias de formulas externas.
Real data leida desde FORECAST FINAL SKU 26-27 V2.xlsx, hoja TD VENTAS Net + Contrib.
La TD se refresca automáticamente como parte del proceso del lunes antes de guardar
el archivo, por lo que los valores cacheados siempre están actualizados.
"""
import json
import datetime
import calendar
import pandas as pd
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FILE  = r'C:\Users\felip\Desktop\UNIONX\PPTO 2026\Metas oficiales 1SEM Nuevo.xlsx'
JSON  = r'C:\Users\felip\Desktop\UnionX Cloude\Seguimiento PPTO 2026\datos_leidos.json'
REAL  = r'C:\Users\felip\Desktop\UNIONX\FORECAST FINAL SKU\FORECAST FINAL SKU 26-27 V2.xlsx'

# =============================================================================
# 1. CARGAR METAS (desde JSON del presupuesto)
# =============================================================================
with open(JSON, encoding='utf-8') as f:
    _json = json.load(f)

canal_vnt          = _json['canal_vnt']
brand_canal_venta  = _json['brand_canal_venta']
brand_canal_contrib= _json['brand_canal_contrib']
mattel_dist_v      = _json['mattel_dist_v']
mattel_mkt_v       = _json['mattel_mkt_v']
mattel_total_c     = _json['mattel_total_c']

def is_header(row):
    return row['canal'] in ('Tipo canal',) or row['marca'] in ('Marca','SUPUESTO')

brand_canal_venta   = [r for r in brand_canal_venta   if not is_header(r)]
brand_canal_contrib = [r for r in brand_canal_contrib if not is_header(r)]

MARCA_MAP = {'Bloom': 'Mattel', 'MMPP': 'UnionX B2B'}
for row in brand_canal_venta + brand_canal_contrib:
    row['marca'] = MARCA_MAP.get(row['marca'], row['marca'])

def sum_by_marca(rows):
    t = defaultdict(lambda: [0.0]*12)
    for r in rows:
        for i, v in enumerate(r['meses']):
            t[r['marca']][i] += v
    return t

def sum_by_canal(rows):
    t = defaultdict(lambda: [0.0]*12)
    for r in rows:
        for i, v in enumerate(r['meses']):
            t[r['canal']][i] += v
    return t

brand_meta_venta   = sum_by_marca(brand_canal_venta)
brand_meta_contrib = sum_by_marca(brand_canal_contrib)

# Unificar "Otras Marcas" dentro de "Prov. Nacionales"
for bm in (brand_meta_venta, brand_meta_contrib):
    if 'Otras Marcas' in bm:
        om = bm.pop('Otras Marcas')
        pn = bm.get('Prov. Nacionales', [0.0]*12)
        bm['Prov. Nacionales'] = [pn[i]+om[i] for i in range(12)]

canal_meta_contrib = sum_by_canal(brand_canal_contrib)
canal_meta_venta   = canal_vnt

# Fusionar Purito en Marketplace (meta venta y contrib)
for _dct in (canal_meta_venta, canal_meta_contrib):
    if 'Purito' in _dct:
        mkt = _dct.get('Marketplace', [0.0]*12)
        pur = _dct.pop('Purito')
        _dct['Marketplace'] = [mkt[i]+pur[i] for i in range(12)]

# Contrib proporcional a la venta de cada canal
mattel_dist_c, mattel_mkt_c = [], []
for i in range(12):
    dv, mv, tc = mattel_dist_v[i], mattel_mkt_v[i], mattel_total_c[i]
    tot = dv + mv
    if tot > 0:
        mattel_dist_c.append(tc * dv / tot)
        mattel_mkt_c.append(tc * mv / tot)
    else:
        mattel_dist_c.append(0.0)
        mattel_mkt_c.append(0.0)

# Sumar Mattel a Distribución y Marketplace (venta)
for key, extra in [('Distribución', mattel_dist_v), ('Marketplace', mattel_mkt_v)]:
    base = list(canal_meta_venta.get(key, [0.0]*12))
    canal_meta_venta[key] = [base[i] + extra[i] for i in range(12)]

# Sumar Mattel a Distribución y Marketplace (contrib)
for key, extra in [('Distribución', mattel_dist_c), ('Marketplace', mattel_mkt_c)]:
    base = list(canal_meta_contrib.get(key, [0.0]*12))
    canal_meta_contrib[key] = [base[i] + extra[i] for i in range(12)]

print('Mattel meta canal leido desde MATTEL sheet')
print(f'  Dist venta (Mar-Abr): {mattel_dist_v[2]:.0f} / {mattel_dist_v[3]:.0f}')
print(f'  Mkt  venta (Mar-Abr): {mattel_mkt_v[2]:.0f} / {mattel_mkt_v[3]:.0f}')

# Reemplazar meta de Mattel en MARCAS con totales de hoja MATTEL (Dist + Mkt)
brand_meta_venta['Mattel']  = [mattel_dist_v[i] + mattel_mkt_v[i] for i in range(12)]
brand_meta_contrib['Mattel'] = list(mattel_total_c)
print(f'  Mattel venta marca (Mar-Abr): {brand_meta_venta["Mattel"][2]:.0f} / {brand_meta_venta["Mattel"][3]:.0f}')
print(f'  Mattel contrib marca (Mar-Abr): {brand_meta_contrib["Mattel"][2]:.0f} / {brand_meta_contrib["Mattel"][3]:.0f}')

# =============================================================================
# 2. CARGAR REAL (desde FORECAST FINAL SKU 26-27 V2.xlsx - hoja TD VENTAS Net + Contrib)
# =============================================================================
# Estructura de la hoja "TD VENTAS Net + Contrib":
#   Cols (0-based): 0=TipoMarca/TipoNeg, 1=Marca/Canal, 2=Ene, 3=Feb, 4=Mar, 5=Abr
#   Tablas:
#     Venta marca:  rows 5-17  (0-based index 4-16 con header=None)
#     Contrib marca: rows 25-37 (0-based index 24-36)
#     Venta canal:  row 44=Corporativo, 61=UnionX B2B, 63=TotalDist, 64=Fidel, 65=Mkt, 66=PWeb
#     Contrib canal: row 73=Corporativo, 90=UnionX B2B, 92=TotalDist, 93=Fidel, 94=Mkt, 95=PWeb

df = pd.read_excel(REAL, sheet_name='TD VENTAS Net + Contrib', header=None)

NR = 5   # meses reales (ene-may, mayo parcial)
N  = 12  # meses totales presupuesto

# Linealidad Mayo: día anterior a hoy
_hoy        = datetime.date.today()
DIA_AYER    = _hoy.day - 1                              # 10
MES_LIN     = _hoy.month                                # 5
AÑO_LIN     = _hoy.year                                 # 2026
TOTAL_DIAS  = calendar.monthrange(AÑO_LIN, MES_LIN)[1] # 31
LINEALIDAD  = DIA_AYER / TOTAL_DIAS                     # 10/31

def pad(lst, length=N):
    return list(lst) + [0.0]*(length - len(lst))

def vals_by_label(col0=None, col1=None):
    """Busca fila por etiqueta en col A y/o col B, devuelve ene-may (cols C-G)."""
    for i, row in df.iterrows():
        v0 = str(row[0]).strip() if pd.notna(row[0]) else ''
        v1 = str(row[1]).strip() if pd.notna(row[1]) else ''
        match0 = (col0 is None) or (col0.lower() in v0.lower())
        match1 = (col1 is None) or (col1.lower() in v1.lower())
        if match0 and match1:
            return df.iloc[i, 2:7].fillna(0).tolist()
    print(f'  ADVERTENCIA: no encontrado col0={col0} col1={col1}')
    return [0.0]*NR

# ---- Venta Neta real por Marca (ene-abr) ------------------------------------
# Busqueda por nombre de marca en columna B
otras_v = vals_by_label(col0='Otras marcas')       # tipo marca agregado
lhotse_v= vals_by_label(col1='Lhotse')
simpl_v = vals_by_label(col1='Simplit')
levo_v  = vals_by_label(col1='LEVO')
xroad_v = vals_by_label(col1='Xroad')
dtl_v   = vals_by_label(col1='Dynamo TL')
bandu_v = vals_by_label(col1='Band')               # Bandu/Bandú
tcare_v = vals_by_label(col1='T-Care')
uma_v   = vals_by_label(col1='UMA')               # Mattel
# ITEK excluido de marcas (va en Corporativo canal)
goya_v  = vals_by_label(col1='Goya')
# Klip Klap excluido

real_venta = {
    'LEVO':             pad(levo_v),
    'Lhotse':           pad(lhotse_v),
    'Simplit':          pad(simpl_v),
    'XROAD':            pad(xroad_v),
    'Marca Flash':      pad([dtl_v[i]+bandu_v[i]+tcare_v[i] for i in range(NR)]),
    # Prov. Nacionales = Otras marcas + Goya  (ITEK excluido)
    'Prov. Nacionales': pad([otras_v[i]+goya_v[i] for i in range(NR)]),
    'Mattel':           pad(uma_v),
    'Purito':           pad([0]*NR),
    'UnionX B2B':       pad(vals_by_label(col0='Distribuci', col1='UnionX B2B')),
}

# ---- Contribucion Frontal real por Marca (ene-abr) --------------------------
# Segunda tabla de marcas: buscar desde la fila de "Suma de Margen Front"
# Para distinguirla de la tabla de venta, buscamos la segunda ocurrencia de cada marca
def vals_by_label_nth(col0=None, col1=None, n=1):
    """Devuelve la n-esima ocurrencia (1=primera, 2=segunda)."""
    count = 0
    for i, row in df.iterrows():
        v0 = str(row[0]).strip() if pd.notna(row[0]) else ''
        v1 = str(row[1]).strip() if pd.notna(row[1]) else ''
        match0 = (col0 is None) or (col0.lower() in v0.lower())
        match1 = (col1 is None) or (col1.lower() in v1.lower())
        if match0 and match1:
            count += 1
            if count == n:
                return df.iloc[i, 2:7].fillna(0).tolist()
    print(f'  ADVERTENCIA: no encontrado (n={n}) col0={col0} col1={col1}')
    return [0.0]*NR

otras_c = vals_by_label_nth(col0='Otras marcas', n=2)
lhotse_c= vals_by_label_nth(col1='Lhotse',    n=2)
simpl_c = vals_by_label_nth(col1='Simplit',    n=2)
levo_c  = vals_by_label_nth(col1='LEVO',       n=2)
xroad_c = vals_by_label_nth(col1='Xroad',      n=2)
dtl_c   = vals_by_label_nth(col1='Dynamo TL',  n=2)
bandu_c = vals_by_label_nth(col1='Band',        n=2)
tcare_c = vals_by_label_nth(col1='T-Care',      n=2)
uma_c   = vals_by_label_nth(col1='UMA',         n=2)
goya_c  = vals_by_label_nth(col1='Goya',        n=2)

real_contrib = {
    'LEVO':             pad(levo_c),
    'Lhotse':           pad(lhotse_c),
    'Simplit':          pad(simpl_c),
    'XROAD':            pad(xroad_c),
    'Marca Flash':      pad([dtl_c[i]+bandu_c[i]+tcare_c[i] for i in range(NR)]),
    'Prov. Nacionales': pad([otras_c[i]+goya_c[i] for i in range(NR)]),
    'Mattel':           pad(uma_c),
    'Purito':           pad([0]*NR),
    'UnionX B2B':       pad(vals_by_label_nth(col0='Distribuci', col1='UnionX B2B', n=2)),
}

# ---- Venta Neta real por Canal (ene-abr) ------------------------------------
# Busqueda por etiqueta en tabla de Venta Neta Canal
# "Total Distribucion" aparece 2 veces (venta y contrib); usamos la primera
corp_v  = vals_by_label_nth(col0='Corporativo', n=1)
totd_v  = vals_by_label_nth(col0='Total Distribu', n=1)
uxb_v   = vals_by_label_nth(col0='Distribuci', col1='UnionX B2B', n=1)
fidel_v = vals_by_label_nth(col0='Fidelizaci', n=1)
mkt_v   = vals_by_label_nth(col0='Marketplace', n=1)
pweb_v  = vals_by_label_nth(col0='ginas propias', n=1)

real_canal_venta = {
    'Corporativo':     pad(corp_v),
    'Distribución':    pad([totd_v[i]-uxb_v[i] for i in range(NR)]),
    'UnionX B2B':      pad(uxb_v),
    'Fidelización':    pad(fidel_v),
    'Marketplace':     pad(mkt_v),
    'P.Web':           pad(pweb_v),
    'Tiendas Propias': pad([0]*NR),
}

# ---- Contribucion Frontal real por Canal (ene-abr) --------------------------
corp_c2  = vals_by_label_nth(col0='Corporativo', n=2)
totd_c   = vals_by_label_nth(col0='Total Distribu', n=2)
uxb_c    = vals_by_label_nth(col0='Distribuci', col1='UnionX B2B', n=2)
fidel_c2 = vals_by_label_nth(col0='Fidelizaci', n=2)
mkt_c2   = vals_by_label_nth(col0='Marketplace', n=2)
pweb_c2  = vals_by_label_nth(col0='ginas propias', n=2)

real_canal_contrib = {
    'Corporativo':     pad(corp_c2),
    'Distribución':    pad([totd_c[i]-uxb_c[i] for i in range(NR)]),
    'UnionX B2B':      pad(uxb_c),
    'Fidelización':    pad(fidel_c2),
    'Marketplace':     pad(mkt_c2),
    'P.Web':           pad(pweb_c2),
    'Tiendas Propias': pad([0]*NR),
}

# Mattel adicional = UMA
real_canal_venta_adic_v  = uma_v
real_canal_contrib_adic_c = uma_c

print('Real data leida desde:', REAL)
print(f'  Corporativo venta:    {corp_v}')
print(f'  Total Dist venta:     {totd_v}')
print(f'  UnionX B2B venta:     {uxb_v}')
print(f'  Fidelizacion venta:   {fidel_v}')
print(f'  Marketplace venta:    {mkt_v}')
print(f'  P.Web venta:          {pweb_v}')
print(f'  Corporativo contrib:  {corp_c2}')
print(f'  Total Dist contrib:   {totd_c}')
print(f'  UnionX B2B contrib:   {uxb_c}')

# =============================================================================
# 3. ESTILOS
# =============================================================================
thin = Side(style='thin', color='BFBFBF')
BRD  = Border(left=thin, right=thin, top=thin, bottom=thin)
MESES = ['Enero','Febrero','Marzo','Abril','Mayo','Junio',
         'Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre']
MESES4 = ['Ene','Feb','Mar','Abr','May']

FILLS = {
    'navy':'1F4E79','blue':'2E75B6','green':'375623',
    'lbl_blue':'D6E4F0','lbl_contrib':'E2EFDA',
    'tot_blue':'BDD7EE','tot_green':'C6E0B4',
    'meta':'D6E4F0','real':'E2EFDA','var':'FFF2CC',
    'tmeta':'BDD7EE','treal':'C6E0B4','tvar':'FFE699',
    'note':'FFF8E1',
}

def cs(cell, val=None, bold=False, fill=None, fmt='#,##0',
       align='right', color='000000', size=10):
    if val is not None:
        cell.value = val
    cell.font = Font(name='Arial', bold=bold, size=size, color=color)
    if fill:
        cell.fill = PatternFill('solid', start_color=fill)
    if fmt is not None:
        cell.number_format = fmt
    cell.alignment = Alignment(horizontal=align, vertical='center')
    cell.border = BRD

def _sum_excl(col_letter, data_start, data_end, excl_rows):
    """Fórmula SUM que salta las filas indicadas (para excluir real de UXB del total)."""
    excl = sorted(excl_rows)
    if not excl:
        return f'=SUM({col_letter}{data_start}:{col_letter}{data_end})'
    parts = []
    prev = data_start
    for er in excl:
        if prev <= er - 1:
            parts.append(f'SUM({col_letter}{prev}:{col_letter}{er-1})')
        prev = er + 1
    if prev <= data_end:
        parts.append(f'SUM({col_letter}{prev}:{col_letter}{data_end})')
    return ('=' + '+'.join(parts)) if parts else '=0'

def banner(ws, row, text, fill, n_cols, size=13):
    ws.merge_cells(f'A{row}:{get_column_letter(n_cols)}{row}')
    c = ws.cell(row, 1)
    c.value = text
    c.font = Font(name='Arial', bold=True, size=size, color='FFFFFF')
    c.fill = PatternFill('solid', start_color=fill)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 26

def note_row(ws, row, text, n_cols, fill=None, color='595959'):
    ws.merge_cells(f'A{row}:{get_column_letter(n_cols)}{row}')
    c = ws.cell(row, 1)
    c.value = text
    c.font = Font(name='Arial', italic=True, size=8, color=color)
    if fill:
        c.fill = PatternFill('solid', start_color=fill)
    c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws.row_dimensions[row].height = 14

def mhdr(ws, r, c1, c2, text, fill):
    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    c = ws.cell(r, c1)
    c.value = text
    c.font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
    c.fill = PatternFill('solid', start_color=fill)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = BRD
    for col in range(c1+1, c2+1):
        ws.cell(r, col).border = BRD

def col_widths(ws, col_a=22, col_rest=14, n_cols=14):
    ws.column_dimensions['A'].width = col_a
    for c in range(2, n_cols+1):
        ws.column_dimensions[get_column_letter(c)].width = col_rest

# Dimensiones comparativo (5 meses: ene-may)
N4 = 5
TC = 1 + N4*3 + 3   # 19 columnas total

# =============================================================================
# HOJA 1: Seguimiento Ventas Macro
# =============================================================================
wb = load_workbook(FILE)

CANAL_ORDER = ['Corporativo','Distribución','Fidelización','Marketplace',
               'P.Web','Tiendas Propias','UnionX B2B']

if 'Seguimiento Ventas Macro' in wb.sheetnames:
    del wb['Seguimiento Ventas Macro']
sv = wb.create_sheet('Seguimiento Ventas Macro')

banner(sv, 1, 'SEGUIMIENTO VENTAS MACRO 2026  --  META VENTA NETA TOTAL', FILLS['navy'], 14)
note_row(sv, 2, 'Fuente: PPTO 2026 REAL PP+M. Valores en pesos chilenos.', 14)

sv.row_dimensions[3].height = 28
cs(sv.cell(3,1), 'CANAL', bold=True, fill=FILLS['blue'], color='FFFFFF', align='center', fmt=None)
for m, mes in enumerate(MESES, start=2):
    cs(sv.cell(3,m), mes, bold=True, fill=FILLS['blue'], color='FFFFFF', align='center', fmt=None)
cs(sv.cell(3,14), 'TOTAL 2026', bold=True, fill=FILLS['navy'], color='FFFFFF', align='center', fmt=None)

D = 4
for r_off, canal in enumerate(CANAL_ORDER):
    row = D + r_off
    sv.row_dimensions[row].height = 18
    cs(sv.cell(row,1), canal, fill=FILLS['lbl_blue'], align='left', fmt=None)
    vals = canal_vnt.get(canal, [0]*12)
    for col, v in enumerate(vals, start=2):
        cs(sv.cell(row,col), v)
    cs(sv.cell(row,14), f'=SUM(B{row}:M{row})', bold=True, fill=FILLS['tot_blue'])

tot_row = D + len(CANAL_ORDER)
sv.row_dimensions[tot_row].height = 20
cs(sv.cell(tot_row,1), 'TOTALES', bold=True, fill=FILLS['navy'], color='FFFFFF', align='left', fmt=None)
for col, v in enumerate(canal_vnt.get('TOTALES', [0]*12), start=2):
    cs(sv.cell(tot_row,col), v, bold=True, fill=FILLS['tot_blue'])
cs(sv.cell(tot_row,14), f'=SUM(B{tot_row}:M{tot_row})', bold=True, fill=FILLS['blue'], color='FFFFFF')

sv.freeze_panes = 'B4'
col_widths(sv)
print('OK: Seguimiento Ventas Macro')

# =============================================================================
# HOJA 2: Seguimiento Ventas Marcas
# =============================================================================
BRAND_ORDER = ['LEVO','Lhotse','Simplit','XROAD','Marca Flash',
               'Prov. Nacionales','Mattel','Purito','UnionX B2B']

if 'Seguimiento Ventas Marcas' in wb.sheetnames:
    del wb['Seguimiento Ventas Marcas']
sm = wb.create_sheet('Seguimiento Ventas Marcas')

def build_marca_section(ws, start_r, title, fill_hdr, data_dict, lbl_fill, tot_fill):
    banner(ws, start_r, title, fill_hdr, 14)
    r = start_r + 1
    ws.row_dimensions[r].height = 28
    cs(ws.cell(r,1), 'MARCA', bold=True, fill=fill_hdr, color='FFFFFF', align='center', fmt=None)
    for m, mes in enumerate(MESES, start=2):
        cs(ws.cell(r,m), mes, bold=True, fill=fill_hdr, color='FFFFFF', align='center', fmt=None)
    cs(ws.cell(r,14), 'TOTAL 2026', bold=True, fill=FILLS['navy'], color='FFFFFF', align='center', fmt=None)
    r += 1
    data_start = r
    for brand in BRAND_ORDER:
        ws.row_dimensions[r].height = 17
        cs(ws.cell(r,1), brand, fill=lbl_fill, align='left', fmt=None)
        vals = data_dict.get(brand, [0]*12)
        for col, v in enumerate(vals, start=2):
            cs(ws.cell(r,col), v)
        cs(ws.cell(r,14), f'=SUM(B{r}:M{r})', bold=True, fill=tot_fill)
        r += 1
    ws.row_dimensions[r].height = 20
    cs(ws.cell(r,1), 'TOTAL', bold=True, fill=fill_hdr, color='FFFFFF', align='left', fmt=None)
    for col in range(2, 14):
        cl = get_column_letter(col)
        cs(ws.cell(r,col), f'=SUM({cl}{data_start}:{cl}{r-1})', bold=True, fill=tot_fill)
    cs(ws.cell(r,14), f'=SUM(B{r}:M{r})', bold=True, fill=fill_hdr, color='FFFFFF')
    return r + 2

nr = build_marca_section(sm, 1,
    'VENTA NETA META 2026 POR MARCA  --  SUMA TODOS LOS CANALES',
    FILLS['blue'], brand_meta_venta, FILLS['lbl_blue'], FILLS['tot_blue'])
nr = build_marca_section(sm, nr,
    'CONTRIBUCION FRONTAL META 2026 POR MARCA  --  SUMA TODOS LOS CANALES',
    FILLS['green'], brand_meta_contrib, FILLS['lbl_contrib'], FILLS['tot_green'])

sm.freeze_panes = 'B3'
col_widths(sm)
print('OK: Seguimiento Ventas Marcas')

# =============================================================================
# HOJA 3: Comp. Marcas  (Real vs Meta, ene-abr)
# =============================================================================
if 'Comp. Marcas' in wb.sheetnames:
    del wb['Comp. Marcas']
cm = wb.create_sheet('Comp. Marcas')

def build_comp_section(ws, start_r, section_title, fill_hdr,
                        meta_dict, real_dict, lbl_fill, order,
                        lbl_col, source_note, exclude_real=None):
    banner(ws, start_r, section_title, fill_hdr, TC, size=12)
    note_row(ws, start_r+1, source_note, TC)

    r = start_r + 2
    ws.row_dimensions[r].height = 20
    mhdr(ws, r, 1, 1, lbl_col, fill_hdr)
    col = 2
    for mes in MESES4:
        mhdr(ws, r, col, col+2, mes, fill_hdr)
        col += 3
    mhdr(ws, r, col, col+2, 'TOTAL ENE-ABR', FILLS['navy'])
    r += 1

    ws.row_dimensions[r].height = 18
    cs(ws.cell(r,1), '', fill=fill_hdr, fmt=None)
    col = 2
    for _ in MESES4:
        cs(ws.cell(r,col),   'META', bold=True, fill=FILLS['meta'],  align='center', fmt=None)
        cs(ws.cell(r,col+1), 'REAL', bold=True, fill=FILLS['real'],  align='center', fmt=None)
        cs(ws.cell(r,col+2), 'VAR%', bold=True, fill=FILLS['var'],   align='center', fmt=None)
        col += 3
    cs(ws.cell(r,col),   'META', bold=True, fill=FILLS['tmeta'], align='center', fmt=None)
    cs(ws.cell(r,col+1), 'REAL', bold=True, fill=FILLS['treal'], align='center', fmt=None)
    cs(ws.cell(r,col+2), 'VAR%', bold=True, fill=FILLS['tvar'],  align='center', fmt=None)
    r += 1
    data_start = r
    excl_rows = []   # filas cuyo real NO entra en el TOTAL

    for item in order:
        ws.row_dimensions[r].height = 16
        cs(ws.cell(r,1), item, fill=lbl_fill, align='left', fmt=None)
        meta_vals = meta_dict.get(item, [0]*12)
        real_vals = real_dict.get(item, [0]*12)
        meta_cols, real_cols = [], []
        col = 2
        for m in range(N4):
            mv = meta_vals[m]; rv = real_vals[m]
            cs(ws.cell(r,col),   mv, fill=FILLS['meta'])
            cs(ws.cell(r,col+1), rv, fill=FILLS['real'])
            mc = get_column_letter(col)+str(r)
            rc = get_column_letter(col+1)+str(r)
            cs(ws.cell(r,col+2), f'=IF({mc}<>0,{rc}/{mc}-1,"")', fill=FILLS['var'], fmt='0.0%')
            meta_cols.append(mc); real_cols.append(rc)
            col += 3
        cs(ws.cell(r,col),   '='+'+'.join(meta_cols), bold=True, fill=FILLS['tmeta'])
        cs(ws.cell(r,col+1), '='+'+'.join(real_cols), bold=True, fill=FILLS['treal'])
        tm = get_column_letter(col)+str(r); tr = get_column_letter(col+1)+str(r)
        cs(ws.cell(r,col+2), f'=IF({tm}<>0,{tr}/{tm}-1,"")', bold=True, fill=FILLS['tvar'], fmt='0.0%')
        if exclude_real and item in exclude_real:
            excl_rows.append(r)
        r += 1

    # Fila TOTAL — meta suma todo; real excluye filas en excl_rows
    ws.row_dimensions[r].height = 20
    c0 = ws.cell(r,1); c0.value = 'TOTAL'
    c0.font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
    c0.fill = PatternFill('solid', start_color=FILLS['navy'])
    c0.alignment = Alignment(horizontal='left', vertical='center'); c0.border = BRD
    col = 2
    for _ in range(N4):
        mc = get_column_letter(col); rc = get_column_letter(col+1)
        cs(ws.cell(r,col),   f'=SUM({mc}{data_start}:{mc}{r-1})', bold=True, fill=FILLS['tmeta'])
        cs(ws.cell(r,col+1), _sum_excl(rc, data_start, r-1, excl_rows), bold=True, fill=FILLS['treal'])
        mref = get_column_letter(col)+str(r); rref = get_column_letter(col+1)+str(r)
        cs(ws.cell(r,col+2), f'=IF({mref}<>0,{rref}/{mref}-1,"")', bold=True, fill=FILLS['tvar'], fmt='0.0%')
        col += 3
    tc2 = get_column_letter(col); rc2 = get_column_letter(col+1)
    cs(ws.cell(r,col),   f'=SUM({tc2}{data_start}:{tc2}{r-1})', bold=True, fill=FILLS['blue'],  color='FFFFFF')
    cs(ws.cell(r,col+1), _sum_excl(rc2, data_start, r-1, excl_rows), bold=True, fill=FILLS['green'], color='FFFFFF')
    mref = get_column_letter(col)+str(r); rref = get_column_letter(col+1)+str(r)
    cs(ws.cell(r,col+2), f'=IF({mref}<>0,{rref}/{mref}-1,"")', bold=True, fill=FILLS['tvar'], fmt='0.0%')
    return r + 2

banner(cm, 1, 'COMPARATIVO REAL VS META 2026  --  ANALISIS POR MARCA', FILLS['navy'], TC, size=13)
note_row(cm, 2,
    f'Ene-May 2026 (Mayo parcial al día {DIA_AYER} de {TOTAL_DIAS} — linealidad {LINEALIDAD:.1%}).  '
    'Marca Flash=DynamoTL+Bandu+T-Care  |  Mattel=UMA  |  '
    'Prov.Nacionales=Otras Marcas+Goya  |  ITEK en Corporativo canal  |  Klip Klap excluido.',
    TC)
cm.row_dimensions[2].height = 16

nr2 = build_comp_section(cm, 3,
    'VENTA NETA POR MARCA', FILLS['blue'],
    brand_meta_venta, real_venta, FILLS['lbl_blue'], BRAND_ORDER, 'MARCA',
    'Meta: PPTO MARCA 2026 + hoja MATTEL.  Real: Ventas Netas - TD VENTAS Net + Contrib (FORECAST FINAL SKU).  '
    'UnionX B2B: meta incluida en total, real solo informativo (canal separado).',
    exclude_real=['UnionX B2B'])

nr2 = build_comp_section(cm, nr2,
    'CONTRIBUCION FRONTAL POR MARCA', FILLS['green'],
    brand_meta_contrib, real_contrib, FILLS['lbl_contrib'], BRAND_ORDER, 'MARCA',
    'Meta: PPTO MARCA 2026 + hoja MATTEL.  Real: Margen Front - TD VENTAS Net + Contrib (FORECAST FINAL SKU).  '
    'UnionX B2B: meta incluida en total, real solo informativo (canal separado).',
    exclude_real=['UnionX B2B'])

cm.freeze_panes = 'B5'
cm.column_dimensions['A'].width = 20
for c in range(2, TC+1):
    cm.column_dimensions[get_column_letter(c)].width = 7.5 if (c-2)%3==2 else 13
print('OK: Comp. Marcas')

# =============================================================================
# HOJA 4: Comp. Canales  (Venta Neta + Contribucion, Real vs Meta, ene-abr)
# Mattel sumado a Distribución y Marketplace (meta). Real ya incluye UMA en canales.
# =============================================================================
CANAL_COMP_MAIN = ['Corporativo','Distribución','UnionX B2B','Fidelización',
                   'Marketplace','P.Web','Tiendas Propias']

def build_comp_section_with_adic(ws, start_r, section_title, fill_hdr,
                                  meta_dict, real_dict, lbl_fill,
                                  order_main, order_adic,
                                  meta_adic, real_adic,
                                  lbl_col, source_note):
    """Igual que build_comp_section pero con filas adicionales fuera del total."""
    banner(ws, start_r, section_title, fill_hdr, TC, size=12)
    note_row(ws, start_r+1, source_note, TC)

    r = start_r + 2
    ws.row_dimensions[r].height = 20
    mhdr(ws, r, 1, 1, lbl_col, fill_hdr)
    col = 2
    for mes in MESES4:
        mhdr(ws, r, col, col+2, mes, fill_hdr)
        col += 3
    mhdr(ws, r, col, col+2, 'TOTAL ENE-ABR', FILLS['navy'])
    r += 1

    ws.row_dimensions[r].height = 18
    cs(ws.cell(r,1), '', fill=fill_hdr, fmt=None)
    col = 2
    for _ in MESES4:
        cs(ws.cell(r,col),   'META', bold=True, fill=FILLS['meta'],  align='center', fmt=None)
        cs(ws.cell(r,col+1), 'REAL', bold=True, fill=FILLS['real'],  align='center', fmt=None)
        cs(ws.cell(r,col+2), 'VAR%', bold=True, fill=FILLS['var'],   align='center', fmt=None)
        col += 3
    cs(ws.cell(r,col),   'META', bold=True, fill=FILLS['tmeta'], align='center', fmt=None)
    cs(ws.cell(r,col+1), 'REAL', bold=True, fill=FILLS['treal'], align='center', fmt=None)
    cs(ws.cell(r,col+2), 'VAR%', bold=True, fill=FILLS['tvar'],  align='center', fmt=None)
    r += 1
    data_start = r

    def write_data_row(ws, r, item, meta_vals, real_vals, lbl_fill, bold=False):
        cs(ws.cell(r,1), item, fill=lbl_fill, align='left', fmt=None, bold=bold)
        meta_cols, real_cols = [], []
        col = 2
        for m in range(N4):
            mv = meta_vals[m]; rv = real_vals[m]
            cs(ws.cell(r,col),   mv, fill=FILLS['meta'], bold=bold)
            cs(ws.cell(r,col+1), rv, fill=FILLS['real'], bold=bold)
            mc = get_column_letter(col)+str(r)
            rc = get_column_letter(col+1)+str(r)
            cs(ws.cell(r,col+2), f'=IF({mc}<>0,{rc}/{mc}-1,"")', fill=FILLS['var'], fmt='0.0%')
            meta_cols.append(mc); real_cols.append(rc)
            col += 3
        cs(ws.cell(r,col),   '='+'+'.join(meta_cols), bold=True, fill=FILLS['tmeta'])
        cs(ws.cell(r,col+1), '='+'+'.join(real_cols), bold=True, fill=FILLS['treal'])
        tm = get_column_letter(col)+str(r); tr = get_column_letter(col+1)+str(r)
        cs(ws.cell(r,col+2), f'=IF({tm}<>0,{tr}/{tm}-1,"")', bold=True, fill=FILLS['tvar'], fmt='0.0%')

    # Canales principales
    for item in order_main:
        ws.row_dimensions[r].height = 16
        write_data_row(ws, r, item,
                       meta_dict.get(item, [0]*12),
                       real_dict.get(item, [0]*12), lbl_fill)
        r += 1

    # Fila TOTAL (solo canales principales)
    ws.row_dimensions[r].height = 20
    c0 = ws.cell(r,1); c0.value = 'TOTAL'
    c0.font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
    c0.fill = PatternFill('solid', start_color=FILLS['navy'])
    c0.alignment = Alignment(horizontal='left', vertical='center'); c0.border = BRD
    col = 2
    for _ in range(N4):
        mc = get_column_letter(col); rc = get_column_letter(col+1)
        cs(ws.cell(r,col),   f'=SUM({mc}{data_start}:{mc}{r-1})', bold=True, fill=FILLS['tmeta'])
        cs(ws.cell(r,col+1), f'=SUM({rc}{data_start}:{rc}{r-1})', bold=True, fill=FILLS['treal'])
        mref = get_column_letter(col)+str(r); rref = get_column_letter(col+1)+str(r)
        cs(ws.cell(r,col+2), f'=IF({mref}<>0,{rref}/{mref}-1,"")', bold=True, fill=FILLS['tvar'], fmt='0.0%')
        col += 3
    tc2 = get_column_letter(col); rc2 = get_column_letter(col+1)
    cs(ws.cell(r,col),   f'=SUM({tc2}{data_start}:{tc2}{r-1})', bold=True, fill=FILLS['blue'],  color='FFFFFF')
    cs(ws.cell(r,col+1), f'=SUM({rc2}{data_start}:{rc2}{r-1})', bold=True, fill=FILLS['green'], color='FFFFFF')
    mref = get_column_letter(col)+str(r); rref = get_column_letter(col+1)+str(r)
    cs(ws.cell(r,col+2), f'=IF({mref}<>0,{rref}/{mref}-1,"")', bold=True, fill=FILLS['tvar'], fmt='0.0%')
    r += 1

    # Separador ADICIONALES
    ws.merge_cells(f'A{r}:{get_column_letter(TC)}{r}')
    sep = ws.cell(r,1); sep.value = 'ADICIONALES (informativo, no incluidos en total)'
    sep.font = Font(name='Arial', bold=True, italic=True, size=9, color='FFFFFF')
    sep.fill = PatternFill('solid', start_color='7F7F7F')
    sep.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[r].height = 16
    r += 1

    # Filas adicionales (Mattel, Purito)
    for item in order_adic:
        ws.row_dimensions[r].height = 16
        write_data_row(ws, r, item,
                       meta_adic.get(item, [0]*12),
                       real_adic.get(item, [0]*12), 'F2F2F2')
        r += 1

    return r + 1

if 'Comp. Canales' in wb.sheetnames:
    del wb['Comp. Canales']
cc = wb.create_sheet('Comp. Canales')

banner(cc, 1,
    'COMPARATIVO REAL VS META 2026  --  VENTA NETA Y CONTRIBUCION POR CANAL',
    FILLS['navy'], TC, size=13)
note_row(cc, 2,
    f'Ene-May 2026 (Mayo parcial al día {DIA_AYER} de {TOTAL_DIAS} — linealidad {LINEALIDAD:.1%}).  '
    'Paginas Propias=P.Web  |  Distribucion REAL = Total Dist - UnionX B2B.  '
    'Meta Dist y Marketplace incluyen meta Mattel (hoja MATTEL).',
    TC)
note_row(cc, 3,
    'Meta venta canal: PPTO 2026 REAL PP+M + Mattel.  Meta contrib canal: PPTO MARCA 2026 + Mattel proporcional.  '
    'ITEK incluido en Corporativo real.',
    TC, fill=FILLS['note'], color='7F4F00')
cc.row_dimensions[3].height = 22

nr3 = build_comp_section(cc, 4,
    'VENTA NETA POR CANAL', FILLS['blue'],
    canal_meta_venta, real_canal_venta, FILLS['lbl_blue'], CANAL_COMP_MAIN, 'CANAL',
    'Meta: PPTO 2026 REAL PP+M + hoja MATTEL (VNT por canal).  Real: TD VENTAS Net + Contrib (FORECAST FINAL SKU).')

nr3 = build_comp_section(cc, nr3,
    'CONTRIBUCION FRONTAL POR CANAL', FILLS['green'],
    canal_meta_contrib, real_canal_contrib, FILLS['lbl_contrib'], CANAL_COMP_MAIN, 'CANAL',
    'Meta: PPTO MARCA 2026 + Mattel proporcional.  Real: TD VENTAS Net + Contrib (FORECAST FINAL SKU).')

cc.freeze_panes = 'B6'
cc.column_dimensions['A'].width = 22
for c in range(2, TC+1):
    cc.column_dimensions[get_column_letter(c)].width = 7.5 if (c-2)%3==2 else 13
print('OK: Comp. Canales')

# =============================================================================
# HOJA 5: Cómo Vamos Mayo  (linealidad al día anterior a hoy)
# =============================================================================
MES_IDX = 4   # Mayo = índice 4 (0-based)
N_CV    = 7   # columnas de la hoja

# Colores % vs Lineal (mismo criterio que analisis planificación)
def color_pct_lineal(pct):
    if pct is None:    return 'BFBFBF', '000000'  # gris, sin dato
    if pct >= 1.10:    return '7030A0', 'FFFFFF'   # morado
    if pct >= 0.90:    return '375623', 'FFFFFF'   # verde
    if pct >= 0.70:    return 'FFC000', '000000'   # amarillo
    if pct >= 0.50:    return 'F4B942', '000000'   # naranja
    return 'FF0000', 'FFFFFF'                       # rojo

def build_como_vamos(ws, start_r, section_title, fill_hdr,
                     meta_dict, real_dict, lbl_fill, order, lbl_col,
                     exclude_real=None):
    """Sección Cómo Vamos con linealidad — 7 columnas."""
    banner(ws, start_r, section_title, fill_hdr, N_CV, size=12)
    r = start_r + 1

    # Encabezados
    ws.row_dimensions[r].height = 22
    hdrs = [lbl_col,
            'Meta Mayo',
            f'Meta Lineal\ndía {DIA_AYER}/{TOTAL_DIAS}',
            'Real Acum.\nMayo',
            'vs Lineal ($)',
            '% vs Lineal',
            '% vs Meta']
    hfills = [fill_hdr, FILLS['meta'], FILLS['meta'],
              FILLS['real'], FILLS['var'], FILLS['var'], 'D9D9D9']
    hcolors= ['FFFFFF','000000','000000','000000','000000','000000','000000']
    for c, (h, hf, hc) in enumerate(zip(hdrs, hfills, hcolors), start=1):
        cell = ws.cell(r, c)
        cell.value = h
        cell.font = Font(name='Arial', bold=True, size=9,
                         color='FFFFFF' if c == 1 else hc)
        cell.fill = PatternFill('solid', start_color=hf)
        cell.alignment = Alignment(horizontal='center', vertical='center',
                                   wrap_text=True)
        cell.border = BRD
    r += 1
    data_start = r

    excl_real = set(exclude_real or [])
    tot_meta = 0.0; tot_meta_lin = 0.0; tot_real = 0.0

    for item in order:
        meta_v = meta_dict.get(item, [0]*12)
        real_v = real_dict.get(item, [0]*12)
        meta_m  = meta_v[MES_IDX]
        real_m  = real_v[MES_IDX]
        meta_lin = meta_m * LINEALIDAD
        vs_lin   = real_m - meta_lin
        pct_lin  = (real_m / meta_lin) if meta_lin > 0 else None
        pct_meta = (real_m / meta_m)   if meta_m   > 0 else None

        # Meta: siempre suma. Meta lineal y real: excluye items en exclude_real
        tot_meta += meta_m
        if item not in excl_real:
            tot_meta_lin += meta_lin
            tot_real     += real_m

        ws.row_dimensions[r].height = 16
        cs(ws.cell(r,1), item,     fill=lbl_fill,       align='left', fmt=None)
        cs(ws.cell(r,2), meta_m,   fill=FILLS['meta'])
        cs(ws.cell(r,3), meta_lin, fill=FILLS['meta'])
        cs(ws.cell(r,4), real_m,   fill=FILLS['real'])
        cs(ws.cell(r,5), vs_lin,   fill=FILLS['var'])

        bg, fc = color_pct_lineal(pct_lin)
        c6 = ws.cell(r, 6)
        c6.value  = pct_lin if pct_lin is not None else '—'
        c6.font   = Font(name='Arial', bold=True, size=10, color=fc)
        c6.fill   = PatternFill('solid', start_color=bg)
        c6.number_format = '0.0%' if pct_lin is not None else '@'
        c6.alignment = Alignment(horizontal='center', vertical='center')
        c6.border = BRD

        c7 = ws.cell(r, 7)
        c7.value  = pct_meta if pct_meta is not None else '—'
        c7.font   = Font(name='Arial', size=10, color='595959')
        c7.fill   = PatternFill('solid', start_color='F2F2F2')
        c7.number_format = '0.0%' if pct_meta is not None else '@'
        c7.alignment = Alignment(horizontal='center', vertical='center')
        c7.border = BRD
        r += 1

    # Fila TOTAL — real excluye items en exclude_real
    tot_vs_lin  = tot_real - tot_meta_lin
    tot_pct_lin = (tot_real / tot_meta_lin) if tot_meta_lin > 0 else None
    tot_pct_meta= (tot_real / tot_meta)     if tot_meta     > 0 else None

    ws.row_dimensions[r].height = 20
    cs(ws.cell(r,1), 'TOTAL', bold=True, fill=FILLS['navy'],
       color='FFFFFF', align='left', fmt=None)
    cs(ws.cell(r,2), tot_meta,     bold=True, fill=FILLS['tmeta'])
    cs(ws.cell(r,3), tot_meta_lin, bold=True, fill=FILLS['tmeta'])
    cs(ws.cell(r,4), tot_real,     bold=True, fill=FILLS['treal'])
    cs(ws.cell(r,5), tot_vs_lin,   bold=True, fill=FILLS['tvar'])

    bg_t, fc_t = color_pct_lineal(tot_pct_lin)
    ct6 = ws.cell(r, 6)
    ct6.value  = tot_pct_lin if tot_pct_lin is not None else '—'
    ct6.font   = Font(name='Arial', bold=True, size=10, color=fc_t)
    ct6.fill   = PatternFill('solid', start_color=bg_t)
    ct6.number_format = '0.0%' if tot_pct_lin is not None else '@'
    ct6.alignment = Alignment(horizontal='center', vertical='center')
    ct6.border = BRD

    c7t = ws.cell(r, 7)
    c7t.value  = tot_pct_meta if tot_pct_meta is not None else '—'
    c7t.font   = Font(name='Arial', bold=True, size=10, color='595959')
    c7t.fill   = PatternFill('solid', start_color='D9D9D9')
    c7t.number_format = '0.0%' if tot_pct_meta is not None else '@'
    c7t.alignment = Alignment(horizontal='center', vertical='center')
    c7t.border = BRD
    return r + 2

if 'Cómo Vamos Mayo' in wb.sheetnames:
    del wb['Cómo Vamos Mayo']
cv = wb.create_sheet('Cómo Vamos Mayo')

mes_nombre = datetime.date(AÑO_LIN, MES_LIN, 1).strftime('%B %Y').capitalize()
banner(cv, 1,
    f'CÓMO VAMOS MAYO 2026  —  LINEALIDAD AL DÍA {DIA_AYER} DE {TOTAL_DIAS}  ({LINEALIDAD:.1%})',
    FILLS['navy'], N_CV, size=13)
note_row(cv, 2,
    f'Linealidad = día {DIA_AYER} / {TOTAL_DIAS} días de Mayo = {LINEALIDAD:.2%}.  '
    f'Meta Lineal = Meta Mayo × {LINEALIDAD:.2%}.  '
    'Colores % vs Lineal: 🟣≥110%  🟢90-110%  🟡70-90%  🟠50-70%  🔴<50%',
    N_CV)
cv.row_dimensions[2].height = 18

nr_cv = 3
nr_cv = build_como_vamos(cv, nr_cv,
    'VENTA NETA POR MARCA', FILLS['blue'],
    brand_meta_venta, real_venta, FILLS['lbl_blue'], BRAND_ORDER, 'MARCA',
    exclude_real=['UnionX B2B'])
nr_cv = build_como_vamos(cv, nr_cv,
    'CONTRIBUCIÓN POR MARCA', FILLS['green'],
    brand_meta_contrib, real_contrib, FILLS['lbl_contrib'], BRAND_ORDER, 'MARCA',
    exclude_real=['UnionX B2B'])
nr_cv = build_como_vamos(cv, nr_cv,
    'VENTA NETA POR CANAL', FILLS['blue'],
    canal_meta_venta, real_canal_venta, FILLS['lbl_blue'], CANAL_COMP_MAIN, 'CANAL')
nr_cv = build_como_vamos(cv, nr_cv,
    'CONTRIBUCIÓN POR CANAL', FILLS['green'],
    canal_meta_contrib, real_canal_contrib, FILLS['lbl_contrib'], CANAL_COMP_MAIN, 'CANAL')

cv.freeze_panes = 'B4'
cv.column_dimensions['A'].width = 22
for c in range(2, N_CV+1):
    cv.column_dimensions[get_column_letter(c)].width = 15
print('OK: Cómo Vamos Mayo')

# =============================================================================
# GUARDAR
# =============================================================================
wb.save(FILE)
print(f'\nGuardado: {FILE}')
print('Hojas:', wb.sheetnames)
